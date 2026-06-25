#!/usr/bin/env python3
"""
(b) Convertisseur : JSON extrait (schéma Cicada) -> classeur Excel au format du
modèle d'export Cicada "Fiche action" (onglets "Action CS" / "Action hors CS").

Produit UN classeur contenant UNE feuille par fiche (clonée du bon onglet modèle,
styles conservés), prête à être relue par un gestionnaire avant import.

Approche "label-driven" : on localise les cellules à remplir en cherchant les
libellés en colonne A, plutôt que des adresses en dur — résilient si la mise en
page du modèle évolue.

Usage:
    pip install openpyxl
    python json_to_excel.py sample_output/extraction_demo.json -o sample_output/fiches_remplies.xlsx
"""

import argparse
import json
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

TEMPLATE = Path(__file__).with_name("templates") / "modele_export_fiche_action.xlsx"
SHEET_CS = "Action CS"
SHEET_HORS_CS = "Action hors CS"


def _norm(s) -> str:
    """Minuscule sans accents, pour comparer des libellés de façon tolérante."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.lower().strip()


def _join(values) -> str:
    return " ; ".join(str(v) for v in (values or []) if v)


def _label_rows(ws, max_row=45):
    """Map { libellé normalisé colonne A -> numéro de ligne }."""
    rows = {}
    for r in range(1, max_row + 1):
        label = _norm(ws.cell(r, 1).value)
        if label:
            rows.setdefault(label, r)
    return rows


def _set_by_label(ws, rows, needle, value, col=2):
    """Écrit `value` en colonne `col` sur la 1re ligne dont le libellé A contient needle."""
    needle = _norm(needle)
    for label, r in rows.items():
        if needle in label:
            ws.cell(r, col).value = value
            return r
    return None


def fill_sheet(ws, fiche):
    rows = _label_rows(ws)
    cadre = fiche.get("cadre", {})
    op = fiche.get("operation", {})
    prog = fiche.get("programmation", {})
    is_cs = fiche.get("type_action") == "CS"

    # En-tête : intitulé en B1, code en regard de "Code Gest Ref", priorité en L1
    ws.cell(1, 2).value = fiche.get("intitule")
    _set_by_label(ws, rows, "code gest ref", fiche.get("code_action"))
    if fiche.get("priorite") is not None:
        ws.cell(1, 12).value = f"Priorité {fiche['priorite']}"  # L1

    # 1) Cadre de l'action
    # CS  -> "Indicateur de réponse" / "Niveau d'exigence"
    # hors CS -> "Indicateur de pression" / "Résultats attendus"
    _set_by_label(ws, rows, "indicateur de", cadre.get("indicateur"))
    if is_cs:
        _set_by_label(ws, rows, "niveau d'exigence", cadre.get("niveau_exigence_ou_resultat"))
    else:
        _set_by_label(ws, rows, "resultats attendus", cadre.get("niveau_exigence_ou_resultat"))
    _set_by_label(ws, rows, "olt", cadre.get("olt_intitule"))
    _set_by_label(ws, rows, "enjeu", cadre.get("enjeu_intitule"))
    _set_by_label(ws, rows, "metriques", _join(cadre.get("metriques")))

    # 2) Détails de l'opération
    details = op.get("details") or ""
    if op.get("nom_protocole"):
        details = f"{details}\n\nProtocole : {op['nom_protocole']}".strip()
    # Needle spécifique pour éviter de matcher l'en-tête de section "2) Détails de l'opération".
    _set_by_label(ws, rows, "details du suivi" if is_cs else "details de l'action", details)
    _set_by_label(ws, rows, "localisation", op.get("localisation"))
    _set_by_label(ws, rows, "operateurs",
                  _join((op.get("operateurs_internes") or []) + (op.get("operateurs_externes") or [])))
    _set_by_label(ws, rows, "partenaires", _join(op.get("partenaires")))

    # 3) Volet administratif et financier
    _set_by_label(ws, rows, "financeurs",
                  _join(f.get("organisme_ou_financeur") for f in (prog.get("financements") or [])))
    _set_by_label(ws, rows, "indicateurs de reponse", _join(fiche.get("indicateurs_reponse")))
    if prog.get("budget_total_eur") is not None:
        _set_by_label(ws, rows, "budget total", prog["budget_total_eur"])

    _fill_year_grid(ws, rows, prog)


def _clear_example_marks(ws, rows):
    """Le modèle contient des 'x' d'exemple : on les efface pour ne garder que les données."""
    for label, r in rows.items():
        if "periodicite" in label:  # lignes Périodicité (annuelle ET mensuelle)
            for c in range(2, 14):  # colonnes B..M
                if _norm(ws.cell(r, c).value) == "x":
                    ws.cell(r, c).value = None


def _fill_year_grid(ws, rows, prog):
    """Renseigne les vraies années en en-tête et coche la ligne 'Périodicité' annuelle."""
    _clear_example_marks(ws, rows)
    annee_debut = prog.get("annee_debut_plan")
    if not annee_debut:
        return
    # Trouver la ligne d'en-tête "Programmation annuelle" et la ligne "Périodicité" juste après.
    header_row = None
    for label, r in rows.items():
        if "programmation annuelle" in label:
            header_row = r
            break
    if header_row is None:
        return
    period_row = header_row + 1  # "Périodicité" dans le modèle
    # Le modèle expose N..N+8 à partir de la colonne D (=4). 9 colonnes.
    first_col, n_cols = 4, 9
    # Remplacer N, N+1... par les vraies années (plus lisible pour le relecteur).
    for k in range(n_cols):
        ws.cell(header_row, first_col + k).value = annee_debut + k
    for entry in prog.get("calendrier_annuel") or []:
        if not entry.get("programme"):
            continue
        offset = entry["annee"] - annee_debut
        if 0 <= offset < n_cols:
            ws.cell(period_row, first_col + offset).value = "x"
        # offsets >= n_cols (plans de 10 ans, modèle à 9 colonnes) : non représentés.


def convert(json_path: Path, out_path: Path):
    data = json.loads(json_path.read_text(encoding="utf-8"))
    fiches = data.get("fiches", [])
    wb = load_workbook(TEMPLATE)
    tpl_cs, tpl_hors = wb[SHEET_CS], wb[SHEET_HORS_CS]

    used = set()
    for fiche in fiches:
        tpl = tpl_cs if fiche.get("type_action") == "CS" else tpl_hors
        ws = wb.copy_worksheet(tpl)
        # Nom d'onglet = code action (unique, <=31 car., caractères interdits retirés)
        base = (fiche.get("code_action") or fiche.get("intitule") or "fiche")[:28]
        for ch in "[]:*?/\\":
            base = base.replace(ch, "-")
        name, i = base, 1
        while name in used:
            i += 1
            name = f"{base[:25]}_{i}"
        used.add(name)
        ws.title = name
        fill_sheet(ws, fiche)

    # Retirer les onglets modèles vierges
    for s in (SHEET_CS, SHEET_HORS_CS):
        del wb[s]

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"✓ {len(fiches)} fiche(s) -> {out_path}")


def main():
    p = argparse.ArgumentParser(description="JSON fiches -> Excel modèle Cicada")
    p.add_argument("json", type=Path)
    p.add_argument("-o", "--out", type=Path, default=Path("fiches_remplies.xlsx"))
    args = p.parse_args()
    convert(args.json, args.out)


if __name__ == "__main__":
    main()
