"""
Tests du moteur d'import / export Excel de l'arborescence d'un plan (#478).

Couvre :
- la construction du classeur modèle (vide et pré-rempli) ;
- le round-trip export → parse → validate → execute (les décomptes doivent
  correspondre à la source, y compris le partage de facteur #552) ;
- la validation (dry-run) : plan non vide, colonnes obligatoires, codes en
  double, références introuvables, nomenclature inconnue, parent d'indicateur ;
- l'exécution : création effective de l'arborescence et des rattachements.
"""

import io

import pytest
from openpyxl import load_workbook

from apps.plans.models import PlanGestion
from apps.plans.models_enjeux import (
    Enjeu,
    FacteurInfluence,
    Pression,
    ObjectifLongTerme,
    NiveauExigence,
    ObjectifOperationnel,
    ResultatAttendu,
)
from apps.plans.models_enjeux import CorEnjeuTaxon, CorEnjeuHabitat
from apps.plans.models_indicateurs import (
    Indicateur,
    Metrique,
    CorIndicateurTaxon,
)
from apps.plans.services_import import (
    build_arborescence_workbook,
    parse_workbook,
    validate_import,
    execute_import,
    ArborescenceImportError,
)

from tests.factories.users import RoleFactory, SuperAdminFactory
from tests.factories.plans import PlanGestionFactory
from tests.factories.enjeux import (
    EnjeuFactory,
    FcrFactory,
    FacteurInfluenceFactory,
    PressionFactory,
    ObjectifLongTermeFactory,
    NiveauExigenceFactory,
    ObjectifOperationnelFactory,
    ResultatAttenduFactory,
    IndicateurFactory,
    MetriqueFactory,
    NomenclatureEnjeuFactory,
    NomenclatureFcrFactory,
    NomenclatureTypeIndicateurFactory,
    NomenclatureTypeMetriqueFactory,
)

pytestmark = [pytest.mark.django_db, pytest.mark.integration]

DATA_SHEETS = (
    "enjeux",
    "facteurs",
    "pressions",
    "olt",
    "ne",
    "oo",
    "ra",
    "indicateurs",
    "metriques",
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _row(**kwargs):
    """Construit une ligne parsée (ajoute un _row factice si absent)."""
    kwargs.setdefault("_row", 3)
    return kwargs


def _base_nomenclatures():
    """Crée les nomenclatures référencées par mnémonique dans les tests."""
    NomenclatureEnjeuFactory()  # CATEGORIE_ENJEU / ENJEU
    NomenclatureFcrFactory()  # CATEGORIE_ENJEU / FCR
    NomenclatureTypeIndicateurFactory(mnemonique="ETAT")
    NomenclatureTypeIndicateurFactory(mnemonique="REPONSE")
    NomenclatureTypeMetriqueFactory(mnemonique="NUMERIQUE")


def _valid_parsed():
    """Un jeu de lignes valide et complet (les deux branches + FCR direct)."""
    return {
        "enjeux": [
            _row(code="E1", categorie="ENJEU", libelle="Qualité des eaux"),
            _row(code="E2", categorie="FCR", libelle="Ancrage territorial", _row=4),
        ],
        "facteurs": [_row(code="F1", libelle="Déprise agricole", enjeux="E1")],
        "pressions": [_row(code="P1", facteur="F1", libelle="Fermeture du milieu")],
        "olt": [_row(code="O1", enjeu="E1", libelle="Maintenir le bon état")],
        "ne": [_row(code="N1", olt="O1", libelle="Seuil 18 ha")],
        "oo": [
            _row(code="OO1", pressions="P1", libelle="Rouvrir les milieux"),
            _row(code="OO2", enjeu="E2", libelle="Objectif du FCR", _row=4),
        ],
        "ra": [_row(code="R1", oo="OO1", libelle="5 ha/an restaurés")],
        "indicateurs": [
            _row(code="I1", parent="N1", type="ETAT", nom_indicateur="Surface"),
            _row(
                code="I2",
                parent="R1",
                type="REPONSE",
                nom_indicateur="Linéaire",
                _row=4,
            ),
        ],
        "metriques": [
            _row(
                code="M1",
                indicateur="I1",
                nom_metrique="Surface (ha)",
                type_metrique="NUMERIQUE",
            ),
        ],
    }


def _build_source_plan(user):
    """Construit un petit plan source avec les deux branches de l'arborescence."""
    plan = PlanGestionFactory(id_utilisateur_ajout=user)
    enjeu = EnjeuFactory(id_pg=plan, id_utilisateur_ajout=user)

    # Branche état : OLT → NE → Indicateur → Métrique
    olt = ObjectifLongTermeFactory(id_enjeu=enjeu, id_utilisateur_ajout=user)
    ne = NiveauExigenceFactory(id_olt=olt, id_utilisateur_ajout=user)
    ind_etat = IndicateurFactory(id_ne=ne, id_utilisateur_ajout=user)
    MetriqueFactory(id_indicateur=ind_etat, id_utilisateur_ajout=user)

    # Branche opérationnelle : Facteur → Pression → OO → RA → Indicateur → Métrique
    facteur = FacteurInfluenceFactory(id_enjeu=enjeu, id_utilisateur_ajout=user)
    pression = PressionFactory(id_facteur_influence=facteur, id_utilisateur_ajout=user)
    oo = ObjectifOperationnelFactory(pressions=[pression], id_utilisateur_ajout=user)
    ra = ResultatAttenduFactory(id_oo=oo, id_utilisateur_ajout=user)
    ind_rep = IndicateurFactory(
        id_ne=None,
        id_resultat_attendu=ra,
        id_utilisateur_ajout=user,
    )
    MetriqueFactory(id_indicateur=ind_rep, id_utilisateur_ajout=user)

    return plan


# ---------------------------------------------------------------------------
# Construction du classeur
# ---------------------------------------------------------------------------


def test_build_empty_workbook_has_all_sheets():
    content = build_arborescence_workbook(plan=None)
    wb = load_workbook(io.BytesIO(content))
    assert "Lisez-moi" in wb.sheetnames
    assert wb["Listes"].sheet_state == "hidden"
    for name in (
        "Enjeux",
        "Facteurs",
        "Pressions",
        "OLT",
        "NE",
        "OO",
        "RA",
        "Indicateurs",
        "Metriques",
    ):
        assert name in wb.sheetnames
    # Un modèle vide n'a pas de ligne de données (au-delà des en-têtes).
    assert wb["Enjeux"].max_row <= 2


def test_build_prefilled_workbook_writes_rows():
    user = RoleFactory()
    _base_nomenclatures()
    plan = _build_source_plan(user)
    content = build_arborescence_workbook(plan=plan)
    wb = load_workbook(io.BytesIO(content))
    # Première ligne de données = ligne 3.
    assert wb["Enjeux"].cell(row=3, column=1).value == "E1"
    assert wb["Indicateurs"].max_row >= 4  # 2 indicateurs


# ---------------------------------------------------------------------------
# Round-trip export → import
# ---------------------------------------------------------------------------


def test_roundtrip_export_import_matches_counts():
    user = SuperAdminFactory()
    _base_nomenclatures()
    source = _build_source_plan(user)

    content = build_arborescence_workbook(plan=source)
    parsed = parse_workbook(content)

    target = PlanGestionFactory(id_utilisateur_ajout=user)
    report = validate_import(target, parsed)
    assert report.can_import, report.errors

    counts = execute_import(target, parsed, user)

    assert counts["enjeux"] == source.enjeux.count()
    assert target.enjeux.count() == source.enjeux.count()
    # Deux indicateurs (état + réponse), deux métriques.
    assert counts["indicateurs"] == 2
    assert counts["metriques"] == 2
    assert counts["oo"] == 1
    assert counts["ra"] == 1


def test_roundtrip_preserves_shared_factor():
    """Un facteur partagé entre deux enjeux (#552) reste partagé après import."""
    user = SuperAdminFactory()
    _base_nomenclatures()
    source = PlanGestionFactory(id_utilisateur_ajout=user)
    e1 = EnjeuFactory(id_pg=source, id_utilisateur_ajout=user)
    e2 = EnjeuFactory(id_pg=source, id_utilisateur_ajout=user)
    FacteurInfluenceFactory(enjeux=[e1, e2], id_utilisateur_ajout=user)

    content = build_arborescence_workbook(plan=source)
    parsed = parse_workbook(content)
    # Un seul facteur exporté, rattaché à deux enjeux.
    assert len(parsed["facteurs"]) == 1
    assert len(parsed["facteurs"][0]["enjeux"].split(",")) == 2

    target = PlanGestionFactory(id_utilisateur_ajout=user)
    execute_import(target, parsed, user)

    facteurs = FacteurInfluence.objects.filter(enjeux__id_pg=target).distinct()
    assert facteurs.count() == 1
    assert facteurs.first().enjeux.count() == 2


# ---------------------------------------------------------------------------
# Validation (dry-run)
# ---------------------------------------------------------------------------


def test_validate_rejects_non_empty_plan():
    user = RoleFactory()
    _base_nomenclatures()
    plan = PlanGestionFactory(id_utilisateur_ajout=user)
    EnjeuFactory(id_pg=plan, id_utilisateur_ajout=user)  # arborescence existante

    report = validate_import(plan, _valid_parsed())
    assert not report.can_import
    assert any("déjà une arborescence" in i["message"] for i in report.errors)


def test_validate_ok_on_valid_parsed():
    user = RoleFactory()
    _base_nomenclatures()
    plan = PlanGestionFactory(id_utilisateur_ajout=user)
    report = validate_import(plan, _valid_parsed())
    assert report.can_import, report.errors
    assert report.summary["enjeux"] == 2
    assert report.summary["indicateurs"] == 2


def test_validate_missing_required_libelle():
    user = RoleFactory()
    _base_nomenclatures()
    plan = PlanGestionFactory(id_utilisateur_ajout=user)
    parsed = _valid_parsed()
    parsed["enjeux"][0]["libelle"] = ""
    report = validate_import(plan, parsed)
    assert not report.can_import
    assert any(
        i["sheet"] == "Enjeux" and i["column"] == "libelle" for i in report.errors
    )


def test_validate_duplicate_code():
    user = RoleFactory()
    _base_nomenclatures()
    plan = PlanGestionFactory(id_utilisateur_ajout=user)
    parsed = _valid_parsed()
    parsed["enjeux"][1]["code"] = "E1"  # doublon
    report = validate_import(plan, parsed)
    assert not report.can_import
    assert any("double" in i["message"] for i in report.errors)


def test_validate_broken_parent_reference():
    user = RoleFactory()
    _base_nomenclatures()
    plan = PlanGestionFactory(id_utilisateur_ajout=user)
    parsed = _valid_parsed()
    parsed["pressions"][0]["facteur"] = "F999"  # facteur inexistant
    report = validate_import(plan, parsed)
    assert not report.can_import
    assert any(
        i["sheet"] == "Pressions" and "F999" in i["message"] for i in report.errors
    )


def test_validate_unknown_nomenclature():
    user = RoleFactory()
    _base_nomenclatures()
    plan = PlanGestionFactory(id_utilisateur_ajout=user)
    parsed = _valid_parsed()
    parsed["enjeux"][0]["categorie"] = "Catégorie inexistante"
    report = validate_import(plan, parsed)
    assert not report.can_import
    assert any(i["column"] == "categorie" for i in report.errors)


def test_validate_indicateur_parent_not_found():
    user = RoleFactory()
    _base_nomenclatures()
    plan = PlanGestionFactory(id_utilisateur_ajout=user)
    parsed = _valid_parsed()
    parsed["indicateurs"][0]["parent"] = "ZZZ"
    report = validate_import(plan, parsed)
    assert not report.can_import
    assert any(
        i["sheet"] == "Indicateurs" and i["column"] == "parent" for i in report.errors
    )


def test_validate_oo_without_parent():
    user = RoleFactory()
    _base_nomenclatures()
    plan = PlanGestionFactory(id_utilisateur_ajout=user)
    parsed = _valid_parsed()
    parsed["oo"][0]["pressions"] = ""
    parsed["oo"][0]["enjeu"] = ""
    report = validate_import(plan, parsed)
    assert not report.can_import
    assert any(i["sheet"] == "OO" for i in report.errors)


def test_validate_intitule_court_too_long():
    """L'intitulé court est plafonné à 25 caractères."""
    user = RoleFactory()
    _base_nomenclatures()
    plan = PlanGestionFactory(id_utilisateur_ajout=user)
    parsed = _valid_parsed()
    parsed["enjeux"][0]["intitule_court"] = "x" * 26
    report = validate_import(plan, parsed)
    assert not report.can_import
    assert any(
        i["sheet"] == "Enjeux" and i["column"] == "intitule_court"
        for i in report.errors
    )


# ---------------------------------------------------------------------------
# Exécution
# ---------------------------------------------------------------------------


def test_execute_creates_full_arborescence_with_links():
    user = SuperAdminFactory()
    _base_nomenclatures()
    plan = PlanGestionFactory(id_utilisateur_ajout=user)

    counts = execute_import(plan, _valid_parsed(), user)

    assert counts == {
        "enjeux": 2,
        "facteurs": 1,
        "pressions": 1,
        "olt": 1,
        "ne": 1,
        "oo": 2,
        "ra": 1,
        "indicateurs": 2,
        "metriques": 1,
        "taxons": 0,
        "habitats": 0,
    }

    # Rattachements : facteur → enjeu E1, OO1 → pression, OO2 → enjeu FCR direct.
    facteur = FacteurInfluence.objects.filter(enjeux__id_pg=plan).first()
    assert facteur is not None and facteur.enjeux.count() == 1

    oo_direct = ObjectifOperationnel.objects.get(libelle="Objectif du FCR")
    assert oo_direct.id_enjeu is not None and oo_direct.pressions.count() == 0

    oo_pression = ObjectifOperationnel.objects.get(libelle="Rouvrir les milieux")
    assert oo_pression.pressions.count() == 1

    # Indicateur XOR : un rattaché au NE, l'autre au RA.
    ind_etat = Indicateur.objects.get(nom_indicateur="Surface")
    assert ind_etat.id_ne is not None and ind_etat.id_resultat_attendu is None
    ind_rep = Indicateur.objects.get(nom_indicateur="Linéaire")
    assert ind_rep.id_resultat_attendu is not None and ind_rep.id_ne is None


def test_execute_refuses_when_invalid():
    user = RoleFactory()
    _base_nomenclatures()
    plan = PlanGestionFactory(id_utilisateur_ajout=user)
    parsed = _valid_parsed()
    parsed["enjeux"][0]["libelle"] = ""  # rend le fichier invalide
    with pytest.raises(ValueError):
        execute_import(plan, parsed, user)
    assert plan.enjeux.count() == 0  # transaction annulée


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------


def test_parse_rejects_non_xlsx():
    with pytest.raises(ArborescenceImportError):
        parse_workbook(b"ceci n'est pas un classeur")


# ---------------------------------------------------------------------------
# Taxons / Habitats (TaxRef / HabRef)
# ---------------------------------------------------------------------------


def test_execute_creates_taxons_and_habitats():
    user = SuperAdminFactory()
    _base_nomenclatures()
    plan = PlanGestionFactory(id_utilisateur_ajout=user)
    parsed = _valid_parsed()
    parsed["taxons"] = [
        {"cible": "E1", "cd_nom": 60585, "nom": "Loup gris", "_row": 3},
        {"cible": "I1", "cd_nom": "99999", "nom": "", "_row": 4},
    ]
    parsed["habitats"] = [
        {"cible": "E1", "cd_hab": "24.1", "nom": "Rivières", "_row": 3},
        {"cible": "I1", "cd_hab": "22.1", "nom": "Eaux douces", "_row": 4},
    ]

    counts = execute_import(plan, parsed, user)
    assert counts["taxons"] == 2
    assert counts["habitats"] == 2

    enjeu = Enjeu.objects.get(id_pg=plan, libelle="Qualité des eaux")
    assert enjeu.taxons.count() == 1
    assert enjeu.taxons.first().cd_nom == 60585
    assert enjeu.habitats.count() == 1

    ind = Indicateur.objects.get(nom_indicateur="Surface")
    assert ind.taxons.count() == 1
    assert ind.habitats.count() == 1


def test_roundtrip_preserves_taxons_habitats():
    user = SuperAdminFactory()
    _base_nomenclatures()
    source = _build_source_plan(user)
    enjeu = source.enjeux.first()
    CorEnjeuTaxon.objects.create(id_enjeu=enjeu, cd_nom=60585, nom_complet="Loup")
    CorEnjeuHabitat.objects.create(id_enjeu=enjeu, cd_hab="24.1", lb_hab_fr="Rivières")

    content = build_arborescence_workbook(plan=source)
    parsed = parse_workbook(content)
    assert len(parsed["taxons"]) >= 1
    assert len(parsed["habitats"]) >= 1

    target = PlanGestionFactory(id_utilisateur_ajout=user)
    execute_import(target, parsed, user)
    assert CorEnjeuTaxon.objects.filter(id_enjeu__id_pg=target, cd_nom=60585).exists()
    assert CorEnjeuHabitat.objects.filter(
        id_enjeu__id_pg=target, cd_hab="24.1"
    ).exists()


def test_validate_taxon_bad_cd_nom():
    user = RoleFactory()
    _base_nomenclatures()
    plan = PlanGestionFactory(id_utilisateur_ajout=user)
    parsed = _valid_parsed()
    parsed["taxons"] = [{"cible": "E1", "cd_nom": "abc", "_row": 3}]
    report = validate_import(plan, parsed)
    assert not report.can_import
    assert any(
        i["sheet"] == "Taxons" and i["column"] == "cd_nom" for i in report.errors
    )


def test_validate_bio_unknown_cible():
    user = RoleFactory()
    _base_nomenclatures()
    plan = PlanGestionFactory(id_utilisateur_ajout=user)
    parsed = _valid_parsed()
    parsed["habitats"] = [{"cible": "Z9", "cd_hab": "24.1", "_row": 3}]
    report = validate_import(plan, parsed)
    assert not report.can_import
    assert any(
        i["sheet"] == "Habitats" and i["column"] == "cible" for i in report.errors
    )


# ---------------------------------------------------------------------------
# Types écologiques / socio-éco (colonnes multi-valeurs → booléens de l'enjeu)
# ---------------------------------------------------------------------------


def test_execute_maps_enjeu_type_flags():
    """Les types écologiques/socio-éco (multi-valeurs) alimentent les booléens."""
    user = SuperAdminFactory()
    _base_nomenclatures()
    plan = PlanGestionFactory(id_utilisateur_ajout=user)
    parsed = _valid_parsed()
    parsed["enjeux"][0]["categorie_ecologique"] = "Oui"
    parsed["enjeux"][0]["types_ecologiques"] = "Habitat,Espèce"
    parsed["enjeux"][1]["categorie_ecologique"] = "Non"
    parsed["enjeux"][1]["types_socioeco"] = "Usages,Développement durable"

    execute_import(plan, parsed, user)

    eco = Enjeu.objects.get(id_pg=plan, libelle="Qualité des eaux")
    assert eco.categorie_ecologique is True
    assert eco.habitat is True and eco.espece is True
    assert eco.patrimoine_geologique is False

    socio = Enjeu.objects.get(id_pg=plan, libelle="Ancrage territorial")
    assert socio.categorie_ecologique is False
    assert socio.usages is True and socio.developpement_durable is True
