"""
Tests du module 2 (actions) de l'import Excel (#478).

Couvre la construction du classeur d'actions (onglet de référence des
indicateurs + onglet Actions), la validation (dry-run) et l'exécution
(création des opérations + programmation annuelle).
"""

import io
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from apps.plans.models_operations import (
    Operation,
    OperationAnnee,
    OperationAnneeRH,
    Poste,
    PosteFonction,
    Fonction,
)
from apps.plans.services_import_actions import (
    build_actions_workbook,
    parse_actions_workbook,
    validate_actions_import,
    execute_actions_import,
    _plan_indicateurs,
)

from tests.factories.users import RoleFactory, SuperAdminFactory
from tests.factories.plans import PlanGestionFactory
from tests.factories.enjeux import (
    EnjeuFactory,
    ObjectifLongTermeFactory,
    NiveauExigenceFactory,
    IndicateurFactory,
    OperationFactory,
    TypeNomenclatureFactory,
)
from tests.factories.core import NomenclatureFactory

pytestmark = [pytest.mark.django_db, pytest.mark.integration]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _plan_with_indicateur(user):
    """Plan minimal avec un indicateur (branche état)."""
    plan = PlanGestionFactory(id_utilisateur_ajout=user)
    enjeu = EnjeuFactory(id_pg=plan, id_utilisateur_ajout=user)
    olt = ObjectifLongTermeFactory(id_enjeu=enjeu, id_utilisateur_ajout=user)
    ne = NiveauExigenceFactory(id_olt=olt, id_utilisateur_ajout=user)
    ind = IndicateurFactory(id_ne=ne, id_utilisateur_ajout=user)
    return plan, ind


def _type_action(label="Chantier CS8"):
    type_nom = TypeNomenclatureFactory(mnemonique="TYPE_ACTION")
    return NomenclatureFactory(id_type=type_nom, mnemonique="CS8", label=label)


def _add_poste(plan, libelle="Garde"):
    """Crée un poste (avec une fonction) rattaché au plan."""
    fonction, _ = Fonction.objects.get_or_create(libelle=libelle)
    poste = Poste.objects.create(id_pg=plan, nombre=1)
    PosteFonction.objects.create(id_poste=poste, id_fonction=fonction)
    return poste


def _parsed(ind_code, ind_id, **overrides):
    """Un jeu « actions » minimal valide."""
    action = {
        "code": "A1",
        "indicateur": ind_code,
        "libelle": "Débroussaillage",
        "annee_min": 2024,
        "annee_max": 2026,
        "_row": 3,
    }
    action.update(overrides)
    return {"actions": [action], "indicateurs": {ind_code: ind_id}}


# ---------------------------------------------------------------------------
# Construction du classeur
# ---------------------------------------------------------------------------


def test_build_actions_workbook_lists_indicateurs():
    user = RoleFactory()
    plan, ind = _plan_with_indicateur(user)
    wb = load_workbook(io.BytesIO(build_actions_workbook(plan=plan)))
    assert wb.sheetnames == [
        "Lisez-moi",
        "Indicateurs",
        "Postes",
        "Listes",
        "Actions",
        "Budgets",
        "RH",
    ]
    assert wb["Listes"].sheet_state == "hidden"
    # L'indicateur du plan est listé, avec son id technique.
    assert wb["Indicateurs"].cell(row=3, column=1).value == "I1"
    assert wb["Indicateurs"].cell(row=3, column=4).value == ind.id_indicateur


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------


def test_parse_reads_reference_and_actions():
    user = RoleFactory()
    plan, ind = _plan_with_indicateur(user)
    wb = load_workbook(io.BytesIO(build_actions_workbook(plan=plan)))
    wb["Actions"].cell(row=3, column=1, value="A1")
    wb["Actions"].cell(row=3, column=2, value="I1")
    wb["Actions"].cell(row=3, column=3, value="Action test")
    buf = io.BytesIO()
    wb.save(buf)
    parsed = parse_actions_workbook(buf.getvalue())
    assert parsed["indicateurs"] == {"I1": ind.id_indicateur}
    assert len(parsed["actions"]) == 1
    assert parsed["actions"][0]["libelle"] == "Action test"


def test_parse_rejects_non_xlsx():
    from apps.plans.services_import import ArborescenceImportError

    with pytest.raises(ArborescenceImportError):
        parse_actions_workbook(b"pas un classeur")


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------


def test_validate_ok():
    user = RoleFactory()
    plan, ind = _plan_with_indicateur(user)
    report = validate_actions_import(plan, _parsed("I1", ind.id_indicateur))
    assert report.can_import, report.errors
    assert report.summary["actions"] == 1


def test_validate_refuses_plan_without_indicateurs():
    user = RoleFactory()
    plan = PlanGestionFactory(id_utilisateur_ajout=user)
    report = validate_actions_import(plan, {"actions": [], "indicateurs": {}})
    assert not report.can_import
    assert any("arborescence" in i["message"] for i in report.errors)


def test_validate_refuses_plan_with_existing_actions():
    user = RoleFactory()
    plan, ind = _plan_with_indicateur(user)
    OperationFactory(id_indicateur=ind, id_utilisateur_ajout=user)
    report = validate_actions_import(plan, _parsed("I1", ind.id_indicateur))
    assert not report.can_import
    assert any("déjà des actions" in i["message"] for i in report.errors)


def test_validate_missing_libelle():
    user = RoleFactory()
    plan, ind = _plan_with_indicateur(user)
    report = validate_actions_import(plan, _parsed("I1", ind.id_indicateur, libelle=""))
    assert not report.can_import
    assert any(i["column"] == "libelle" for i in report.errors)


def test_validate_unknown_indicateur_code():
    user = RoleFactory()
    plan, ind = _plan_with_indicateur(user)
    parsed = _parsed("I1", ind.id_indicateur)
    parsed["actions"][0]["indicateur"] = "I999"
    report = validate_actions_import(plan, parsed)
    assert not report.can_import
    assert any(i["column"] == "indicateur" for i in report.errors)


def test_validate_indicateur_not_in_plan():
    """Un id de référence qui n'appartient pas au plan est rejeté."""
    user = RoleFactory()
    plan, ind = _plan_with_indicateur(user)
    parsed = _parsed("I1", ind.id_indicateur + 99999)
    report = validate_actions_import(plan, parsed)
    assert not report.can_import
    assert any("n'appartient pas" in i["message"] for i in report.errors)


def test_validate_bad_year_order():
    user = RoleFactory()
    plan, ind = _plan_with_indicateur(user)
    parsed = _parsed("I1", ind.id_indicateur, annee_min=2028, annee_max=2024)
    report = validate_actions_import(plan, parsed)
    assert not report.can_import
    assert any(i["column"] == "annee_max" for i in report.errors)


def test_validate_unknown_type_action():
    user = RoleFactory()
    plan, ind = _plan_with_indicateur(user)
    _type_action()  # crée la nomenclature TYPE_ACTION (label « Chantier CS8 »)
    parsed = _parsed("I1", ind.id_indicateur, type_action="Type inexistant")
    report = validate_actions_import(plan, parsed)
    assert not report.can_import
    assert any(i["column"] == "type_action" for i in report.errors)


# ---------------------------------------------------------------------------
# Exécution
# ---------------------------------------------------------------------------


def test_execute_creates_operations_and_years():
    user = SuperAdminFactory()
    plan, ind = _plan_with_indicateur(user)
    type_action = _type_action()
    parsed = _parsed("I1", ind.id_indicateur, type_action=type_action.label)

    counts = execute_actions_import(plan, parsed, user)
    assert counts == {"actions": 1, "annees": 3, "budgets": 0, "rh": 0}

    op = Operation.objects.get(libelle="Débroussaillage")
    assert op.id_indicateur_id == ind.id_indicateur
    assert op.statut == "draft"
    assert op.id_type_action_id == type_action.id_nomenclature
    annees = set(
        OperationAnnee.objects.filter(id_operation=op).values_list("annee", flat=True)
    )
    assert annees == {2024, 2025, 2026}


def test_execute_refuses_when_invalid():
    user = RoleFactory()
    plan, ind = _plan_with_indicateur(user)
    parsed = _parsed("I1", ind.id_indicateur, libelle="")
    with pytest.raises(ValueError):
        execute_actions_import(plan, parsed, user)
    assert not Operation.objects.filter(id_indicateur=ind).exists()


# ---------------------------------------------------------------------------
# Budgets et RH (#560)
# ---------------------------------------------------------------------------


def _parsed_with_budget_rh(ind_code, ind_id, poste_code, poste_id):
    """Jeu actions + budgets + RH (parsé) valide."""
    return {
        "actions": [
            {
                "code": "A1",
                "indicateur": ind_code,
                "libelle": "Débroussaillage",
                "annee_min": 2024,
                "annee_max": 2026,
                "_row": 3,
            },
        ],
        "indicateurs": {ind_code: ind_id},
        "postes": {poste_code: poste_id},
        "budgets": [
            {
                "action": "A1",
                "annee": 2024,
                "budget_fonctionnement": "1000",
                "budget_investissement": "500",
                "_row": 3,
            },
        ],
        "rh": [
            {
                "action": "A1",
                "annee": 2024,
                "poste": poste_code,
                "jours": "10",
                "finance": "Oui",
                "_row": 3,
            },
        ],
    }


def test_build_actions_workbook_lists_postes():
    user = RoleFactory()
    plan, ind = _plan_with_indicateur(user)
    poste = _add_poste(plan)
    wb = load_workbook(io.BytesIO(build_actions_workbook(plan=plan)))
    assert (
        "Postes" in wb.sheetnames
        and "Budgets" in wb.sheetnames
        and "RH" in wb.sheetnames
    )
    assert wb["Postes"].cell(row=3, column=1).value == "Q1"
    assert wb["Postes"].cell(row=3, column=4).value == poste.id_poste


def test_execute_creates_budgets_and_rh():
    user = SuperAdminFactory()
    plan, ind = _plan_with_indicateur(user)
    poste = _add_poste(plan)
    parsed = _parsed_with_budget_rh("I1", ind.id_indicateur, "Q1", poste.id_poste)

    counts = execute_actions_import(plan, parsed, user)
    assert counts == {"actions": 1, "annees": 3, "budgets": 1, "rh": 1}

    op = Operation.objects.get(libelle="Débroussaillage")
    assert op.ventilation_mode == "by_type"
    assert op.declinaison_par_poste is True

    oa = OperationAnnee.objects.get(id_operation=op, annee=2024)
    assert oa.budget_fonctionnement == Decimal("1000")
    assert oa.budget_investissement == Decimal("500")

    rh = OperationAnneeRH.objects.get(id_operation_annee=oa)
    assert rh.id_poste_id == poste.id_poste
    assert rh.jours == Decimal("10")
    assert rh.finance is True


def test_roundtrip_budget_rh_via_workbook():
    """Injection dans le classeur généré puis import de bout en bout."""
    user = SuperAdminFactory()
    plan, ind = _plan_with_indicateur(user)
    poste = _add_poste(plan)
    wb = load_workbook(io.BytesIO(build_actions_workbook(plan=plan)))
    ind_code = wb["Indicateurs"].cell(row=3, column=1).value
    poste_code = wb["Postes"].cell(row=3, column=1).value
    a = wb["Actions"]
    a.cell(row=3, column=1, value="A1")
    a.cell(row=3, column=2, value=ind_code)
    a.cell(row=3, column=3, value="Action")
    a.cell(row=3, column=6, value=2024)
    a.cell(row=3, column=7, value=2024)
    b = wb["Budgets"]
    b.cell(row=3, column=1, value="A1")
    b.cell(row=3, column=2, value=2024)
    b.cell(row=3, column=3, value=800)
    rh = wb["RH"]
    rh.cell(row=3, column=1, value="A1")
    rh.cell(row=3, column=2, value=2024)
    rh.cell(row=3, column=3, value=poste_code)
    rh.cell(row=3, column=4, value=5)
    rh.cell(row=3, column=5, value="Non")
    buf = io.BytesIO()
    wb.save(buf)

    parsed = parse_actions_workbook(buf.getvalue())
    report = validate_actions_import(plan, parsed)
    assert report.can_import, report.errors
    counts = execute_actions_import(plan, parsed, user)
    assert counts["budgets"] == 1 and counts["rh"] == 1

    rh_line = OperationAnneeRH.objects.get()
    assert rh_line.finance is False and rh_line.jours == Decimal("5")


def test_validate_rh_unknown_poste():
    user = RoleFactory()
    plan, ind = _plan_with_indicateur(user)
    poste = _add_poste(plan)
    parsed = _parsed_with_budget_rh("I1", ind.id_indicateur, "Q1", poste.id_poste)
    parsed["rh"][0]["poste"] = "Q999"  # code de poste inexistant
    report = validate_actions_import(plan, parsed)
    assert not report.can_import
    assert any(i["sheet"] == "RH" and i["column"] == "poste" for i in report.errors)


def test_validate_budget_negative():
    user = RoleFactory()
    plan, ind = _plan_with_indicateur(user)
    poste = _add_poste(plan)
    parsed = _parsed_with_budget_rh("I1", ind.id_indicateur, "Q1", poste.id_poste)
    parsed["budgets"][0]["budget_fonctionnement"] = "-100"
    report = validate_actions_import(plan, parsed)
    assert not report.can_import
    assert any(
        i["sheet"] == "Budgets" and "négatif" in i["message"] for i in report.errors
    )


def test_validate_budget_unknown_action():
    user = RoleFactory()
    plan, ind = _plan_with_indicateur(user)
    parsed = _parsed_with_budget_rh("I1", ind.id_indicateur, "Q1", 1)
    parsed["budgets"][0]["action"] = "A999"
    report = validate_actions_import(plan, parsed)
    assert not report.can_import
    assert any(
        i["sheet"] == "Budgets" and i["column"] == "action" for i in report.errors
    )


def test_validate_budget_year_outside_span_is_warning():
    user = RoleFactory()
    plan, ind = _plan_with_indicateur(user)
    poste = _add_poste(plan)
    parsed = _parsed_with_budget_rh("I1", ind.id_indicateur, "Q1", poste.id_poste)
    parsed["budgets"][0]["annee"] = 2099  # hors [2024, 2026]
    report = validate_actions_import(plan, parsed)
    assert report.can_import  # avertissement, pas erreur
    assert any(
        i["level"] == "warning" and i["sheet"] == "Budgets" for i in report.issues
    )
