"""
Tests de l'export Excel « fiche action » (#626).

Couvre les manques remontés en recette :
- code action local du plan (CS1…) **et** code du référentiel Gestref ;
- indicateurs de réponse exclus du cadre « indicateur d'état / de pression » ;
- intitulé de l'indicateur de réponse suivi de sa métrique entre parenthèses
  quand la métrique n'utilise pas de grille ;
- grille de scoring 5 paliers rendue quand la métrique est au format GRILLE ;
- descriptif et objectifs du protocole du suivi ;
- petite carte du contour de l'emprise de l'action.
"""

import io

import pytest
from django.contrib.gis.geos import MultiPolygon, Polygon
from openpyxl import load_workbook

from apps.core.models import Nomenclature, TypeNomenclature
from apps.plans import services_export_fiche_action as fiche
from tests.factories.enjeux import (
    EnjeuFactory, IndicateurFactory, MetriqueFactory, NiveauExigenceFactory,
    ObjectifLongTermeFactory, OperationFactory,
)
from tests.factories.plans import PlanGestionFactory


def _nomenclature(type_mnemo, mnemo, label):
    t, _ = TypeNomenclature.objects.get_or_create(
        mnemonique=type_mnemo, defaults={'label': type_mnemo})
    n, _ = Nomenclature.objects.get_or_create(
        id_type=t, mnemonique=mnemo,
        defaults={'cd_nomenclature': mnemo[:10], 'label': label})
    return n


@pytest.mark.django_db
class TestExportFicheAction:
    """#626 — codes local/Gestref, protocole détaillé, indicateurs de réponse,
    grille de métrique et carte de localisation."""

    @pytest.fixture
    def plan_fiche(self):
        """Plan avec une action CS portant :

        - un indicateur d'**état** (le seul attendu dans le cadre de l'action) ;
        - un indicateur de **réponse** avec une métrique simple et une métrique
          en grille (type TEXTE, 5 libellés) ;
        - un suivi avec un protocole standardisé documenté ;
        - une emprise (pour la carte).
        """
        from django.contrib.gis.geos import Polygon
        from tests.factories.enjeux import ProtocoleFactory, SuiviInventaireFactory

        plan = PlanGestionFactory(annee_debut=2024, annee_fin=2026)
        t_etat = _nomenclature('TYPE_INDICATEUR', 'ETAT', 'État')
        t_rep = _nomenclature('TYPE_INDICATEUR', 'REPONSE', 'Réponse')
        t_num = _nomenclature('TYPE_METRIQUE', 'NUMERIQUE', 'Intervalle numérique')
        t_texte = _nomenclature('TYPE_METRIQUE', 'TEXTE', 'Texte')
        f_simple = _nomenclature('FORMAT_METRIQUE', 'SIMPLE', 'Simple')
        f_grille = _nomenclature('FORMAT_METRIQUE', 'GRILLE', 'Grille')
        cs = _nomenclature('CATEGORIE_ACTION_RESERVE', 'CS', 'Connaissance et suivi')
        type_action = _nomenclature('TYPE_ACTION', 'CS8', 'Inventaire de la faune')

        enjeu = EnjeuFactory(id_pg=plan)
        ne = NiveauExigenceFactory(id_olt=ObjectifLongTermeFactory(id_enjeu=enjeu))

        ind_etat = IndicateurFactory(id_ne=ne, type_indicateur=t_etat,
                                     nom_indicateur='État de la lande')
        met_etat = MetriqueFactory(id_indicateur=ind_etat, nom_metrique='Recouvrement',
                                   type_metrique=t_num)

        ind_rep = IndicateurFactory(id_ne=ne, type_indicateur=t_rep,
                                    nom_indicateur='Panneaux posés')
        met_simple = MetriqueFactory(id_indicateur=ind_rep, type_metrique=t_num,
                                     nom_metrique='Nombre de panneaux',
                                     format_metrique=f_simple)
        met_grille = MetriqueFactory(
            id_indicateur=ind_rep, type_metrique=t_texte, unite='',
            nom_metrique='Qualité de la pose', format_metrique=f_grille,
            score_1_label='Aucun panneau', score_2_label='Quelques panneaux',
            score_3_label='Moitié posée', score_4_label='Presque tout',
            score_5_label='Tout posé',
        )

        proto = ProtocoleFactory(
            protocole_dans_campanule=True, protocole_campanule_nom='STOC EPS',
            description_protocole='Écoutes ponctuelles standardisées',
            objectif_protocole='Suivre les tendances des oiseaux communs',
        )
        suivi = SuiviInventaireFactory(id_pg=plan, protocoles=[proto])

        op = OperationFactory(
            metriques=[met_etat, met_simple, met_grille],
            id_categorie_action_reserve=cs, id_type_action=type_action,
            id_suivi=suivi, code_operation=None, numero_manuel=None,
            geom=Polygon(((2.0, 43.0), (2.1, 43.0), (2.1, 43.1),
                          (2.0, 43.1), (2.0, 43.0)), srid=4326),
        )
        return {'plan': plan, 'op': op}

    @pytest.fixture
    def sheet(self, plan_fiche):
        from apps.plans.services_export_fiche_action import build_fiche_action_workbook
        wb = load_workbook(io.BytesIO(build_fiche_action_workbook(plan_fiche['plan'])))
        # sans `code_operation`, l'onglet prend le code local calculé
        assert 'CS1' in wb.sheetnames
        return wb['CS1']

    @staticmethod
    def _values(ws):
        """{label colonne A: valeur colonne D} (première occurrence)."""
        out = {}
        for r in range(1, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if label and label not in out:
                out[label] = ws.cell(r, 4).value
        return out

    @staticmethod
    def _all_text(ws):
        return [ws.cell(r, c).value
                for r in range(1, ws.max_row + 1)
                for c in range(1, ws.max_column + 1)
                if isinstance(ws.cell(r, c).value, str)]

    def test_code_local_et_code_gestref(self, sheet):
        vals = self._values(sheet)
        assert vals['Code action (local)'] == 'CS1'
        assert 'CS8' in (vals['Code action (référentiel Gestref)'] or '')

    def test_indicateur_reponse_exclu_du_cadre(self, sheet):
        vals = self._values(sheet)
        assert vals["Indicateur d'état"] == 'État de la lande'
        assert vals['Métriques'] == 'Recouvrement'

    def test_indicateur_reponse_avec_metrique_entre_parentheses(self, sheet):
        vals = self._values(sheet)
        assert vals['Indicateurs de réponse'] == 'Panneaux posés (Nombre de panneaux)'

    def test_grille_de_metrique_rendue(self, sheet):
        texts = self._all_text(sheet)
        assert 'Panneaux posés — Qualité de la pose' in texts
        assert 'Très mauvais\n= 1' in texts
        assert 'Très bon\n= 5' in texts
        assert 'Aucun panneau' in texts
        assert 'Tout posé' in texts

    def test_protocole_dans_details_du_suivi(self, sheet):
        """#626 — protocole (nom, descriptif, objectifs) intégré textuellement
        dans « Détails du suivi », sans lignes séparées."""
        vals = self._values(sheet)
        details = vals['Détails du suivi'] or ''
        assert 'STOC EPS' in details, details
        assert 'Écoutes ponctuelles standardisées' in details, details
        assert 'Suivre les tendances des oiseaux communs' in details, details
        # plus de lignes séparées pour le protocole
        assert 'Descriptif du protocole' not in vals
        assert 'Objectifs du protocole' not in vals

    def test_indicateur_reponse_ne_contamine_pas_les_autres_actions(self, plan_fiche):
        """#626 — un indicateur de réponse lié à UNE action n'apparaît pas sur une
        autre action qui ne partage que le NE (via son indicateur d'état)."""
        from apps.plans.services_export_fiche_action import _reponse_indicateurs
        from tests.factories.enjeux import OperationFactory
        op_rep = plan_fiche['op']
        met_etat = op_rep.metriques.get(nom_metrique='Recouvrement')
        # action liée uniquement à la métrique d'état (même NE, aucune réponse)
        op_etat_seul = OperationFactory(metriques=[met_etat])
        # l'action liée à la réponse la voit toujours...
        assert any(i.nom_indicateur == 'Panneaux posés'
                   for i in _reponse_indicateurs(op_rep))
        # ...mais l'action « état seul » n'hérite pas de l'indicateur du NE
        assert _reponse_indicateurs(op_etat_seul) == []

    def test_carte_de_localisation(self, sheet):
        assert "Emprise de l'action" in self._values(sheet)
        assert len(sheet._images) == 1

    def test_carte_placee_dans_le_champ_de_son_libelle(self, sheet):
        """#626 — la vignette occupe le champ « valeur », en face du libellé.

        Ancrée en colonne A, elle s'affichait sous « Emprise de l'action », dans
        la marge des libellés, et non dans le cadre attendu à droite.
        """
        ligne_emprise = next(
            r for r in range(1, sheet.max_row + 1)
            if sheet.cell(r, 1).value == "Emprise de l'action"
        )
        ancre = sheet._images[0].anchor._from
        # openpyxl indexe l'ancre à partir de 0 : colonne D = 3, ligne N = N-1.
        assert ancre.col == 3, ancre.col
        assert ancre.row == ligne_emprise - 1, (ancre.row, ligne_emprise)


@pytest.mark.django_db
class TestCarteFondDeCarte:
    """#629 — la carte de localisation embarque le fond de carte (tuiles XYZ)
    et se replie proprement sur l'aplat quand le réseau est indisponible."""

    GEOM = MultiPolygon(Polygon(((4.60, 43.50), (4.62, 43.50),
                                 (4.62, 43.52), (4.60, 43.52), (4.60, 43.50))))

    @staticmethod
    def _tile_png(color):
        from PIL import Image as PILImage

        buf = io.BytesIO()
        PILImage.new('RGB', (256, 256), color).save(buf, format='PNG')
        return buf.getvalue()

    @staticmethod
    def _pixels(png):
        from PIL import Image as PILImage

        img = PILImage.open(io.BytesIO(png)).convert('RGB')
        return img, list(img.getdata())

    def test_tuiles_collees_sur_le_fond(self, monkeypatch):
        """Les tuiles téléchargées composent le fond : plus d'aplat beige."""
        urls = []

        def fake_fetch(url):
            urls.append(url)
            return self._tile_png('#8899AA')

        monkeypatch.setattr(fiche, '_TILE_CACHE', {})
        monkeypatch.setattr(fiche, '_fetch_tile', fake_fetch)

        png = fiche._geom_png(self.GEOM)
        img, pixels = self._pixels(png)

        assert urls, 'aucune tuile demandée'
        assert all(u.startswith('https://tile.openstreetmap.org/') for u in urls)
        # Le coin haut-gauche est du fond de carte, pas l'aplat beige (#F3EFEA).
        assert img.getpixel((3, 3)) == (136, 153, 170)
        assert (243, 239, 234) not in pixels
        # L'emprise ne couvre pas toute l'image (le « carré vert » de #629).
        assert pixels.count((136, 153, 170)) > len(pixels) * 0.3

    def test_repli_sans_reseau(self, monkeypatch):
        """Tuiles inaccessibles → aplat + contour, l'export ne casse pas."""
        monkeypatch.setattr(fiche, '_TILE_CACHE', {})
        monkeypatch.setattr(fiche, '_fetch_tile', lambda url: None)

        png = fiche._geom_png(self.GEOM)
        img, pixels = self._pixels(png)

        assert png
        assert img.getpixel((3, 3)) == (243, 239, 234)
        assert any(p != (243, 239, 234) for p in pixels)  # contour dessiné

    def test_fond_desactivable_par_configuration(self, monkeypatch, settings):
        """`EXPORT_MAP_TILE_URL` vide → aucun appel réseau."""
        calls = []
        monkeypatch.setattr(fiche, '_TILE_CACHE', {})
        monkeypatch.setattr(fiche, '_fetch_tile', lambda url: calls.append(url))
        settings.EXPORT_MAP_TILE_URL = ''

        assert fiche._geom_png(self.GEOM)
        assert calls == []
