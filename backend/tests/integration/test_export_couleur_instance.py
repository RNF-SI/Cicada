"""
Couleur des exports paramétrable par instance (#601).

CICADA est déployé par plusieurs structures ; les classeurs et documents qu'elles
produisent doivent pouvoir porter **leur** couleur, choisie par l'administrateur
de l'instance. Ce que ces tests verrouillent :

- la couleur choisie ressort réellement dans les fichiers produits (Excel ET
  Word), et pas seulement dans la configuration ;
- une instance qui n'a rien choisi garde la couleur de CICADA ;
- les couleurs de **score** ne bougent pas : elles font partie de la légende.
"""

import io

import pytest
from docx import Document
from openpyxl import load_workbook

from apps.core.models import SiteConfiguration
from apps.plans.export_theme import COULEUR_DEFAUT, argb, rgb
from tests.factories.enjeux import (
    EnjeuFactory, IndicateurFactory, MetriqueFactory, NiveauExigenceFactory,
    ObjectifLongTermeFactory, OperationFactory,
)
from tests.factories.plans import PlanGestionFactory

#: Une couleur franchement différente du bleu-vert CICADA, pour que l'assertion
#: ne puisse pas passer par accident.
VIOLET = '#7A1FA2'


@pytest.fixture
def plan_export(db):
    plan = PlanGestionFactory(annee_debut=2024, annee_fin=2026)
    enjeu = EnjeuFactory(id_pg=plan, libelle="Landes sèches")
    ne = NiveauExigenceFactory(id_olt=ObjectifLongTermeFactory(id_enjeu=enjeu))
    met = MetriqueFactory(
        id_indicateur=IndicateurFactory(id_ne=ne),
        nom_metrique='Recouvrement', unite='%',
        score_1_inf=0, score_1_sup=10,
    )
    OperationFactory(metriques=[met], code_operation=None, numero_manuel=None)
    return plan


def _choisir(couleur):
    config = SiteConfiguration.get_instance()
    config.export_color = couleur
    config.save()


def _couleurs_de_police(ws):
    return {
        (cell.font.color.rgb if cell.font and cell.font.color else None)
        for row in ws.iter_rows()
        for cell in row
    }


@pytest.mark.integration
class TestCouleurExportInstance:

    def test_par_defaut_la_couleur_de_cicada(self, plan_export):
        assert SiteConfiguration.get_instance().export_color == COULEUR_DEFAUT
        assert argb() == 'FF025359'

    def test_la_couleur_choisie_est_convertie_pour_les_deux_formats(self, db):
        _choisir(VIOLET)

        assert argb() == 'FF7A1FA2'      # openpyxl : alpha en tête
        assert rgb() == (0x7A, 0x1F, 0xA2)  # python-docx

    def test_arborescence_prend_la_couleur_de_linstance(self, plan_export):
        from apps.plans.services_export_arbo import build_presentation_workbook

        _choisir(VIOLET)
        wb = load_workbook(io.BytesIO(build_presentation_workbook(plan_export)))
        couleurs = _couleurs_de_police(wb[wb.sheetnames[0]])

        assert 'FF7A1FA2' in couleurs
        assert 'FF025359' not in couleurs

    def test_les_couleurs_de_score_ne_bougent_pas(self, plan_export):
        """La grille de lecture est une légende : elle doit rester lisible
        d'une structure à l'autre."""
        from apps.plans.services_export_arbo import build_presentation_workbook

        _choisir(VIOLET)
        wb = load_workbook(io.BytesIO(build_presentation_workbook(plan_export)))
        ws = wb[wb.sheetnames[0]]
        fonds = {
            cell.fill.fgColor.rgb
            for row in ws.iter_rows() for cell in row
            if cell.fill and cell.fill.fgColor
        }

        assert 'FFFF0000' in fonds   # score 1 — rouge
        assert 'FF00CCFF' in fonds   # score 5 — cyan

    def test_fiche_action_prend_la_couleur_de_linstance(self, plan_export):
        from apps.plans.services_export_fiche_action import (
            build_fiche_action_workbook,
        )

        _choisir(VIOLET)
        wb = load_workbook(io.BytesIO(build_fiche_action_workbook(plan_export)))
        ws = wb[wb.sheetnames[0]]
        fonds = {
            cell.fill.fgColor.rgb
            for row in ws.iter_rows() for cell in row
            if cell.fill and cell.fill.fgColor
        }

        assert 'FF7A1FA2' in fonds, "le bandeau de la fiche garde l'ancienne couleur"
        assert 'FF7A1FA2' in _couleurs_de_police(ws)

    def test_budget_prend_la_couleur_de_linstance(self, plan_export):
        from apps.plans.services_export_budget_rh import (
            build_budget_previsionnel_workbook,
        )

        _choisir(VIOLET)
        wb = load_workbook(
            io.BytesIO(build_budget_previsionnel_workbook(plan_export))
        )
        couleurs = set()
        for nom in wb.sheetnames:
            couleurs |= _couleurs_de_police(wb[nom])

        assert 'FF7A1FA2' in couleurs

    def test_document_word_prend_la_couleur_de_linstance(self, plan_export):
        from apps.plans.services_export_word import build_plan_docx

        _choisir(VIOLET)
        doc = Document(io.BytesIO(build_plan_docx(plan_export)))
        couleurs = {
            run.font.color.rgb
            for para in doc.paragraphs for run in para.runs
            if run.font.color and run.font.color.rgb
        }

        assert '7A1FA2' in {str(couleur) for couleur in couleurs}

    def test_revenir_a_la_couleur_par_defaut_est_pris_en_compte(self, plan_export):
        """Les styles sont des variables de module : le retour arrière doit
        les réaligner, pas rester bloqué sur la dernière couleur choisie."""
        from apps.plans.services_export_arbo import build_presentation_workbook

        _choisir(VIOLET)
        build_presentation_workbook(plan_export)
        _choisir(COULEUR_DEFAUT)

        wb = load_workbook(io.BytesIO(build_presentation_workbook(plan_export)))
        couleurs = _couleurs_de_police(wb[wb.sheetnames[0]])

        assert 'FF025359' in couleurs
        assert 'FF7A1FA2' not in couleurs
