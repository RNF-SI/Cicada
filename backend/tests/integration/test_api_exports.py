"""
Tests des exports Excel/Word d'un plan de gestion.

Couvre :
- la logique financière partagée (``services_export_finance``) : coût salarial
  recalculé, ventilation fonctionnement/investissement, prestataire fonct+invest,
  bénévoles, réalisé, modes avec/sans ventilation par organisme ;
- la génération des classeurs (fiche action, budget/RH prév. et suivi) et de la
  fiche Word : validité du fichier, onglets attendus, quelques valeurs clés.
"""

import io

import pytest
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.core.models import Nomenclature, TypeNomenclature
from apps.plans.models_operations import (
    Fonction, Operation, OperationAnnee, OperationAnneeOrganisme,
    OperationAnneeRH, Poste, PosteFonction, RealisationOperationAnnee,
    RealisationOperationAnneeRH,
)
from tests.factories.enjeux import (
    EnjeuFactory, IndicateurFactory, MetriqueFactory, NiveauExigenceFactory,
    ObjectifLongTermeFactory, OperationFactory,
)
from apps.plans.models import CorRolePlan
from tests.factories.plans import PlanGestionFactory
from tests.factories.users import OrganismeFactory, RoleFactory, SuperAdminFactory


def _nomenclature(type_mnemo, mnemo, label):
    t, _ = TypeNomenclature.objects.get_or_create(
        mnemonique=type_mnemo, defaults={'label': type_mnemo})
    n, _ = Nomenclature.objects.get_or_create(
        id_type=t, mnemonique=mnemo,
        defaults={'cd_nomenclature': mnemo[:10], 'label': label})
    return n


@pytest.fixture
def api_client():
    return APIClient()


@pytest.fixture
def plan_finance():
    """Plan 2024-2026 avec une action CS ventilée par organisme + RH + coûts.

    Valeurs contrôlées (année 2024) :
      - RH : 10 jours fonctionnement @ 300 €/j → salarial fonct = 3000 €
      - RH : 2 jours bénévolat
      - Organisme : prestataire fonct 500 €, prestataire invest 200 €,
        autre_cout 100 €, autre_cout_invest 50 €
    """
    plan = PlanGestionFactory(annee_debut=2024, annee_fin=2026)
    org = OrganismeFactory(nom_organisme="Org Alpha")

    enjeu = EnjeuFactory(id_pg=plan)
    ne = NiveauExigenceFactory(id_olt=ObjectifLongTermeFactory(id_enjeu=enjeu))
    ind = IndicateurFactory(id_ne=ne)
    met = MetriqueFactory(id_indicateur=ind)

    cs = _nomenclature('CATEGORIE_ACTION_RESERVE', 'CS', 'Connaissance et suivi')
    op = OperationFactory(
        metriques=[met], id_categorie_action_reserve=cs,
        ventilation_mode='by_org_type', declinaison_par_poste=True,
        code_operation='CS01', annee_min=2024, annee_max=2026,
    )

    fonction, _ = Fonction.objects.get_or_create(
        libelle='Conservateur (export test)', defaults={'type_poste': 'salarie'})
    poste = Poste.objects.create(id_pg=plan, id_organisme=org, nombre=1, cout_jour=300)
    PosteFonction.objects.create(id_poste=poste, id_fonction=fonction, pourcentage=100)

    oa = OperationAnnee.objects.create(id_operation=op, annee=2024, periodicite=True)
    OperationAnneeRH.objects.create(
        id_operation_annee=oa, id_poste=poste, id_organisme=org,
        jours=10, categorie_depense='fonctionnement')
    OperationAnneeRH.objects.create(
        id_operation_annee=oa, id_poste=poste, id_organisme=org,
        jours=2, categorie_depense='benevolat_partenariat')
    OperationAnneeOrganisme.objects.create(
        id_operation_annee=oa, id_organisme=org,
        cout_prestataire=500, cout_prestataire_invest=200,
        autre_cout=100, autre_cout_invest=50)

    # Réalisé (pour les exports « suivi »)
    real = RealisationOperationAnnee.objects.create(
        id_operation_annee=oa, periodicite_realisee=True)
    RealisationOperationAnneeRH.objects.create(
        id_realisation_operation_annee=real, id_poste=poste, id_organisme=org,
        jours=8, categorie_depense='fonctionnement')

    return {'plan': plan, 'org': org, 'op': op, 'poste': poste}


@pytest.fixture
def plan_arbo():
    """Plan avec une arborescence complète (branche NE + branche OO/RA).

    - branche état : OLT → NE → indicateur d'état → métrique multi-blocs (#619)
      → action (sans `code_operation`, pour vérifier le code local du plan) ;
    - branche pression : facteur → pression → OO → RA → indicateur de pression
      → métrique.
    """
    from apps.plans.models_indicateurs import MetriqueScoreBlock
    from tests.factories.enjeux import (
        FacteurInfluenceFactory, IndicateurPressionFactory,
        ObjectifOperationnelFactory, PressionFactory, ResultatAttenduFactory,
    )

    plan = PlanGestionFactory(annee_debut=2024, annee_fin=2026)
    t_etat = _nomenclature('TYPE_INDICATEUR', 'ETAT', 'État')
    t_pression = _nomenclature('TYPE_INDICATEUR', 'PRESSION', 'Pression')
    cs = _nomenclature('CATEGORIE_ACTION_RESERVE', 'CS', 'Connaissance et suivi')

    enjeu = EnjeuFactory(id_pg=plan, etat_enjeu="État actuel de l'enjeu")

    # -- branche « vision à long terme »
    ne = NiveauExigenceFactory(id_olt=ObjectifLongTermeFactory(id_enjeu=enjeu))
    ind = IndicateurFactory(id_ne=ne, type_indicateur=t_etat)
    met = MetriqueFactory(
        id_indicateur=ind, nom_metrique='Recouvrement', unite='%',
        bloc_intitule='Surface', sens_variation='CROISSANT',
        score_1_inf=0, score_1_sup=10, score_5_inf=90, score_5_sup=100,
    )
    bloc2 = MetriqueScoreBlock.objects.create(
        id_metrique=met, position=1, intitule='Hauteur', unite='cm',
        logical_op='AND', score_1_inf=0, score_1_sup=5,
        score_5_inf=50, score_5_sup=100,
    )
    op = OperationFactory(
        metriques=[met], id_categorie_action_reserve=cs,
        code_operation=None, numero_manuel=None,
    )

    # -- branche « stratégie d'action »
    facteur = FacteurInfluenceFactory(id_enjeu=enjeu, libelle='Fréquentation')
    pression = PressionFactory(id_facteur_influence=facteur, libelle='Piétinement')
    oo = ObjectifOperationnelFactory(pressions=[pression])
    ra = ResultatAttenduFactory(id_oo=oo)
    ind_p = IndicateurPressionFactory(id_resultat_attendu=ra, type_indicateur=t_pression)
    met_p = MetriqueFactory(id_indicateur=ind_p, nom_metrique='Sentiers', unite='m')

    return {
        'plan': plan, 'enjeu': enjeu, 'met': met, 'bloc2': bloc2,
        'op': op, 'met_pression': met_p,
    }


# ---------------------------------------------------------------------------
# Logique financière partagée
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestExportFinance:
    def test_cout_salarial_recalcule(self, plan_finance):
        from apps.plans.services_export_finance import build_plan_finance
        pf = build_plan_finance(plan_finance['plan'])
        assert pf.years == [2024, 2025, 2026]
        assert len(pf.actions) == 1
        af = pf.actions[0]
        assert af.is_cs is True
        assert af.categorie == 'CS'
        assert af.is_org_ventilated is True
        oid = plan_finance['org'].id_organisme
        c = af.cell(oid, 2024)
        # 10 jours × 300 €/j
        assert c.sal_fonct == 3000
        assert c.j_fonct == 10
        assert c.j_benevole == 2

    def test_ventilation_fonct_invest_prestataire(self, plan_finance):
        from apps.plans.services_export_finance import build_plan_finance
        af = build_plan_finance(plan_finance['plan']).actions[0]
        c = af.cell(plan_finance['org'].id_organisme, 2024)
        # prestataire présent dans les DEUX blocs (#607 Q2)
        assert c.prest_fonct == 500
        assert c.prest_invest == 200
        assert c.autre_fonct == 100
        assert c.autre_invest == 50
        # totaux
        assert c.tot_fonct == 3000 + 500 + 100     # salarial + presta + autres
        assert c.tot_invest == 0 + 200 + 50
        assert c.tot == c.tot_fonct + c.tot_invest

    def test_realise(self, plan_finance):
        from apps.plans.services_export_finance import build_plan_finance
        af = build_plan_finance(plan_finance['plan']).actions[0]
        c = af.cell(plan_finance['org'].id_organisme, 2024)
        assert c.rj_fonct == 8
        assert c.rsal_fonct == 8 * 300

    def test_mode_none_non_ventile(self, plan_finance):
        """Une action en mode 'none' n'est pas ventilée par organisme (#607 Q3)."""
        from apps.plans.services_export_finance import build_action_finance, poste_entry_factory
        from collections import defaultdict
        op = plan_finance['op']
        op.ventilation_mode = 'none'
        op.save(update_fields=['ventilation_mode'])
        af = build_action_finance(op, {}, defaultdict(poste_entry_factory))
        assert af.is_org_ventilated is False

    def test_mode_by_type_poste_detail_des_couts_sur_lannee(self, plan_finance):
        """#624 — mode « par type de budget + type de poste » : le détail des
        coûts est porté par l'ANNÉE (pas par l'organisme) et alimente les
        totaux fonctionnement / investissement comme la ventilation maximale.
        """
        from collections import defaultdict
        from decimal import Decimal
        from apps.plans.services_export_finance import (
            build_action_finance, poste_entry_factory,
        )
        op = plan_finance['op']
        op.ventilation_mode = 'by_type_poste'
        op.save(update_fields=['ventilation_mode'])
        # Plus de ventilation par organisme : le détail migre sur l'année.
        OperationAnneeOrganisme.objects.filter(
            id_operation_annee__id_operation=op).delete()
        oa = OperationAnnee.objects.get(id_operation=op, annee=2024)
        oa.cout_stage = 200
        oa.cout_prestataire = 500
        oa.autre_cout = 100
        oa.cout_prestataire_invest = 300
        oa.autre_cout_invest = 50
        oa.save()

        af = build_action_finance(op, {}, defaultdict(poste_entry_factory))
        assert af.is_org_ventilated is False
        # Les coûts saisis sur l'année tombent dans la cellule « non ventilé »…
        c = af.cell(0, 2024)
        assert c.prest_fonct == Decimal('500')
        assert c.prest_invest == Decimal('300')
        # stage + autres coûts, l'enveloppe budget_fonctionnement étant vide.
        assert c.autre_fonct == Decimal('300')
        assert c.autre_invest == Decimal('50')
        # …et le total de l'année cumule le salarial calculé des lignes RH
        # (10 j × 300 €), rattaché au poste.
        total = af.year_total(2024)
        assert total.sal_fonct == Decimal('3000')
        assert total.tot_fonct == Decimal('3800')   # 3000 + 500 + 300
        assert total.tot_invest == Decimal('350')   # 300 + 50

    def test_code_action_pg_est_le_code_affichage(self, plan_finance):
        """#618 — la colonne « Code action PG » porte le code calculé (CS1…)."""
        from apps.plans.services_export_finance import build_plan_finance
        af = build_plan_finance(plan_finance['plan']).actions[0]
        assert af.code == 'CS1'

    def test_mode_none_non_ventile(self, plan_finance):
        """Une action en mode 'none' n'est pas ventilée par organisme (#607 Q3)."""
        from apps.plans.services_export_finance import build_action_finance, poste_entry_factory
        from collections import defaultdict
        op = plan_finance['op']
        op.ventilation_mode = 'none'
        op.save(update_fields=['ventilation_mode'])
        af = build_action_finance(op, {}, defaultdict(poste_entry_factory))
        assert af.is_org_ventilated is False

# ---------------------------------------------------------------------------
# Génération des classeurs
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestExportWorkbooks:
    def _load(self, content):
        return load_workbook(io.BytesIO(content))

    def test_fiche_action(self, plan_finance):
        from apps.plans.services_export_fiche_action import build_fiche_action_workbook
        wb = self._load(build_fiche_action_workbook(plan_finance['plan']))
        assert 'CS01' in wb.sheetnames
        ws = wb['CS01']
        labels = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
        # variante CS
        assert any(l == "Indicateur d'état" for l in labels)
        # prévisionnel uniquement (#607 Q4) : pas de ligne « réalisée »
        assert not any('réalisée' in (l or '').lower() for l in labels)
        # prestataire dans les deux blocs (#607 Q2)
        assert any('prestataire — investissement' in (l or '').lower() for l in labels)

    def test_budget_previsionnel(self, plan_finance):
        from apps.plans.services_export_budget_rh import build_budget_previsionnel_workbook
        wb = self._load(build_budget_previsionnel_workbook(plan_finance['plan']))
        assert any('TOTAL par poste de dépense' == s for s in wb.sheetnames)
        # une feuille par organisme gestionnaire
        assert any('Org Alpha' in s for s in wb.sheetnames)

    def test_budget_suivi_a_colonnes_prevu_realise(self, plan_finance):
        from apps.plans.services_export_budget_rh import build_budget_suivi_workbook
        wb = self._load(build_budget_suivi_workbook(plan_finance['plan']))
        ws = next(wb[s] for s in wb.sheetnames if 'Org Alpha' in s)
        row4 = [ws.cell(4, c).value for c in range(1, ws.max_column + 1)]
        assert 'Prévu' in row4 and 'Réalisé' in row4

    def test_rh_previsionnel(self, plan_finance):
        from apps.plans.services_export_budget_rh import build_rh_previsionnel_workbook
        wb = self._load(build_rh_previsionnel_workbook(plan_finance['plan']))
        assert 'RH par type de poste' in wb.sheetnames
        # 10 jours prévus en 2024 pour l'action, colonne TOTAL = 10
        ws = next(wb[s] for s in wb.sheetnames if 'Org Alpha' in s)
        data_row = [ws.cell(4, c).value for c in range(1, ws.max_column + 1)]
        # #618 — colonne « Code action PG » = code local du plan (et non le
        # champ libre `code_operation`)
        assert 'CS1' in data_row

    def test_budget_suivi_totaux_fonct_invest(self, plan_finance):
        """#618 — sous-totaux Fonctionnement / Investissement sur la feuille TOTAL."""
        from apps.plans.services_export_budget_rh import build_budget_suivi_workbook
        wb = self._load(build_budget_suivi_workbook(plan_finance['plan']))
        ws = wb['Total par type de dépense']
        labels = {ws.cell(r, 1).value: r for r in range(1, ws.max_row + 1)}
        assert 'TOTAL Fonctionnement' in labels
        assert 'TOTAL Investissement' in labels
        # colonnes TOTAL (prévu / réalisé) en fin de feuille
        c_prev, c_real = ws.max_column - 1, ws.max_column
        # fonctionnement 2024 : 3000 (salarial) + 500 (presta) + 100 (autres)
        assert ws.cell(labels['TOTAL Fonctionnement'], c_prev).value == '3 600'
        # réalisé : 8 jours × 300 €/j
        assert ws.cell(labels['TOTAL Fonctionnement'], c_real).value == '2 400'
        # investissement 2024 : 200 (presta) + 50 (autres)
        assert ws.cell(labels['TOTAL Investissement'], c_prev).value == '250'
        assert ws.cell(labels['TOTAL Investissement'], c_real).value == '0'
        # cohérence avec le TOTAL général
        assert ws.cell(labels['TOTAL'], c_prev).value == '3 850'


    def test_budget_suivi_sous_totaux_par_categorie(self, plan_finance):
        """#618 — les en-têtes de catégorie portent leur sous-total (plus de ligne vide)."""
        from apps.plans.services_export_budget_rh import build_budget_suivi_workbook
        wb = self._load(build_budget_suivi_workbook(plan_finance['plan']))
        ws = wb['Total par type de dépense']
        labels = {ws.cell(r, 1).value: r for r in range(1, ws.max_row + 1)}
        c_prev = ws.max_column - 1
        # Fonctionnement — Coût salarial 2024 = 10 j × 300 = 3000
        assert ws.cell(labels['► Fonctionnement — Coût salarial'], c_prev).value == '3 000'
        # Fonctionnement — autres coûts = 500 (presta) + 100 (autres) = 600
        assert ws.cell(labels['► Fonctionnement — autres coûts'], c_prev).value == '600'
        # Investissement — autres coûts = 200 (presta) + 50 (autres) = 250
        assert ws.cell(labels['► Investissement — autres coûts'], c_prev).value == '250'

    def test_rh_suivi(self, plan_finance):
        from apps.plans.services_export_budget_rh import build_rh_suivi_workbook
        wb = self._load(build_rh_suivi_workbook(plan_finance['plan']))
        assert 'Total par poste' in wb.sheetnames

    def test_word_et_arborescence(self, plan_finance):
        from apps.plans.services_export_word import build_plan_docx
        from apps.plans.services_export_arbo import build_presentation_workbook
        docx = build_plan_docx(plan_finance['plan'])
        assert docx[:2] == b'PK'      # conteneur zip Office
        wb = self._load(build_presentation_workbook(plan_finance['plan']))
        assert len(wb.sheetnames) >= 1

    def test_word_prefixe_numero_enjeu_et_olt(self, plan_arbo):
        """#628 — intitulés préfixés « Enjeu N : … » et « OLT N : … »."""
        from docx import Document
        from apps.plans.services_export_word import build_plan_docx
        doc = Document(io.BytesIO(build_plan_docx(plan_arbo['plan'])))
        texts = [p.text for p in doc.paragraphs]
        assert any(t.startswith('Enjeu 1 : ') for t in texts), texts
        assert any(t.startswith('OLT 1 : ') for t in texts), texts


# ---------------------------------------------------------------------------
# Export arborescence « de présentation » (#619 / #620)
# ---------------------------------------------------------------------------

def _find_row(ws, label, col=1):
    """Numéro de la première ligne dont la cellule `col` vaut `label`."""
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, col).value == label:
            return r
    return None


def _row_values(ws, row):
    return [ws.cell(row, c).value for c in range(1, ws.max_column + 1)]


def _fill(ws, row, col):
    return ws.cell(row, col).fill.fgColor.rgb


@pytest.mark.django_db
class TestExportArborescencePresentation:
    """#619 / #620 — mise en forme et contenu du classeur d'arborescence."""

    def _sheet(self, plan):
        from apps.plans.services_export_arbo import build_presentation_workbook
        wb = load_workbook(io.BytesIO(build_presentation_workbook(plan)))
        return wb[wb.sheetnames[0]]

    # ---- #620 : colonnes et couleurs --------------------------------------

    def test_pas_de_colonne_niveau_exigence_en_face_des_facteurs(self, plan_arbo):
        """Le bloc bas commence par « Facteurs d'influence » (#620)."""
        ws = self._sheet(plan_arbo['plan'])
        b_h2 = _find_row(ws, "Facteurs d'influence")
        assert b_h2 is not None, "en-tête « Facteurs d'influence » absent"
        assert ws.cell(b_h2, 1).value == "Facteurs d'influence"
        assert ws.cell(b_h2, 2).value == "Pressions à gérer"
        # « Niveau d'exigence » ne subsiste que dans le bloc haut
        niveaux = [
            (r, c)
            for r in range(1, ws.max_row + 1)
            for c in range(1, 12)
            if (ws.cell(r, c).value or '') == "Niveau d'exigence"
        ]
        assert niveaux == [], f"colonne « Niveau d'exigence » résiduelle : {niveaux}"

    def test_couleurs_du_modele(self, plan_arbo):
        """Bandeaux bleu foncé, données « stratégie » en orange (#620)."""
        from apps.plans.services_export_arbo import (
            _C_ENJEU, _C_ETAT_DATA, _C_INFLU, _C_STRAT_DATA,
        )
        ws = self._sheet(plan_arbo['plan'])
        # bandeau « Influences sur l'enjeu » = même bleu foncé que « Enjeu »
        r_enjeu = _find_row(ws, "Enjeu")
        r_influ = _find_row(ws, "Influences sur l'enjeu")
        assert _fill(ws, r_enjeu, 1) == _C_ENJEU == _C_INFLU
        assert _fill(ws, r_influ, 1) == _C_INFLU
        # données facteur / pression : bleu clair (bloc « influences »), pas orange
        b_data = _find_row(ws, "Facteurs d'influence") + 1
        assert _fill(ws, b_data, 1) == _C_ETAT_DATA
        assert _fill(ws, b_data, 2) == _C_ETAT_DATA   # ancre de la fusion B:C
        # données « stratégie d'action » (objectifs opérationnels et suivantes) :
        # orange du modèle, plus rouge/rose
        assert _fill(ws, b_data, 4) == _C_STRAT_DATA
        assert _C_STRAT_DATA == 'FFFCD5B5', _C_STRAT_DATA

    # ---- #619 : grille de lecture et code action --------------------------

    def test_grille_a_une_colonne_unite(self, plan_arbo):
        """La grille de lecture porte une colonne « Unité » (#619)."""
        from apps.plans.services_export_arbo import GR_MET, GR_UNITE
        ws = self._sheet(plan_arbo['plan'])
        r = _find_row(ws, "Métriques", col=GR_MET)
        assert r is not None, "sous-en-tête « Métriques » de la grille absent"
        assert ws.cell(r, GR_UNITE).value == "Unité"
        # l'unité de la métrique apparaît en face de sa ligne
        unites = [ws.cell(rr, GR_UNITE).value
                  for rr in range(r + 1, r + 4)]
        assert '%' in unites, unites

    def test_grille_multibloc_dans_une_seule_case(self, plan_arbo):
        """Une métrique multi-blocs tient sur UNE ligne, blocs combinés par
        colonne de score dans la même cellule (#619)."""
        from apps.plans.services_export_arbo import GR_MET, GR_UNITE, GR_S1
        ws = self._sheet(plan_arbo['plan'])
        r = _find_row(ws, "Métriques", col=GR_MET)
        row = r + 1
        # une seule ligne de données pour la métrique
        assert ws.cell(row, GR_MET).value == 'Recouvrement'
        assert ws.cell(row, GR_UNITE).value == '%'
        # le bloc complémentaire ne crée plus de ligne « Hauteur » séparée
        assert 'Hauteur' not in (ws.cell(row + 1, GR_MET).value or '')
        # colonne score 1 : les deux blocs dans la MÊME cellule, sur deux lignes
        cell = ws.cell(row, GR_S1).value or ''
        assert 'surface' in cell.lower(), cell
        assert 'hauteur' in cell.lower(), cell
        assert '\n' in cell, cell

    def test_grille_seuils_sans_zeros_superflus(self, plan_arbo):
        """Les seuils reprennent les décimales saisies : « 10 », pas « 10.0000 » (#619)."""
        from apps.plans.services_export_arbo import GR_MET, GR_S1
        ws = self._sheet(plan_arbo['plan'])
        r = _find_row(ws, "Métriques", col=GR_MET)
        cell = ws.cell(r + 1, GR_S1).value or ''
        assert '.0000' not in cell, cell
        assert '10' in cell, cell   # score_1_sup=10 du bloc principal

    def test_grille_note_les_bornes_incluses_et_exclues(self, plan_arbo):
        """Les paliers gardent la notation par crochets de la saisie (#619).

        « 0 – 10 » ne disait pas de quel côté la borne appartenait : l'export
        doit écrire « [0 ; 10] » ou « [0 ; 10[ » comme l'écran de saisie, dans
        le bloc principal COMME dans les blocs complémentaires.
        """
        from apps.plans.models_indicateurs import Metrique
        from apps.plans.services_export_arbo import GR_MET, GR_S1, GR_S5
        met = plan_arbo['met']
        # borne haute du palier 1 exclue : la valeur 10 appartient au palier 2
        Metrique.objects.filter(pk=met.pk).update(score_1_sup_inclusive=False)
        met.score_blocks.update(score_1_sup_inclusive=False)

        ws = self._sheet(plan_arbo['plan'])
        r = _find_row(ws, "Métriques", col=GR_MET)
        s1 = ws.cell(r + 1, GR_S1).value or ''
        assert '[0 ; 10[' in s1, s1          # bloc principal (Surface)
        assert '[0 ; 5[' in s1, s1           # bloc complémentaire (Hauteur)
        assert '–' not in s1, s1             # plus d'ancienne notation « 0 – 10 »

        # Palier 5 : sa borne basse dépend du palier 4, pas de lui-même. Le
        # palier 4 étant inclusif par défaut, 90 lui appartient et le palier 5
        # démarre donc exclu.
        s5 = ws.cell(r + 1, GR_S5).value or ''
        assert ']90 ; 100]' in s5, s5

        # Et si le palier 4 devient exclusif, 90 bascule dans le palier 5.
        Metrique.objects.filter(pk=met.pk).update(score_4_sup_inclusive=False)
        ws = self._sheet(plan_arbo['plan'])
        r = _find_row(ws, "Métriques", col=GR_MET)
        assert '[90 ; 100]' in (ws.cell(r + 1, GR_S5).value or '')

    def test_grille_multibloc_reste_sur_une_ligne(self, plan_arbo):
        """Ajouter des blocs n'ajoute plus de lignes à la grille (#619) : ils sont
        empilés dans la même cellule, pas en lignes séparées."""
        from apps.plans.models_indicateurs import MetriqueScoreBlock
        from apps.plans.services_export_arbo import GR_MET, GR_S1
        met = plan_arbo['met']
        for pos in range(2, 6):   # 4 blocs de plus
            MetriqueScoreBlock.objects.create(
                id_metrique=met, position=pos, intitule=f'Bloc {pos}',
                logical_op='OR', score_1_inf=0, score_1_sup=pos)
        ws = self._sheet(plan_arbo['plan'])
        r = _find_row(ws, "Métriques", col=GR_MET)
        # toujours une seule ligne de données pour la métrique
        assert ws.cell(r + 1, GR_MET).value == 'Recouvrement'
        assert (ws.cell(r + 2, GR_MET).value or '') == ''
        # les 5 blocs empilés dans la cellule de score 1 (principal + 4 « Bloc N »)
        cell = ws.cell(r + 1, GR_S1).value or ''
        assert cell.count('\n') >= 4, cell

    def test_code_action_est_le_code_local_du_plan(self, plan_arbo):
        """La colonne « Code » porte le code calculé du plan (CS1…) — #619.

        L'action de la fixture n'a ni `code_operation` ni `numero_manuel` :
        la colonne restait vide.
        """
        ws = self._sheet(plan_arbo['plan'])
        h2 = _find_row(ws, "Code", col=8)
        codes = [ws.cell(r, 8).value for r in range(h2 + 1, h2 + 4)]
        assert 'CS1' in codes, codes



# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@pytest.mark.django_db
class TestExportEndpoints:
    ENDPOINTS = [
        'export-arborescence-presentation-xlsx',
        'export-fiches-actions-xlsx',
        'export-plan-docx',
        'export-rh-previsionnel-xlsx',
        'export-rh-suivi-xlsx',
        'export-budget-previsionnel-xlsx',
        'export-budget-suivi-xlsx',
    ]

    def test_endpoints_return_file(self, api_client, plan_finance):
        api_client.force_authenticate(user=SuperAdminFactory())
        plan = plan_finance['plan']
        for ep in self.ENDPOINTS:
            resp = api_client.get(f'/api/plans/plans/{plan.id_pg}/{ep}/')
            assert resp.status_code == 200, ep
            assert resp['Content-Type'].startswith('application/vnd.openxmlformats'), ep
            content = b''.join(resp.streaming_content) if resp.streaming else resp.content
            assert content[:2] == b'PK', ep

    def test_endpoints_require_auth(self, api_client, plan_finance):
        plan = plan_finance['plan']
        resp = api_client.get(f'/api/plans/plans/{plan.id_pg}/export-budget-previsionnel-xlsx/')
        assert resp.status_code in (401, 403)

    def test_membre_non_referent_ne_peut_pas_exporter(self, api_client, plan_finance):
        """Un utilisateur simplement lié au plan (membre) consulte mais n'exporte pas."""
        plan = plan_finance['plan']
        membre = RoleFactory()
        CorRolePlan.objects.create(id_role=membre, plan_de_gestion=plan, referent=False)
        api_client.force_authenticate(user=membre)

        # Il voit bien le plan (lecture seule) …
        assert api_client.get(f'/api/plans/plans/{plan.id_pg}/').status_code == 200
        # … mais tous les exports lui sont refusés.
        for ep in self.ENDPOINTS + ['export-arborescence-xlsx', 'export-actions-xlsx']:
            resp = api_client.get(f'/api/plans/plans/{plan.id_pg}/{ep}/')
            assert resp.status_code == 403, ep

    def test_referent_du_plan_peut_exporter(self, api_client, plan_finance):
        plan = plan_finance['plan']
        referent = RoleFactory()
        plan.referents.add(referent)
        api_client.force_authenticate(user=referent)

        for ep in self.ENDPOINTS:
            resp = api_client.get(f'/api/plans/plans/{plan.id_pg}/{ep}/')
            assert resp.status_code == 200, ep


# ---------------------------------------------------------------------------
# Tableau de bord (#638) — mise en forme du tableau affiché
# ---------------------------------------------------------------------------

def _payload_tableau_bord():
    """Deux lignes minimales : un indicateur scoré et sa métrique."""
    return {
        'titre': 'Tableau de bord — Plan test',
        'meta': [['Onglet', 'État'], ['Recherche', 'Balbuzard']],
        'entetes': ['Enjeu', 'Objectif', 'Niveau', 'Indicateur', 'Métrique',
                    '2024', 'Évaluation globale', 'Actions'],
        'lignes': [
            {'type': 'indicateur', 'cellules': [
                'Enjeu 1', 'OLT 1 : Objectif A', 'NE 1', 'Indicateur 1', '',
                {'t': 'Bon', 's': 'good'}, {'t': 'Très mauvais', 's': 'very-bad'}, 'CS01',
            ]},
            {'type': 'metrique', 'cellules': [
                'Enjeu 1', 'OLT 1 : Objectif A', 'NE 1', 'Indicateur 1', 'Surface (ha)',
                {'t': 'Moyen', 's': 'neutral'}, '', '',
            ]},
        ],
    }


@pytest.mark.django_db
class TestExportTableauDeBord:
    """
    Retour de recette : le tableau sortait en CSV, sans mise en forme. Le
    classeur doit reprendre les couleurs de l'interface — en-tête à la couleur
    de l'instance et cases de score à la palette du design system.
    """

    URL = '/api/plans/plans/{}/export-tableau-de-bord-xlsx/'

    def _workbook(self, api_client, plan, payload=None):
        resp = api_client.post(
            self.URL.format(plan.id_pg), payload or _payload_tableau_bord(), format='json')
        assert resp.status_code == 200
        assert resp['Content-Type'].startswith(
            'application/vnd.openxmlformats-officedocument')
        return load_workbook(io.BytesIO(resp.content))

    def test_colore_les_cases_de_score_avec_la_palette(self, api_client, plan_finance):
        api_client.force_authenticate(user=SuperAdminFactory())
        ws = self._workbook(api_client, plan_finance['plan']).active

        entete = next(
            r for r in ws.iter_rows() if r[0].value == 'Enjeu'
        )
        ligne_indic = ws[entete[0].row + 1]
        ligne_metrique = ws[entete[0].row + 2]

        # 6e colonne = année 2024, 7e = évaluation globale.
        assert ligne_indic[5].value == 'Bon'
        assert ligne_indic[5].fill.fgColor.rgb == 'FF82DB8A'        # good
        assert ligne_indic[6].fill.fgColor.rgb == 'FFFF7579'        # very-bad
        assert ligne_metrique[5].fill.fgColor.rgb == 'FFF7D35C'     # neutral

    def test_entete_a_la_couleur_de_l_instance(self, api_client, plan_finance):
        api_client.force_authenticate(user=SuperAdminFactory())
        ws = self._workbook(api_client, plan_finance['plan']).active

        entete = next(r for r in ws.iter_rows() if r[0].value == 'Enjeu')
        assert entete[0].fill.fgColor.rgb == 'FF025359'
        assert entete[0].font.color.rgb == 'FFFFFFFF'
        assert entete[0].font.bold

    def test_reprend_le_titre_et_le_rappel_des_filtres(self, api_client, plan_finance):
        api_client.force_authenticate(user=SuperAdminFactory())
        ws = self._workbook(api_client, plan_finance['plan']).active

        valeurs = [c.value for col in ws.iter_cols(min_col=1, max_col=2) for c in col]
        assert 'Tableau de bord — Plan test' in valeurs
        assert 'Balbuzard' in valeurs

    def test_case_sans_score_reste_incolore(self, api_client, plan_finance):
        api_client.force_authenticate(user=SuperAdminFactory())
        ws = self._workbook(api_client, plan_finance['plan']).active

        entete = next(r for r in ws.iter_rows() if r[0].value == 'Enjeu')
        # Évaluation globale de la ligne métrique : vide, donc pas d'aplat de score.
        case = ws[entete[0].row + 2][6]
        assert case.value in (None, '')
        assert case.fill.fgColor.rgb != 'FF82DB8A'

    def test_exportable_meme_quand_le_plan_est_valide(self, api_client, plan_finance):
        """
        Le POST ne modifie pas le plan : le verrou « hors brouillon » (#248) ne
        doit pas l'intercepter, sinon aucun plan validé ne serait exportable —
        or c'est justement l'état dans lequel on lit un tableau de bord.
        """
        plan = plan_finance['plan']
        plan.statut = 'valide'
        plan.save(update_fields=['statut'])
        api_client.force_authenticate(user=SuperAdminFactory())

        resp = api_client.post(
            self.URL.format(plan.id_pg), _payload_tableau_bord(), format='json')
        assert resp.status_code == 200

    def test_membre_non_referent_ne_peut_pas_exporter(self, api_client, plan_finance):
        plan = plan_finance['plan']
        membre = RoleFactory()
        CorRolePlan.objects.create(id_role=membre, plan_de_gestion=plan, referent=False)
        api_client.force_authenticate(user=membre)

        resp = api_client.post(
            self.URL.format(plan.id_pg), _payload_tableau_bord(), format='json')
        assert resp.status_code == 403
