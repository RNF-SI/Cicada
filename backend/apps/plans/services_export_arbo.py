"""
Export Excel « de présentation » de l'arborescence d'un plan de gestion.

Distinct du format d'import round-trip (``services_import.py``) : ce module
produit un classeur **lisible / imprimable**, calqué sur le modèle
« Modèle_Export_CICADA_Arborescence » fourni par l'équipe :

- **un onglet par enjeu et par FCR** ;
- un bloc haut « Vision à long terme » (état de conservation) :
  OLT → niveau d'exigence → indicateurs d'état → métriques → actions ;
- un bloc bas « Stratégie d'action » (pressions) :
  facteurs d'influence → pressions → OO → résultats attendus →
  indicateurs de pression → métriques → actions ;
- une « grille de lecture des métriques » (scores 1..5) à droite de chaque bloc.

Le point d'entrée public est :func:`build_presentation_workbook`.
"""

from __future__ import annotations

import io
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Couleurs — résolues en ARGB concrets pour être indépendantes du thème du
# classeur (le modèle d'origine s'appuyait sur des couleurs de thème Office).
# Bases issues du thème du modèle : accent1 bleu, accent2 rouge, accent3 vert,
# lt2 beige. Les teintes (tint) sont appliquées par le formule OOXML linéaire.
# ---------------------------------------------------------------------------

_THEME = {
    "blue": (0x4F, 0x81, 0xBD),   # accent1 — colonne Enjeu / influences
    "red": (0xC0, 0x50, 0x4D),    # accent2 — bandeau stratégie d'action
    "green": (0x9B, 0xBB, 0x59),  # accent3 — bandeau vision long terme
    "orange": (0xF7, 0x96, 0x46), # accent6 — corps du bloc stratégie d'action
    "beige": (0xEE, 0xEC, 0xE1),  # lt2 — bandeau grille de lecture
}


def _tint(base: tuple[int, int, int], tint: float) -> str:
    """Applique une teinte OOXML (approximation linéaire par canal) → 'FFrrggbb'."""
    out = []
    for c in base:
        if tint < 0:
            v = c * (1 + tint)
        else:
            v = c * (1 - tint) + 255 * tint
        out.append(max(0, min(255, int(round(v)))))
    return "FF{:02X}{:02X}{:02X}".format(*out)


# Palette dérivée du modèle (#620 — bandeaux « Enjeu » et « Influences sur
# l'enjeu » au même bleu foncé, corps du bloc « Stratégie d'action » en orange).
_C_ENJEU = _tint(_THEME["blue"], -0.5)        # A6 — en-tête Enjeu
_C_ENJEU_DATA = _tint(_THEME["blue"], 0.6)    # cellule nom d'enjeu
_C_ETAT_DATA = _tint(_THEME["blue"], 0.8)     # état actuel / facteurs & pressions
_C_VLT = _tint(_THEME["green"], -0.25)        # D6 — VISION A LONG TERME
_C_VLT_HDR = _tint(_THEME["green"], 0.4)      # sous-en-têtes bloc haut
_C_VLT_DATA = _tint(_THEME["green"], 0.6)     # cellules données bloc haut
_C_STRAT = _tint(_THEME["red"], -0.5)         # D19 — STRATEGIE D'ACTION
_C_STRAT_HDR = _tint(_THEME["orange"], 0.4)   # sous-en-têtes bloc bas
_C_STRAT_DATA = _tint(_THEME["orange"], 0.6)  # cellules données bloc bas
_C_INFLU = _C_ENJEU                           # A19 — Influences sur l'enjeu
_C_INFLU_HDR = _tint(_THEME["blue"], 0.4)     # sous-en-têtes bleus (haut & bas)
_C_GRILLE = _tint(_THEME["beige"], 0.0)       # bandeau grille de lecture
_WHITE = "FFFFFFFF"

# Couleurs des scores (palette du modèle, indexées héritées Excel)
_SCORE_FILLS = {
    "indet": "FFC0C0C0",  # Indéterminé — gris
    1: "FFFF0000",        # très mauvais — rouge
    2: "FFFFCC00",        # mauvais — ambre
    3: "FFFFFF00",        # moyen — jaune
    4: "FF99CC00",        # bon — vert
    5: "FF00CCFF",        # très bon — cyan
}

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

_FONT_TITLE = Font(name="Calibri", bold=True, size=14, color="FF025359")
_FONT_FLAG = Font(name="Calibri", bold=True, size=16, color="FFB74D5D")
_FONT_HDR = Font(name="Calibri", bold=True, size=11, color="FF000000")
_FONT_HDR_LIGHT = Font(name="Calibri", bold=True, size=11, color=_WHITE)
_FONT_DATA = Font(name="Calibri", size=10, color="FF000000")
_FONT_ACTION = Font(name="Calibri", size=10, color="FF000000")
_FONT_CODE = Font(name="Calibri", bold=True, size=10, color="FF025359")

_thin = Side(style="thin", color="FF9A8F86")
_med = Side(style="medium", color="FF746F6E")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_BORDER_MED = Border(left=_med, right=_med, top=_med, bottom=_med)

_AL_CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
_AL_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
_AL_CTR_TOP = Alignment(horizontal="center", vertical="top", wrap_text=True)

# Largeurs de colonnes (issues du modèle)
_COL_WIDTHS = {
    1: 22.5, 2: 18.5, 3: 20.0, 4: 34.0, 5: 30.0, 6: 27.5, 7: 38.0,
    8: 10.0, 9: 46.0, 10: 30.0, 11: 12.5, 12: 4.0, 13: 4.0,
    14: 26.0, 15: 10.0, 16: 18.0, 17: 18.0, 18: 18.0, 19: 18.0, 20: 18.0,
    21: 20.0,
}

# Indices de colonnes (1-based)
COL_A, COL_B, COL_C, COL_D, COL_E = 1, 2, 3, 4, 5
COL_F, COL_G, COL_H, COL_I, COL_J, COL_K = 6, 7, 8, 9, 10, 11
# Grille de lecture — #619 : colonne « Unité » intercalée après « Métriques »
GR_MET, GR_UNITE, GR_INDET = 14, 15, 16
GR_S1, GR_S2, GR_S3, GR_S4, GR_S5 = 17, 18, 19, 20, 21


def _set(ws, row, col, value, *, fill=None, font=None, align=None, border=_BORDER):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = font or _FONT_DATA
    cell.alignment = align or _AL_LEFT
    cell.border = border
    return cell


def _merge(ws, r1, c1, r2, c2, value=None, **kw):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = _set(ws, r1, c1, value, **kw)
    # borde toutes les cellules de la plage fusionnée
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = kw.get("border", _BORDER)
    return cell


def _sanitize_sheet_title(title: str, used: set) -> str:
    """Titre d'onglet valide Excel (<=31 car., pas de []:*?/\\), unique."""
    for ch in '[]:*?/\\':
        title = title.replace(ch, " ")
    title = title.strip() or "Feuille"
    title = title[:31]
    base = title
    i = 2
    while title.lower() in used:
        suffix = f" ({i})"
        title = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(title.lower())
    return title


# ---------------------------------------------------------------------------
# Extraction des données
# ---------------------------------------------------------------------------

def _txt(value) -> str:
    return (value or "").strip() if isinstance(value, str) else (value or "")


def _ind_type(ind) -> str:
    t = getattr(ind, "type_indicateur", None)
    return (getattr(t, "mnemonique", "") or "").upper()


def _score_labels(block) -> list[str]:
    """Retourne [s1..s5] pour la grille de lecture (label sinon plage de seuils).

    ``block`` est indifféremment une :class:`Metrique` (bloc principal) ou un
    :class:`MetriqueScoreBlock` (bloc complémentaire) : les deux exposent les
    mêmes champs ``score_N_inf/sup`` et ``inactive_levels`` (#619).
    """
    inactifs = set(getattr(block, "inactive_levels", None) or [])
    labels = []
    for i in range(1, 6):
        if i in inactifs:
            labels.append("")
            continue
        lbl = _txt(getattr(block, f"score_{i}_label", None))
        if not lbl:
            inf = getattr(block, f"score_{i}_inf", None)
            sup = getattr(block, f"score_{i}_sup", None)
            if inf is not None or sup is not None:
                lbl = f"{'' if inf is None else inf} – {'' if sup is None else sup}".strip(" –")
        labels.append(lbl or "")
    return labels


def _grille_lignes(met) -> list[tuple[str, str, list[str]]]:
    """Lignes de la grille de lecture d'une métrique : (libellé, unité, scores).

    #619 — une métrique multi-blocs (#247) produit une ligne PAR bloc : le bloc
    principal porté par la métrique, puis chaque bloc complémentaire avec son
    opérateur logique (ET / OU) et sa propre unité.
    """
    principal = _txt(met.nom_metrique)
    intitule = _txt(getattr(met, "bloc_intitule", ""))
    lignes = [(
        f"{principal} — {intitule}" if intitule else principal,
        _txt(getattr(met, "unite", "")),
        _score_labels(met),
    )]
    for bloc in met.score_blocks.all():
        op = "ET" if bloc.logical_op == "AND" else "OU"
        libelle = _txt(bloc.intitule) or f"bloc {bloc.position + 1}"
        lignes.append((
            f"    {op} — {libelle}",
            _txt(bloc.unite),
            _score_labels(bloc),
        ))
    return lignes


def _priorite(op) -> str:
    p = getattr(op, "id_priorite", None)
    if not p:
        return ""
    # « Priorité 1 » → « 1 »
    lbl = _txt(getattr(p, "label", "")) or ""
    return lbl.replace("Priorité", "").strip() or lbl


def _actions_for(metrique=None, indicateur=None):
    """Actions (opérations) rattachées à une métrique et/ou directement à l'indicateur."""
    ops = []
    if metrique is not None:
        ops.extend(list(metrique.operations.all()))
    if indicateur is not None:
        ops.extend(list(indicateur.operations.all()))
    # dédoublonne en conservant l'ordre
    seen, out = set(), []
    for o in ops:
        if o.id_operation in seen:
            continue
        seen.add(o.id_operation)
        out.append(o)
    return out


def _code_action(op, codes) -> str:
    """Code affiché d'une action — #619.

    Le « code action » du plan est le code **calculé** (CS1, IP2…), celui que
    l'application affiche partout ; `code_operation` est un champ libre quasi
    toujours vide, d'où la colonne « Code » vide dans les exports.
    Même convention que ``services_export_finance`` (#618).
    """
    return (
        _txt(codes.get(op.id_operation))
        or _txt(op.code_operation)
        or (f"n°{op.numero_manuel}" if op.numero_manuel else "")
    )


def _leaf_rows_for_indicateurs(indicateurs, group_prefix, codes):
    """Génère les lignes feuilles (indicateur → métrique → action) d'un groupe.

    ``group_prefix`` : tuple identifiant le parent (olt/ne ou facteur/pression/oo/ra)
    utilisé pour calculer les fusions verticales sans coller des valeurs
    identiques de parents distincts.
    ``codes`` : {id_operation: code local du plan} (#619).
    Retourne une liste de dicts {path, ind, met, code, action, priorite}.
    """
    rows = []
    metriques_grille = []  # (ordre) pour la grille de lecture
    for ind in indicateurs:
        mets = list(ind.metriques.all())
        if not mets:
            rows.append({
                "path": group_prefix + (("i", ind.id_indicateur),),
                "ind": ind.nom_indicateur, "met": "", "code": "",
                "action": "", "priorite": "",
            })
            continue
        for met in mets:
            metriques_grille.append(met)
            actions = _actions_for(metrique=met)
            base = group_prefix + (("i", ind.id_indicateur), ("m", met.id_metrique))
            if not actions:
                rows.append({
                    "path": base, "ind": ind.nom_indicateur,
                    "met": met.nom_metrique, "code": "", "action": "",
                    "priorite": "",
                })
                continue
            for op in actions:
                rows.append({
                    "path": base,
                    "ind": ind.nom_indicateur, "met": met.nom_metrique,
                    "code": _code_action(op, codes),
                    "action": op.libelle, "priorite": _priorite(op),
                })
        # actions rattachées directement à l'indicateur (sans métrique)
        for op in _actions_for(indicateur=ind):
            rows.append({
                "path": group_prefix + (("i", ind.id_indicateur), ("m", 0)),
                "ind": ind.nom_indicateur, "met": "",
                "code": _code_action(op, codes),
                "action": op.libelle, "priorite": _priorite(op),
            })
    return rows, metriques_grille


def _split_indicateurs(indicateurs):
    """(indicateurs non-réponse, indicateurs réponse)."""
    non_rep = [i for i in indicateurs if _ind_type(i) != "REPONSE"]
    rep = [i for i in indicateurs if _ind_type(i) == "REPONSE"]
    return non_rep, rep


def _collect_top(enjeu, codes):
    """Bloc haut (vision long terme / état). Retourne (rows, grille)."""
    rows = []
    grille = []
    for olt in enjeu.objectifs_long_terme.all():
        for ne in olt.niveaux_exigence.all():
            inds = list(ne.indicateurs.all())
            non_rep, rep = _split_indicateurs(inds)
            reponse = " ; ".join(i.nom_indicateur for i in rep)
            prefix = (("o", olt.id_olt), ("n", ne.id_ne))
            leaf, mets = _leaf_rows_for_indicateurs(non_rep, prefix, codes)
            if not leaf:
                leaf = [{
                    "path": prefix, "ind": "", "met": "", "code": "",
                    "action": "", "priorite": "",
                }]
            for r in leaf:
                r["olt"] = olt.libelle
                r["ne"] = ne.libelle
                r["reponse"] = reponse
                r["ne_key"] = prefix
            rows.extend(leaf)
            grille.extend(mets)
    return rows, grille


def _collect_bottom(enjeu, codes):
    """Bloc bas (stratégie d'action / pressions). Retourne (rows, grille)."""
    from .models_enjeux import ObjectifOperationnel

    rows = []
    grille = []
    seen_oo = set()

    def emit_oo(oo, facteur_lbl, pression_lbl, keyprefix):
        for ra in oo.resultats_attendus.all():
            inds = list(ra.indicateurs.all())
            non_rep, rep = _split_indicateurs(inds)
            reponse = " ; ".join(i.nom_indicateur for i in rep)
            prefix = keyprefix + (("oo", oo.id_oo), ("ra", ra.id_ra))
            leaf, mets = _leaf_rows_for_indicateurs(non_rep, prefix, codes)
            if not leaf:
                leaf = [{
                    "path": prefix, "ind": "", "met": "", "code": "",
                    "action": "", "priorite": "",
                }]
            for r in leaf:
                r["facteur"] = facteur_lbl
                r["pression"] = pression_lbl
                r["oo"] = oo.libelle
                r["ra"] = ra.libelle
                r["reponse"] = reponse
            rows.extend(leaf)
            grille.extend(mets)
        if not oo.resultats_attendus.all():
            prefix = keyprefix + (("oo", oo.id_oo),)
            rows.append({
                "path": prefix, "facteur": facteur_lbl, "pression": pression_lbl,
                "oo": oo.libelle, "ra": "", "ind": "", "met": "", "code": "",
                "action": "", "priorite": "", "reponse": "",
            })

    # Via facteurs → pressions → OO
    for facteur in enjeu.facteurs_influence.all():
        for pression in facteur.pressions.all():
            oos = list(pression.objectifs_operationnels.all())
            if not oos:
                rows.append({
                    "path": (("f", facteur.id_facteur_influence), ("p", pression.id_pression)),
                    "facteur": facteur.libelle, "pression": pression.libelle,
                    "oo": "", "ra": "", "ind": "", "met": "", "code": "",
                    "action": "", "priorite": "", "reponse": "",
                })
                continue
            for oo in oos:
                seen_oo.add(oo.id_oo)
                emit_oo(oo, facteur.libelle, pression.libelle,
                        (("f", facteur.id_facteur_influence), ("p", pression.id_pression)))

    # OO rattachés directement à l'enjeu (cas FCR notamment)
    direct = ObjectifOperationnel.objects.filter(id_enjeu=enjeu).exclude(
        id_oo__in=seen_oo
    ).prefetch_related(
        "resultats_attendus__indicateurs__metriques",
        "resultats_attendus__indicateurs__type_indicateur",
    )
    for oo in direct:
        if oo.id_oo in seen_oo:
            continue
        seen_oo.add(oo.id_oo)
        emit_oo(oo, "", "", (("oo", oo.id_oo),))

    return rows, grille


# ---------------------------------------------------------------------------
# Rendu d'un onglet
# ---------------------------------------------------------------------------

def _merge_runs(ws, col, start_row, keys, values, *, fill, font=None, col_end=None):
    """Fusionne verticalement les lignes consécutives de même clé et écrit la valeur.

    ``keys`` : clé de regroupement par ligne (identité du parent).
    ``values`` : valeur affichée par ligne.
    ``col_end`` : dernière colonne quand la valeur s'étale sur plusieurs colonnes.
    """
    col_end = col if col_end is None else col_end
    n = len(keys)
    i = 0
    while i < n:
        j = i
        while j + 1 < n and keys[j + 1] == keys[i]:
            j += 1
        r1, r2 = start_row + i, start_row + j
        val = values[i]
        if r2 > r1 or col_end > col:
            _merge(ws, r1, col, r2, col_end, val, fill=fill, font=font or _FONT_DATA,
                   align=_AL_LEFT)
        else:
            _set(ws, r1, col, val, fill=fill, font=font or _FONT_DATA, align=_AL_LEFT)
        i = j + 1


def _grille_hauteur(metriques) -> int:
    """Nombre de lignes qu'occupera la grille de lecture (#619).

    Une métrique multi-blocs pesant plusieurs lignes, la grille peut être plus
    haute que le bloc de données qu'elle accompagne : le bloc suivant doit en
    tenir compte pour ne pas la recouvrir.
    """
    return sum(len(_grille_lignes(met)) for met in metriques)


def _write_grille(ws, start_row, metriques, *, is_etat):
    """Écrit la grille de lecture (N..U) à droite d'un bloc.

    #619 — une ligne par bloc de scoring (et non par métrique) : les métriques
    multi-blocs sortent désormais en entier.
    """
    if not metriques:
        return
    r = start_row
    for met in metriques:
        for libelle, unite, labels in _grille_lignes(met):
            _set(ws, r, GR_MET, libelle, fill=_C_GRILLE, font=_FONT_DATA,
                 align=_AL_LEFT)
            _set(ws, r, GR_UNITE, unite, fill=_C_GRILLE, font=_FONT_DATA,
                 align=_AL_CTR)
            _set(ws, r, GR_INDET, "", fill=_SCORE_FILLS["indet"], align=_AL_CTR)
            for s, col in zip(range(1, 6), (GR_S1, GR_S2, GR_S3, GR_S4, GR_S5)):
                _set(ws, r, col, labels[s - 1], fill=_SCORE_FILLS[s], align=_AL_CTR)
            r += 1


def _grille_header(ws, row, title):
    _merge(ws, row, GR_MET, row, GR_S5, title, fill=_C_GRILLE, font=_FONT_HDR,
           align=_AL_CTR)


def _grille_subheader(ws, row, met_label):
    _set(ws, row, GR_MET, met_label, fill=_C_GRILLE, font=_FONT_HDR, align=_AL_CTR)
    _set(ws, row, GR_UNITE, "Unité", fill=_C_GRILLE, font=_FONT_HDR, align=_AL_CTR)
    _set(ws, row, GR_INDET, "Indéterminé", fill=_SCORE_FILLS["indet"],
         font=_FONT_HDR, align=_AL_CTR)
    heads = [
        (GR_S1, "très mauvais\nScore = 1", 1),
        (GR_S2, "Mauvais\nScore = 2", 2),
        (GR_S3, "Score moyen\n= 3", 3),
        (GR_S4, "Bon\nScore = 4", 4),
        (GR_S5, "Très bon\nScore = 5", 5),
    ]
    for col, label, s in heads:
        _set(ws, row, col, label, fill=_SCORE_FILLS[s], font=_FONT_HDR, align=_AL_CTR)


def _render_enjeu_sheet(ws, enjeu, *, is_fcr, codes):
    for col, w in _COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    name = _txt(enjeu.intitule_court) or _txt(enjeu.libelle) or f"Enjeu {enjeu.id_enjeu}"
    importance = getattr(getattr(enjeu, "id_importance", None), "label", "") or ""
    is_prioritaire = "1" in importance

    # Ligne 1 : drapeau prioritaire
    if is_prioritaire:
        _set(ws, 1, COL_A, "Enjeu prioritaire", font=_FONT_FLAG, align=_AL_LEFT,
             border=Border())
    # Ligne 2 : titre
    cat = "FCR" if is_fcr else "enjeu"
    fcr_cat = getattr(getattr(enjeu, "id_categorie_fcr", None), "label", "") or ""
    title = (
        f"Tableau d'arborescence pour un {cat}"
        + (f" « {fcr_cat} »" if is_fcr and fcr_cat else "")
        + " avec grille de lecture des indicateurs"
    )
    _merge(ws, 2, COL_A, 2, COL_K, title, font=_FONT_TITLE, align=_AL_LEFT,
           border=Border())

    top_rows, top_grille = _collect_top(enjeu, codes)
    bottom_rows, bottom_grille = _collect_bottom(enjeu, codes)

    # ============================ BLOC HAUT ============================
    h1 = 4  # bandeau
    h2 = 5  # sous-en-têtes
    data_start = 6
    _set(ws, h1, COL_A, "Enjeu", fill=_C_ENJEU, font=_FONT_HDR_LIGHT, align=_AL_CTR,
         border=_BORDER_MED)
    _merge(ws, h1, COL_B, h1, COL_C, "État de l'enjeu", fill=_C_ENJEU,
           font=_FONT_HDR_LIGHT, align=_AL_CTR, border=_BORDER_MED)
    _merge(ws, h1, COL_D, h1, COL_K, "VISION À LONG TERME", fill=_C_VLT,
           font=_FONT_HDR_LIGHT, align=_AL_CTR, border=_BORDER_MED)
    _grille_header(ws, h1, "Grille de lecture des métriques des indicateurs d'état de conservation")

    _set(ws, h2, COL_A, "", fill=_C_ENJEU, border=_BORDER_MED)
    # #620 — sous-en-tête du bloc « enjeu » en bleu (et non en vert du bloc VLT)
    _merge(ws, h2, COL_B, h2, COL_C, "État actuel de l'enjeu", fill=_C_INFLU_HDR,
           font=_FONT_HDR, align=_AL_CTR)
    top_heads = [
        (COL_D, "Objectifs à long terme"),
        (COL_E, "Niveau d'exigence\n(État visé sur le LT)"),
        (COL_F, "Indicateurs d'état"),
        (COL_G, "Métriques"),
        (COL_H, "Code"),
        (COL_I, "Actions de gestion"),
        (COL_J, "Indicateurs de réponse\n(réalisation)"),
        (COL_K, "Priorité"),
    ]
    for col, label in top_heads:
        _set(ws, h2, col, label, fill=_C_VLT_HDR, font=_FONT_HDR, align=_AL_CTR)
    _grille_subheader(ws, h2, "Métriques")

    n_top = max(len(top_rows), 1)
    # Colonne A (enjeu) + B:C (état actuel) fusionnées sur tout le bloc
    _merge(ws, data_start, COL_A, data_start + n_top - 1, COL_A, name,
           fill=_C_ENJEU_DATA, font=_FONT_HDR, align=_AL_CTR)
    _merge(ws, data_start, COL_B, data_start + n_top - 1, COL_C,
           _txt(enjeu.etat_enjeu), fill=_C_ETAT_DATA, font=_FONT_DATA, align=_AL_LEFT)

    if top_rows:
        keys_olt = [r["path"][0] for r in top_rows]
        keys_ne = [r["ne_key"] for r in top_rows]
        _merge_runs(ws, COL_D, data_start, keys_olt, [r["olt"] for r in top_rows],
                    fill=_C_VLT_DATA)
        _merge_runs(ws, COL_E, data_start, keys_ne, [r["ne"] for r in top_rows],
                    fill=_C_VLT_DATA)
        keys_ind = [r["path"][:3] if len(r["path"]) >= 3 else r["path"] for r in top_rows]
        _merge_runs(ws, COL_F, data_start, keys_ind, [r["ind"] for r in top_rows],
                    fill=_C_VLT_DATA)
        keys_met = [r["path"] for r in top_rows]
        _merge_runs(ws, COL_G, data_start, keys_met, [r["met"] for r in top_rows],
                    fill=_C_VLT_DATA)
        for i, r in enumerate(top_rows):
            rr = data_start + i
            _set(ws, rr, COL_H, r["code"], fill=_C_VLT_DATA, font=_FONT_CODE, align=_AL_CTR)
            _set(ws, rr, COL_I, r["action"], fill=_C_VLT_DATA, font=_FONT_ACTION)
            _set(ws, rr, COL_K, r["priorite"], fill=_C_VLT_DATA, align=_AL_CTR)
        _merge_runs(ws, COL_J, data_start, keys_ne, [r["reponse"] for r in top_rows],
                    fill=_C_VLT_DATA)

    _write_grille(ws, data_start, top_grille, is_etat=True)

    # ============================ BLOC BAS ============================
    # #619 — la grille de lecture peut dépasser le bloc de données (métriques
    # multi-blocs) : le bloc bas démarre après la plus haute des deux.
    gap = 1
    h_top = max(n_top, _grille_hauteur(top_grille))
    b_h1 = data_start + h_top + gap
    b_h2 = b_h1 + 1
    b_data = b_h2 + 1

    _merge(ws, b_h1, COL_A, b_h1, COL_C, "Influences sur l'enjeu", fill=_C_INFLU,
           font=_FONT_HDR_LIGHT, align=_AL_CTR, border=_BORDER_MED)
    _merge(ws, b_h1, COL_D, b_h1, COL_K, "STRATÉGIE D'ACTION (Durée du plan)",
           fill=_C_STRAT, font=_FONT_HDR_LIGHT, align=_AL_CTR, border=_BORDER_MED)
    _grille_header(ws, b_h1, "Grille de lecture des métriques des indicateurs de pression")

    # #620 — plus de colonne « Niveau d'exigence » en face des facteurs
    # d'influence : le bloc « influences » démarre sur les facteurs, et les
    # pressions occupent B:C pour rester alignées sur le bloc haut.
    bottom_heads = [
        (COL_A, "Facteurs d'influence"),
        (COL_B, "Pressions à gérer"),
        (COL_D, "Objectifs opérationnels"),
        (COL_E, "Résultats attendus"),
        (COL_F, "Indicateurs de pression"),
        (COL_G, "Métriques"),
        (COL_H, "Code"),
        (COL_I, "Actions de gestion"),
        (COL_J, "Indicateurs de réponse\n(réalisation)"),
        (COL_K, "Priorité"),
    ]
    for col, label in bottom_heads:
        if col == COL_B:  # « Pressions à gérer » s'étale sur B:C
            _merge(ws, b_h2, COL_B, b_h2, COL_C, label, fill=_C_INFLU_HDR,
                   font=_FONT_HDR, align=_AL_CTR)
            continue
        fill = _C_INFLU_HDR if col == COL_A else _C_STRAT_HDR
        _set(ws, b_h2, col, label, fill=fill, font=_FONT_HDR, align=_AL_CTR)
    _grille_subheader(ws, b_h2, "Métriques")

    if bottom_rows:
        def keyslice(r, depth):
            p = r["path"]
            return p[:depth] if len(p) >= depth else p
        keys_f = [keyslice(r, 1) for r in bottom_rows]
        keys_p = [keyslice(r, 2) for r in bottom_rows]
        keys_oo = [keyslice(r, 3) for r in bottom_rows]
        keys_ra = [keyslice(r, 4) for r in bottom_rows]
        # Facteurs / pressions relèvent du bloc bleu « Influences sur l'enjeu »
        _merge_runs(ws, COL_A, b_data, keys_f, [r.get("facteur", "") for r in bottom_rows],
                    fill=_C_ETAT_DATA)
        _merge_runs(ws, COL_B, b_data, keys_p, [r.get("pression", "") for r in bottom_rows],
                    fill=_C_ETAT_DATA, col_end=COL_C)
        _merge_runs(ws, COL_D, b_data, keys_oo, [r.get("oo", "") for r in bottom_rows],
                    fill=_C_STRAT_DATA)
        _merge_runs(ws, COL_E, b_data, keys_ra, [r.get("ra", "") for r in bottom_rows],
                    fill=_C_STRAT_DATA)
        keys_ind_b = [r["path"] if r.get("met") else keyslice(r, 5) for r in bottom_rows]
        keys_met_b = [r["path"] for r in bottom_rows]
        _merge_runs(ws, COL_F, b_data, keys_ind_b, [r.get("ind", "") for r in bottom_rows],
                    fill=_C_STRAT_DATA)
        _merge_runs(ws, COL_G, b_data, keys_met_b, [r.get("met", "") for r in bottom_rows],
                    fill=_C_STRAT_DATA)
        for i, r in enumerate(bottom_rows):
            rr = b_data + i
            _set(ws, rr, COL_H, r.get("code", ""), fill=_C_STRAT_DATA, font=_FONT_CODE,
                 align=_AL_CTR)
            _set(ws, rr, COL_I, r.get("action", ""), fill=_C_STRAT_DATA, font=_FONT_ACTION)
            _set(ws, rr, COL_K, r.get("priorite", ""), fill=_C_STRAT_DATA, align=_AL_CTR)
        _merge_runs(ws, COL_J, b_data, keys_ra, [r.get("reponse", "") for r in bottom_rows],
                    fill=_C_STRAT_DATA)

    _write_grille(ws, b_data, bottom_grille, is_etat=False)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"


# ---------------------------------------------------------------------------
# Point d'entrée
# ---------------------------------------------------------------------------

def _prefetched_enjeux(plan):
    from .models_enjeux import Enjeu
    qs = (
        Enjeu.objects.filter(id_pg=plan)
        .select_related("id_categorie", "id_categorie_fcr", "id_importance")
        .prefetch_related(
            "objectifs_long_terme__niveaux_exigence__indicateurs__type_indicateur",
            "objectifs_long_terme__niveaux_exigence__indicateurs__metriques__operations",
            # #619 — blocs de scoring complémentaires (grille multi-blocs)
            "objectifs_long_terme__niveaux_exigence__indicateurs__metriques__score_blocks",
            "objectifs_long_terme__niveaux_exigence__indicateurs__operations",
            "facteurs_influence__pressions__objectifs_operationnels__resultats_attendus__indicateurs__type_indicateur",
            "facteurs_influence__pressions__objectifs_operationnels__resultats_attendus__indicateurs__metriques__operations",
            "facteurs_influence__pressions__objectifs_operationnels__resultats_attendus__indicateurs__metriques__score_blocks",
        )
    )
    enjeux = list(qs)
    enjeux.sort(key=lambda e: (
        1 if (e.id_categorie and e.id_categorie.mnemonique == "FCR") else 0,
        e.ordre if e.ordre is not None else 0,
        e.id_enjeu,
    ))
    return enjeux


def build_presentation_workbook(plan) -> bytes:
    """Construit le classeur de présentation de l'arborescence du plan."""
    # #619 — codes d'action locaux au plan (CS1, IP2…), calculés une seule fois.
    from .serializers_operations import compute_operation_codes_for_plan

    wb = Workbook()
    wb.remove(wb.active)
    used_titles: set = set()

    codes = compute_operation_codes_for_plan(plan.pk)
    enjeux = _prefetched_enjeux(plan)
    if not enjeux:
        ws = wb.create_sheet(_sanitize_sheet_title("Arborescence", used_titles))
        ws["A1"] = "Ce plan ne contient pas encore d'arborescence."
        ws["A1"].font = _FONT_TITLE
    else:
        for enjeu in enjeux:
            is_fcr = bool(enjeu.id_categorie and enjeu.id_categorie.mnemonique == "FCR")
            label = _txt(enjeu.intitule_court) or _txt(enjeu.libelle) or f"Enjeu {enjeu.id_enjeu}"
            prefix = "FCR" if is_fcr else "Enjeu"
            title = _sanitize_sheet_title(f"{prefix} - {label}", used_titles)
            ws = wb.create_sheet(title)
            _render_enjeu_sheet(ws, enjeu, is_fcr=is_fcr, codes=codes)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
