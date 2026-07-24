"""
Exports Excel **Budget** et **RH**, en **prévisionnel** et en **suivi**.

Quatre classeurs calqués sur les modèles CICADA :

- ``build_rh_previsionnel_workbook`` — jours prévisionnels par organisme
  gestionnaire (1 onglet / OG), TOTAL, et synthèse par type de poste ;
- ``build_rh_suivi_workbook`` — idem avec colonnes Prévu / Réalisé par année ;
- ``build_budget_previsionnel_workbook`` — budget prévisionnel par OG
  (sections Fonctionnement / Investissement), TOTAL, synthèse par poste de dépense ;
- ``build_budget_suivi_workbook`` — idem avec colonnes Prévu / Réalisé.

Toute la logique de coûts provient du module partagé
:mod:`apps.plans.services_export_finance`.
"""

from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .services_export_finance import build_plan_finance

_ZERO = Decimal(0)

# --- styles ---
_PRIMARY = "FF025359"
_WHITE = "FFFFFFFF"
_HDR_FILL = "FF025359"
_YEAR_FILL = "FFFEC180"
_SECTION_FILL = "FFDCE6E7"
_TOTAL_FILL = "FFC0E3CF"

_F_TITLE = Font(name="Calibri", bold=True, size=13, color=_PRIMARY)
_F_SUB = Font(name="Calibri", bold=True, size=10, color="FF746F6E")
_F_HDR = Font(name="Calibri", bold=True, size=9, color=_WHITE)
_F_YEAR = Font(name="Calibri", bold=True, size=9, color=_PRIMARY)
_F_CELL = Font(name="Calibri", size=9, color="FF343433")
_F_SECTION = Font(name="Calibri", bold=True, size=10, color=_PRIMARY)
_F_TOTAL = Font(name="Calibri", bold=True, size=9, color=_PRIMARY)

_thin = Side(style="thin", color="FFBFC9C9")
_B = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
_AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_AL_R = Alignment(horizontal="right", vertical="center", wrap_text=True)


def _fmt_jours(v) -> str:
    v = v or _ZERO
    if v == 0:
        return "0"
    return f"{v:g}"


def _fmt_euro(v) -> str:
    v = v or _ZERO
    return f"{v:,.0f}".replace(",", " ")


def _title(name: str, used: set) -> str:
    for ch in '[]:*?/\\':
        name = name.replace(ch, " ")
    name = (name.strip() or "Feuille")[:31]
    base, i = name, 2
    while name.lower() in used:
        sfx = f" ({i})"
        name = base[: 31 - len(sfx)] + sfx
        i += 1
    used.add(name.lower())
    return name


def _plan_title(plan) -> str:
    nom = (plan.nom or "").strip() or f"Plan {plan.id_pg}"
    return f"PLAN DE GESTION {plan.annee_debut}-{plan.annee_fin} — {nom}".upper()


def _set(ws, r, c, value, *, fill=None, font=_F_CELL, align=_AL_R):
    cell = ws.cell(r, c, value)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = font
    cell.alignment = align
    cell.border = _B
    return cell


# ---------------------------------------------------------------------------
# Matrice « actions » (A-D + colonnes années) avec sections
# ---------------------------------------------------------------------------

class _ActionMatrix:
    """Feuille : Enjeu/FCR | Code simplifié | Code action PG | Libellé | années | TOTAL."""

    HEADERS = ["Enjeu / FCR", "Code simplifié", "Code action PG", "Libellé de l'action"]

    def __init__(self, ws, years, *, suivi, fmt):
        self.ws = ws
        self.years = years
        self.suivi = suivi
        self.fmt = fmt
        self.ncols = 4 + (2 * len(years) + 2 if suivi else len(years) + 1)
        self.r = 1
        widths = {1: 18, 2: 13, 3: 15, 4: 44}
        for c in range(5, self.ncols + 1):
            widths[c] = 10
        for c, w in widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w

    # --- positions colonnes ---
    def _yc(self, i, real=False):
        if self.suivi:
            return 5 + 2 * i + (1 if real else 0)
        return 5 + i

    def _tc(self, real=False):
        if self.suivi:
            return 5 + 2 * len(self.years) + (1 if real else 0)
        return 5 + len(self.years)

    def head(self, title, subtitle):
        ws = self.ws
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=self.ncols)
        c = ws.cell(1, 1, title); c.font = _F_TITLE; c.alignment = _AL_L
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=self.ncols)
        c = ws.cell(2, 1, subtitle); c.font = _F_SUB; c.alignment = _AL_L
        self.r = 3
        hr = self.r
        # colonnes A-D (fusion sur 2 lignes en mode suivi)
        for i, h in enumerate(self.HEADERS):
            col = i + 1
            if self.suivi:
                ws.merge_cells(start_row=hr, start_column=col, end_row=hr + 1, end_column=col)
            _set(ws, hr, col, h, fill=_HDR_FILL, font=_F_HDR, align=_AL_C)
        # années
        for i, y in enumerate(self.years):
            if self.suivi:
                ws.merge_cells(start_row=hr, start_column=self._yc(i), end_row=hr, end_column=self._yc(i, True))
            _set(ws, hr, self._yc(i), str(y), fill=_YEAR_FILL, font=_F_YEAR, align=_AL_C)
        if self.suivi:
            ws.merge_cells(start_row=hr, start_column=self._tc(), end_row=hr, end_column=self._tc(True))
        _set(ws, hr, self._tc(), "TOTAL", fill=_YEAR_FILL, font=_F_YEAR, align=_AL_C)
        if self.suivi:
            # sous-en-tête Prévu / Réalisé
            for i in range(len(self.years)):
                _set(ws, hr + 1, self._yc(i), "Prévu", fill=_SECTION_FILL, font=_F_TOTAL, align=_AL_C)
                _set(ws, hr + 1, self._yc(i, True), "Réalisé", fill=_SECTION_FILL, font=_F_TOTAL, align=_AL_C)
            _set(ws, hr + 1, self._tc(), "Prévu", fill=_SECTION_FILL, font=_F_TOTAL, align=_AL_C)
            _set(ws, hr + 1, self._tc(True), "Réalisé", fill=_SECTION_FILL, font=_F_TOTAL, align=_AL_C)
            self.r = hr + 2
        else:
            self.r = hr + 1

    def section(self, label):
        ws = self.ws
        ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=self.ncols)
        _set(ws, self.r, 1, f"► {label}", fill=_SECTION_FILL, font=_F_SECTION, align=_AL_L)
        self.r += 1

    def row(self, enjeu, cat, code, libelle, prev, real=None):
        ws = self.ws
        _set(ws, self.r, 1, enjeu, font=_F_CELL, align=_AL_L)
        _set(ws, self.r, 2, cat, font=_F_CELL, align=_AL_C)
        _set(ws, self.r, 3, code, font=_F_CELL, align=_AL_C)
        _set(ws, self.r, 4, libelle, font=_F_CELL, align=_AL_L)
        tp = _ZERO
        tr = _ZERO
        for i, y in enumerate(self.years):
            pv = prev.get(y, _ZERO)
            tp += pv
            _set(ws, self.r, self._yc(i), self.fmt(pv))
            if self.suivi:
                rv = (real or {}).get(y, _ZERO)
                tr += rv
                _set(ws, self.r, self._yc(i, True), self.fmt(rv))
        _set(ws, self.r, self._tc(), self.fmt(tp), fill=_TOTAL_FILL, font=_F_TOTAL)
        if self.suivi:
            _set(ws, self.r, self._tc(True), self.fmt(tr), fill=_TOTAL_FILL, font=_F_TOTAL)
        self.r += 1

    def total(self, label, prev, real=None):
        ws = self.ws
        ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=4)
        _set(ws, self.r, 1, label, fill=_TOTAL_FILL, font=_F_TOTAL, align=_AL_L)
        tp = tr = _ZERO
        for i, y in enumerate(self.years):
            pv = prev.get(y, _ZERO); tp += pv
            _set(ws, self.r, self._yc(i), self.fmt(pv), fill=_TOTAL_FILL, font=_F_TOTAL)
            if self.suivi:
                rv = (real or {}).get(y, _ZERO); tr += rv
                _set(ws, self.r, self._yc(i, True), self.fmt(rv), fill=_TOTAL_FILL, font=_F_TOTAL)
        _set(ws, self.r, self._tc(), self.fmt(tp), fill=_TOTAL_FILL, font=_F_TOTAL)
        if self.suivi:
            _set(ws, self.r, self._tc(True), self.fmt(tr), fill=_TOTAL_FILL, font=_F_TOTAL)
        self.r += 1

    def freeze(self):
        self.ws.freeze_panes = self.ws.cell(4 if not self.suivi else 5, 5).coordinate
        self.ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Matrice « synthèse » (libellé [+ coût/jour] + colonnes années)
# ---------------------------------------------------------------------------

class _LabelMatrix:
    def __init__(self, ws, years, *, suivi, fmt, label_header="Type de poste", cout_col=False):
        self.ws = ws
        self.years = years
        self.suivi = suivi
        self.fmt = fmt
        self.cout_col = cout_col
        self.base = 2 + (1 if cout_col else 0)   # 1ère colonne année
        self.ncols = (self.base - 1) + (2 * len(years) + 2 if suivi else len(years) + 1)
        self.r = 1
        ws.column_dimensions["A"].width = 32
        if cout_col:
            ws.column_dimensions["B"].width = 11
        for c in range(self.base, self.ncols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 10
        self.label_header = label_header

    def _yc(self, i, real=False):
        if self.suivi:
            return self.base + 2 * i + (1 if real else 0)
        return self.base + i

    def _tc(self, real=False):
        if self.suivi:
            return self.base + 2 * len(self.years) + (1 if real else 0)
        return self.base + len(self.years)

    def head(self, title, subtitle):
        ws = self.ws
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=self.ncols)
        c = ws.cell(1, 1, title); c.font = _F_TITLE; c.alignment = _AL_L
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=self.ncols)
        c = ws.cell(2, 1, subtitle); c.font = _F_SUB; c.alignment = _AL_L
        hr = 3
        if self.suivi:
            ws.merge_cells(start_row=hr, start_column=1, end_row=hr + 1, end_column=1)
        _set(ws, hr, 1, self.label_header, fill=_HDR_FILL, font=_F_HDR, align=_AL_C)
        if self.cout_col:
            if self.suivi:
                ws.merge_cells(start_row=hr, start_column=2, end_row=hr + 1, end_column=2)
            _set(ws, hr, 2, "Coût/jour", fill=_HDR_FILL, font=_F_HDR, align=_AL_C)
        for i, y in enumerate(self.years):
            if self.suivi:
                ws.merge_cells(start_row=hr, start_column=self._yc(i), end_row=hr, end_column=self._yc(i, True))
            _set(ws, hr, self._yc(i), str(y), fill=_YEAR_FILL, font=_F_YEAR, align=_AL_C)
        if self.suivi:
            ws.merge_cells(start_row=hr, start_column=self._tc(), end_row=hr, end_column=self._tc(True))
        _set(ws, hr, self._tc(), "TOTAL", fill=_YEAR_FILL, font=_F_YEAR, align=_AL_C)
        if self.suivi:
            for i in range(len(self.years)):
                _set(ws, hr + 1, self._yc(i), "Prévu", fill=_SECTION_FILL, font=_F_TOTAL, align=_AL_C)
                _set(ws, hr + 1, self._yc(i, True), "Réalisé", fill=_SECTION_FILL, font=_F_TOTAL, align=_AL_C)
            _set(ws, hr + 1, self._tc(), "Prévu", fill=_SECTION_FILL, font=_F_TOTAL, align=_AL_C)
            _set(ws, hr + 1, self._tc(True), "Réalisé", fill=_SECTION_FILL, font=_F_TOTAL, align=_AL_C)
            self.r = hr + 2
        else:
            self.r = hr + 1

    def section(self, label):
        ws = self.ws
        ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=self.ncols)
        _set(ws, self.r, 1, f"► {label}", fill=_SECTION_FILL, font=_F_SECTION, align=_AL_L)
        self.r += 1

    def row(self, label, prev, real=None, *, cout=None, total_style=False):
        ws = self.ws
        fill = _TOTAL_FILL if total_style else None
        font = _F_TOTAL if total_style else _F_CELL
        _set(ws, self.r, 1, label, fill=fill, font=font, align=_AL_L)
        if self.cout_col:
            _set(ws, self.r, 2, ("" if cout in (None, "") else self.fmt(cout) if False else cout), fill=fill, font=font, align=_AL_C)
        tp = tr = _ZERO
        for i, y in enumerate(self.years):
            pv = prev.get(y, _ZERO); tp += pv
            _set(ws, self.r, self._yc(i), self.fmt(pv), fill=fill, font=font)
            if self.suivi:
                rv = (real or {}).get(y, _ZERO); tr += rv
                _set(ws, self.r, self._yc(i, True), self.fmt(rv), fill=fill, font=font)
        _set(ws, self.r, self._tc(), self.fmt(tp), fill=_TOTAL_FILL, font=_F_TOTAL)
        if self.suivi:
            _set(ws, self.r, self._tc(True), self.fmt(tr), fill=_TOTAL_FILL, font=_F_TOTAL)
        self.r += 1

    def freeze(self):
        self.ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# Agrégations plan
# ---------------------------------------------------------------------------

def _sum_years(seq_of_dicts, years):
    out = {y: _ZERO for y in years}
    for d in seq_of_dicts:
        for y in years:
            out[y] += d.get(y, _ZERO)
    return out


def _poste_labels(pf, benevole_last=True):
    """Libellés de poste distincts (agrégés tous organismes), bénévoles/partenaires en fin."""
    labels = []
    for (oid, label, y) in pf.poste_jours:
        if label not in labels:
            labels.append(label)
    specials = ["Bénévoles", "Partenaires"]
    normals = [l for l in labels if l not in specials]
    normals.sort(key=lambda s: s.lower())
    ordered = normals + [s for s in specials if s in labels]
    return ordered


def _poste_year(pf, label, key):
    """{année: valeur} agrégé tous organismes pour un type de poste et une clé."""
    out = {y: _ZERO for y in pf.years}
    for (oid, lab, y), v in pf.poste_jours.items():
        if lab == label and y in out:
            out[y] += v.get(key, _ZERO)
    return out


def _poste_cout_jour(pf, label):
    for (oid, lab, y), v in pf.poste_jours.items():
        if lab == label and v.get("cout_jour"):
            return v["cout_jour"]
    return None


# ---------------------------------------------------------------------------
# Builders
# ---------------------------------------------------------------------------

def _org_sheets(pf):
    """Organismes à exporter (exclut « Non ventilé » vide) dans un ordre stable."""
    return pf.org_ids()


def build_rh_previsionnel_workbook(plan) -> bytes:
    pf = build_plan_finance(plan)
    wb = Workbook(); wb.remove(wb.active); used = set()
    title = _plan_title(plan)

    def org_rows(oid):
        rows = []
        for af in pf.actions:
            prev = {y: af.cell(oid, y).jours for y in pf.years}
            if sum(prev.values()) != 0:
                rows.append((af, prev))
        return rows

    for oid in _org_sheets(pf):
        rows = org_rows(oid)
        if not rows:
            continue
        name = pf.org_names.get(oid, "Non ventilé")
        m = _ActionMatrix(wb.create_sheet(_title(f"RH Prév. — {name}", used)), pf.years, suivi=False, fmt=_fmt_jours)
        m.head(title, f"RESSOURCES HUMAINES — PRÉVISIONNEL (jours) — {name}")
        for af, prev in rows:
            m.row(af.enjeu_label, af.categorie, af.code, af.libelle, prev)
        m.total(f"TOTAL jours prévus — {name}", _sum_years([p for _, p in rows], pf.years))
        m.freeze()

    # TOTAL tous organismes
    m = _ActionMatrix(wb.create_sheet(_title("TOTAL", used)), pf.years, suivi=False, fmt=_fmt_jours)
    m.head(title, "RESSOURCES HUMAINES — PRÉVISIONNEL (jours) — TOTAL")
    all_rows = []
    for af in pf.actions:
        prev = {y: af.year_total(y).jours for y in pf.years}
        if sum(prev.values()) != 0:
            m.row(af.enjeu_label, af.categorie, af.code, af.libelle, prev)
            all_rows.append(prev)
    m.total("TOTAL jours prévus", _sum_years(all_rows, pf.years))
    m.freeze()

    # Par type de poste
    _rh_par_poste_sheet(wb, pf, used, suivi=False)

    return _save(wb)


def build_rh_suivi_workbook(plan) -> bytes:
    pf = build_plan_finance(plan)
    wb = Workbook(); wb.remove(wb.active); used = set()
    title = _plan_title(plan)

    for oid in _org_sheets(pf):
        rows = []
        for af in pf.actions:
            prev = {y: af.cell(oid, y).jours for y in pf.years}
            real = {y: af.cell(oid, y).rjours for y in pf.years}
            if sum(prev.values()) != 0 or sum(real.values()) != 0:
                rows.append((af, prev, real))
        if not rows:
            continue
        name = pf.org_names.get(oid, "Non ventilé")
        m = _ActionMatrix(wb.create_sheet(_title(f"RH Suivi — {name}", used)), pf.years, suivi=True, fmt=_fmt_jours)
        m.head(title, f"RESSOURCES HUMAINES — SUIVI (jours) — {name}")
        for af, prev, real in rows:
            m.row(af.enjeu_label, af.categorie, af.code, af.libelle, prev, real)
        m.total(f"TOTAL — {name}", _sum_years([p for _, p, _ in rows], pf.years), _sum_years([r for _, _, r in rows], pf.years))
        m.freeze()

    m = _ActionMatrix(wb.create_sheet(_title("TOTAL", used)), pf.years, suivi=True, fmt=_fmt_jours)
    m.head(title, "RESSOURCES HUMAINES — SUIVI (jours) — TOTAL")
    pr, rr = [], []
    for af in pf.actions:
        prev = {y: af.year_total(y).jours for y in pf.years}
        real = {y: af.year_total(y).rjours for y in pf.years}
        if sum(prev.values()) != 0 or sum(real.values()) != 0:
            m.row(af.enjeu_label, af.categorie, af.code, af.libelle, prev, real)
            pr.append(prev); rr.append(real)
    m.total("TOTAL", _sum_years(pr, pf.years), _sum_years(rr, pf.years))
    m.freeze()

    _rh_par_poste_sheet(wb, pf, used, suivi=True)
    return _save(wb)


def _rh_par_poste_sheet(wb, pf, used, *, suivi):
    label = "Total par poste" if suivi else "RH par type de poste"
    m = _LabelMatrix(wb.create_sheet(_title(label, used)), pf.years, suivi=suivi, fmt=_fmt_jours)
    m.head(_plan_title(pf.plan), f"RESSOURCES HUMAINES — {'SUIVI' if suivi else 'PRÉVISIONNEL'} (jours) — TOTAL (tous organismes)")
    for lab in _poste_labels(pf):
        prev = _poste_year(pf, lab, "prev")
        real = _poste_year(pf, lab, "real") if suivi else None
        if sum(prev.values()) == 0 and (not real or sum(real.values()) == 0):
            continue
        m.row(lab, prev, real)
    m.freeze()


def build_budget_previsionnel_workbook(plan) -> bytes:
    pf = build_plan_finance(plan)
    wb = Workbook(); wb.remove(wb.active); used = set()
    title = _plan_title(plan)

    def sheet_for(oid, name):
        rows_f, rows_i = [], []
        for af in pf.actions:
            f = {y: af.cell(oid, y).tot_fonct for y in pf.years}
            i = {y: af.cell(oid, y).tot_invest for y in pf.years}
            if sum(f.values()) != 0:
                rows_f.append((af, f))
            if sum(i.values()) != 0:
                rows_i.append((af, i))
        if not rows_f and not rows_i:
            return
        m = _ActionMatrix(wb.create_sheet(_title(f"Budget Prév. — {name}", used)), pf.years, suivi=False, fmt=_fmt_euro)
        m.head(title, f"BUDGET PRÉVISIONNEL (€) — {name}")
        m.section("Fonctionnement")
        for af, f in rows_f:
            m.row(af.enjeu_label, af.categorie, af.code, af.libelle, f)
        m.total(f"TOTAL — Fonctionnement {name}", _sum_years([f for _, f in rows_f], pf.years))
        m.section("Investissement")
        for af, i in rows_i:
            m.row(af.enjeu_label, af.categorie, af.code, af.libelle, i)
        m.total(f"TOTAL — Investissement {name}", _sum_years([i for _, i in rows_i], pf.years))
        m.freeze()

    for oid in _org_sheets(pf):
        sheet_for(oid, pf.org_names.get(oid, "Non ventilé"))

    # TOTAL OG (tous organismes)
    m = _ActionMatrix(wb.create_sheet(_title("TOTAL OG", used)), pf.years, suivi=False, fmt=_fmt_euro)
    m.head(title, "BUDGET PRÉVISIONNEL (€) — TOTAL (tous organismes)")
    m.section("Fonctionnement")
    rf = []
    for af in pf.actions:
        f = {y: af.year_total(y).tot_fonct for y in pf.years}
        if sum(f.values()) != 0:
            m.row(af.enjeu_label, af.categorie, af.code, af.libelle, f); rf.append(f)
    m.total("TOTAL — Fonctionnement", _sum_years(rf, pf.years))
    m.section("Investissement")
    ri = []
    for af in pf.actions:
        i = {y: af.year_total(y).tot_invest for y in pf.years}
        if sum(i.values()) != 0:
            m.row(af.enjeu_label, af.categorie, af.code, af.libelle, i); ri.append(i)
    m.total("TOTAL — Investissement", _sum_years(ri, pf.years))
    m.freeze()

    _budget_par_poste_sheet(wb, pf, used, suivi=False)
    return _save(wb)


def build_budget_suivi_workbook(plan) -> bytes:
    pf = build_plan_finance(plan)
    wb = Workbook(); wb.remove(wb.active); used = set()
    title = _plan_title(plan)

    def sheet_for(oid, name):
        rf, ri = [], []
        for af in pf.actions:
            fp = {y: af.cell(oid, y).tot_fonct for y in pf.years}
            fr = {y: af.cell(oid, y).rtot_fonct for y in pf.years}
            ip = {y: af.cell(oid, y).tot_invest for y in pf.years}
            ir = {y: af.cell(oid, y).rtot_invest for y in pf.years}
            if sum(fp.values()) or sum(fr.values()):
                rf.append((af, fp, fr))
            if sum(ip.values()) or sum(ir.values()):
                ri.append((af, ip, ir))
        if not rf and not ri:
            return
        m = _ActionMatrix(wb.create_sheet(_title(f"Budget Suivi — {name}", used)), pf.years, suivi=True, fmt=_fmt_euro)
        m.head(title, f"BUDGET SUIVI (€) — {name}")
        m.section("Fonctionnement")
        for af, fp, fr in rf:
            m.row(af.enjeu_label, af.categorie, af.code, af.libelle, fp, fr)
        m.total(f"TOTAL — Fonctionnement {name}", _sum_years([x for _, x, _ in rf], pf.years), _sum_years([x for _, _, x in rf], pf.years))
        m.section("Investissement")
        for af, ip, ir in ri:
            m.row(af.enjeu_label, af.categorie, af.code, af.libelle, ip, ir)
        m.total(f"TOTAL — Investissement {name}", _sum_years([x for _, x, _ in ri], pf.years), _sum_years([x for _, _, x in ri], pf.years))
        m.freeze()

    for oid in _org_sheets(pf):
        sheet_for(oid, pf.org_names.get(oid, "Non ventilé"))

    m = _ActionMatrix(wb.create_sheet(_title("TOTAL", used)), pf.years, suivi=True, fmt=_fmt_euro)
    m.head(title, "BUDGET SUIVI (€) — TOTAL (tous organismes)")
    m.section("Fonctionnement")
    fp_all, fr_all = [], []
    for af in pf.actions:
        fp = {y: af.year_total(y).tot_fonct for y in pf.years}
        fr = {y: af.year_total(y).rtot_fonct for y in pf.years}
        if sum(fp.values()) or sum(fr.values()):
            m.row(af.enjeu_label, af.categorie, af.code, af.libelle, fp, fr); fp_all.append(fp); fr_all.append(fr)
    m.total("TOTAL — Fonctionnement", _sum_years(fp_all, pf.years), _sum_years(fr_all, pf.years))
    m.section("Investissement")
    ip_all, ir_all = [], []
    for af in pf.actions:
        ip = {y: af.year_total(y).tot_invest for y in pf.years}
        ir = {y: af.year_total(y).rtot_invest for y in pf.years}
        if sum(ip.values()) or sum(ir.values()):
            m.row(af.enjeu_label, af.categorie, af.code, af.libelle, ip, ir); ip_all.append(ip); ir_all.append(ir)
    m.total("TOTAL — Investissement", _sum_years(ip_all, pf.years), _sum_years(ir_all, pf.years))
    m.freeze()

    _budget_par_poste_sheet(wb, pf, used, suivi=True)
    return _save(wb)


def _plan_cost_by_year(pf, attr):
    """{année: Σ sur actions/organismes d'un composant de coût}."""
    out = {y: _ZERO for y in pf.years}
    for af in pf.actions:
        for y in pf.years:
            out[y] += getattr(af.year_total(y), attr)
    return out


def _budget_par_poste_sheet(wb, pf, used, *, suivi):
    name = "Total par type de dépense" if suivi else "TOTAL par poste de dépense"
    m = _LabelMatrix(wb.create_sheet(_title(name, used)), pf.years, suivi=suivi,
                     fmt=_fmt_euro, label_header="Poste", cout_col=True)
    m.head(_plan_title(pf.plan), f"BUDGET {'SUIVI' if suivi else 'PRÉVISIONNEL'} (€) — par poste de dépense")
    salaried = [l for l in _poste_labels(pf) if l not in ("Bénévoles", "Partenaires")]

    def poste_prev_real(lab, key_prev, key_real):
        prev = _poste_year(pf, lab, key_prev)
        real = _poste_year(pf, lab, key_real) if suivi else None
        return prev, real

    def _nonzero(prev, real):
        return sum(prev.values()) != 0 or (real and sum(real.values()) != 0)

    # Fonctionnement - salarial
    m.section("Fonctionnement — Coût salarial")
    for lab in salaried:
        prev, real = poste_prev_real(lab, "salf_prev", "salf_real")
        if _nonzero(prev, real):
            m.row(lab, prev, real, cout=_poste_cout_jour(pf, lab))
    # Fonctionnement - autres
    m.section("Fonctionnement — autres coûts")
    m.row("Prestataire", _plan_cost_by_year(pf, "prest_fonct"),
          _plan_cost_by_year_real(pf, "rprest_fonct") if suivi else None, cout="/")
    m.row("Autres coûts de fonctionnement", _plan_cost_by_year(pf, "autre_fonct"),
          _plan_cost_by_year_real(pf, "rautre_fonct") if suivi else None, cout="/")
    # Investissement - salarial
    m.section("Investissement — Coût salarial")
    for lab in salaried:
        prev, real = poste_prev_real(lab, "sali_prev", "sali_real")
        if _nonzero(prev, real):
            m.row(lab, prev, real, cout=_poste_cout_jour(pf, lab))
    # Investissement - autres
    m.section("Investissement — autres coûts")
    m.row("Prestataire", _plan_cost_by_year(pf, "prest_invest"),
          _plan_cost_by_year_real(pf, "rprest_invest") if suivi else None, cout="/")
    m.row("Autres coûts d'investissement", _plan_cost_by_year(pf, "autre_invest"),
          _plan_cost_by_year_real(pf, "rautre_invest") if suivi else None, cout="/")
    # TOTAL
    tot_prev = {y: _ZERO for y in pf.years}
    for af in pf.actions:
        for y in pf.years:
            tot_prev[y] += af.year_total(y).tot
    tot_real = None
    if suivi:
        tot_real = {y: _ZERO for y in pf.years}
        for af in pf.actions:
            for y in pf.years:
                tot_real[y] += af.year_total(y).rtot
    m.row("TOTAL", tot_prev, tot_real, cout="", total_style=True)
    m.freeze()


def _plan_cost_by_year_real(pf, attr):
    out = {y: _ZERO for y in pf.years}
    for af in pf.actions:
        for y in pf.years:
            out[y] += getattr(af.year_total(y), attr)
    return out


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
