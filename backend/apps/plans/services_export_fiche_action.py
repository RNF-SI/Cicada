"""
Export Excel « Fiche action » — une fiche (onglet) par opération d'un plan.

Calqué sur le modèle « Modèle_Export_CICADA_Fiche_action », qui distingue :

- **Action CS** (catégorie « Connaissance et suivi », `id_categorie_action_reserve`
  mnémonique ``CS``) : cadre indicateur d'état → niveau d'exigence → OLT → enjeu,
  avec détails du suivi / protocole ;
- **Action hors CS** (toutes les autres catégories) : cadre indicateur de pression
  → résultats attendus → OO → enjeu.

Les deux variantes partagent le volet administratif et financier (programmation
annuelle, blocs budgétaires par organisme gestionnaire, programmation mensuelle,
financeurs) puis la rubrique des indicateurs de réponse.

#626 — la fiche porte en plus : les deux codes de l'action (code local du plan,
type « CS1 », et code du référentiel Gestref), le descriptif et les objectifs de
chaque protocole du suivi, une petite carte du contour de l'emprise, et — pour
les indicateurs de réponse — la métrique entre parenthèses hors grille ou la
grille de scoring 5 paliers quand la métrique est au format GRILLE.

Point d'entrée public : :func:`build_fiche_action_workbook`.

Volet financier — note d'implémentation : le modèle de données ne stocke pas le
« coût salarial » ; on le recalcule (jours des postes × ``Poste.cout_jour``), comme
le front. La ventilation fonctionnement / investissement des jours suit
``OperationAnneeRH.categorie_depense`` ; prestataire / autres coûts / budgets
proviennent de ``OperationAnneeOrganisme`` (ou de ``OperationAnnee`` en mode sans
ventilation par organisme).
"""

from __future__ import annotations

import io
import math
from collections import defaultdict
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import export_theme
from .metrique_seuils import intervalle_palier
from .services_export_finance import build_action_finance, poste_entry_factory

# ---------------------------------------------------------------------------
# Styles (palette CICADA)
# ---------------------------------------------------------------------------

_PRIMARY = "FF025359"
_TERRA = "FFB74D5D"
_GREEN = "FF04854B"
_BEIGE = "FFF3EFEA"
_SECTION = "FF025359"
_LABEL_FILL = "FFDCE6E7"
_SUBLABEL_FILL = "FFEDF4F4"
_TOTAL_FILL = "FFC0E3CF"
_YEARHDR_FILL = "FFFEC180"
_WHITE = "FFFFFFFF"

_F_TITLE = Font(name="Calibri", bold=True, size=13, color=_WHITE)
_F_SECTION = Font(name="Calibri", bold=True, size=11, color=_WHITE)
_F_LABEL = Font(name="Calibri", bold=True, size=10, color=_PRIMARY)
_F_SUBLABEL = Font(name="Calibri", size=10, color="FF343433")
_F_VALUE = Font(name="Calibri", size=10, color="FF343433")
_F_TOTAL = Font(name="Calibri", bold=True, size=10, color=_PRIMARY)
_F_PRIO = Font(name="Calibri", bold=True, size=12, color=_WHITE)
_F_YEAR = Font(name="Calibri", bold=True, size=9, color=_PRIMARY)
_F_X = Font(name="Calibri", bold=True, size=11, color=_GREEN)


def _appliquer_couleur_instance():
    """
    Réaligne les styles sur la couleur d'export de l'instance (#601).

    Appelé au début de chaque construction de classeur : cf. `export_theme`
    pour pourquoi ces styles restent des variables de module.
    """
    global _PRIMARY, _SECTION, _F_LABEL, _F_TOTAL, _F_YEAR

    couleur = export_theme.argb()
    if couleur == _PRIMARY:
        return
    _PRIMARY = _SECTION = couleur
    _F_LABEL = Font(name="Calibri", bold=True, size=10, color=couleur)
    _F_TOTAL = Font(name="Calibri", bold=True, size=10, color=couleur)
    _F_YEAR = Font(name="Calibri", bold=True, size=9, color=couleur)

# Palette des paliers de score (design system) — texte noir uniquement (#626).
_SCORE_FILLS = {
    1: "FFFF7579", 2: "FFFA9965", 3: "FFF7D35C", 4: "FF82DB8A", 5: "FF81C9D8",
}
_SCORE_LABELS = {
    1: "Très mauvais", 2: "Mauvais", 3: "Moyen", 4: "Bon", 5: "Très bon",
}
_F_SCOREHDR = Font(name="Calibri", bold=True, size=10, color="FF343433")

_thin = Side(style="thin", color="FFBFC9C9")
_med = Side(style="medium", color="FF9DB3B4")
_B = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_BM = Border(left=_med, right=_med, top=_med, bottom=_med)

_AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
_AL_LT = Alignment(horizontal="left", vertical="top", wrap_text=True)
_AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
_AL_R = Alignment(horizontal="right", vertical="center", wrap_text=True)


def _n(value) -> Decimal:
    if value in (None, ""):
        return Decimal(0)
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal(0)


def _txt(value) -> str:
    if value is None:
        return ""
    return value.strip() if isinstance(value, str) else str(value)


def _euro(value) -> str:
    d = _n(value)
    if d == 0:
        return "/"
    return f"{d:,.0f} €".replace(",", " ")


def _jours(value) -> str:
    d = _n(value)
    if d == 0:
        return "/"
    return f"{d:g}"


def _sanitize_title(title: str, used: set) -> str:
    for ch in '[]:*?/\\':
        title = title.replace(ch, " ")
    title = (title.strip() or "Action")[:31]
    base, i = title, 2
    while title.lower() in used:
        sfx = f" ({i})"
        title = base[: 31 - len(sfx)] + sfx
        i += 1
    used.add(title.lower())
    return title


# ---------------------------------------------------------------------------
# Grilles de métrique (#626) — même formatage que MetriqueGridDisplayComponent
# ---------------------------------------------------------------------------

def _fmt_dec(value) -> str:
    """Nombre sans zéros terminaux (≤ 4 décimales, comme le front)."""
    try:
        s = f"{Decimal(str(value)):f}"
    except Exception:
        return _txt(value)
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s or "0"


def _metrique_kind(m) -> str:
    """TEXTE / CHIFFRE / NUMERIQUE (déduit si le type n'est pas renseigné)."""
    mnemo = (getattr(getattr(m, "type_metrique", None), "mnemonique", "") or "").upper()
    if mnemo:
        return mnemo
    levels = range(1, 6)
    has_labels = any(_txt(getattr(m, f"score_{i}_label", "")) for i in levels)
    has_vals = any(getattr(m, f"score_{i}_val", None) is not None for i in levels)
    has_bounds = any(
        getattr(m, f"score_{i}_inf", None) is not None
        or getattr(m, f"score_{i}_sup", None) is not None
        for i in levels
    )
    if has_labels and not has_bounds:
        return "TEXTE"
    if has_vals and not has_bounds:
        return "CHIFFRE"
    return "NUMERIQUE"


def _is_grille(m) -> bool:
    """#452 — la grille de scoring est opt-in (format GRILLE) côté réponse."""
    fmt = (getattr(getattr(m, "format_metrique", None), "mnemonique", "") or "").upper()
    if fmt == "GRILLE":
        return True
    if fmt == "SIMPLE":
        return False
    # Format non renseigné : grille seulement si des données de grille existent
    # (donnée héritée d'avant #452).
    if m.score_blocks.all():
        return True
    return any(
        getattr(m, f"score_{i}_inf", None) is not None
        or getattr(m, f"score_{i}_sup", None) is not None
        or getattr(m, f"score_{i}_val", None) is not None
        or _txt(getattr(m, f"score_{i}_label", ""))
        for i in range(1, 6)
    )


def _interval(obj, level, sens, inactive) -> str:
    """
    Intervalle d'un palier (inclusivité sens-aware, cf. #545/#554).

    La règle vit dans `metrique_seuils` : elle est partagée avec l'export
    arborescence, où les deux notations avaient divergé (#619).
    """
    return intervalle_palier(obj, level, sens=sens, inactifs=inactive)


def _block_label(m, idx) -> str:
    if idx == 0:
        intitule, unite = getattr(m, "bloc_intitule", "") or "", m.unite or ""
    else:
        block = list(m.score_blocks.all())[idx - 1]
        intitule, unite = block.intitule or "", block.unite or ""
    intitule = intitule.strip()
    if intitule:
        return f"{intitule} ({unite.strip()})" if unite.strip() else intitule
    return "Bloc " + chr(ord("A") + idx)


def _grid_cell(m, level) -> str:
    """Contenu d'une cellule de palier (multi-lignes si blocs ET/OU)."""
    kind = _metrique_kind(m)
    inactive = list(m.inactive_levels or [])
    blocks = list(m.score_blocks.all())

    if kind == "TEXTE":
        main = "" if level in inactive else _txt(getattr(m, f"score_{level}_label", ""))
    elif kind == "CHIFFRE":
        val = getattr(m, f"score_{level}_val", None)
        main = "" if (level in inactive or val is None) else _fmt_dec(val)
    else:
        main = _interval(m, level, m.sens_variation, inactive)

    if not blocks:
        return main

    lines = []
    if main:
        lines.append(_paren(f"{_block_label(m, 0)} : {main}", m.group_open, m.group_close))
    for idx, block in enumerate(blocks, start=1):
        text = _interval(block, level, block.sens_variation, list(block.inactive_levels or []))
        if not text:
            continue
        op = "ET" if block.logical_op == "AND" else "OU"
        prefix = f"{op} " if lines else ""
        lines.append(prefix + _paren(f"{_block_label(m, idx)} : {text}",
                                     block.group_open, block.group_close))
    return "\n".join(lines)


def _paren(text, group_open, group_close) -> str:
    return f"{'(' * (group_open or 0)}{text}{')' * (group_close or 0)}"


# ---------------------------------------------------------------------------
# Carte de localisation (#626, #629) — emprise sur fond de carte, rendu PNG
# ---------------------------------------------------------------------------

# #629 : le contour seul, étiré sur toute l'image, se lisait comme « un carré
# vert ». On rend désormais une vraie mini-carte : projection Web Mercator,
# tuiles XYZ du fond de carte, puis l'emprise en surcouche semi-transparente.
# Le fond est optionnel : réseau indisponible (prod isolée, CI) → repli sur le
# fond uni, l'export n'échoue jamais pour une tuile manquante.

_MAP_W, _MAP_H = 520, 360        # taille de la vignette (px)
_MAP_ROWS = 14                   # lignes Excel fusionnées pour former la case
_MAP_PAD = 4                     # marge entre la vignette et le bord de la case
_TILE_SIZE = 256
_TILE_ZOOM_MAX = 16
_TILE_ZOOM_MIN = 3
_TILE_PAD = 0.14                 # marge autour de l'emprise (fraction de l'image)
_TILE_TIMEOUT = 4                # secondes, par tuile
_TILE_MAX = 24                   # garde-fou : nombre de tuiles téléchargées
_TILE_CACHE: dict[str, bytes | None] = {}


def _tile_url_template() -> str:
    from django.conf import settings

    return getattr(
        settings, "EXPORT_MAP_TILE_URL",
        "https://tile.openstreetmap.org/{z}/{x}/{y}.png",
    )


def _tile_attribution() -> str:
    from django.conf import settings

    return getattr(settings, "EXPORT_MAP_ATTRIBUTION", "© OpenStreetMap")


def _fetch_tile(url):
    """Télécharge une tuile (bytes) ou None. Mémoïsé : les actions d'un même
    plan partagent le plus souvent la même emprise (site)."""
    if url in _TILE_CACHE:
        return _TILE_CACHE[url]
    data = None
    try:
        import requests

        resp = requests.get(
            url, timeout=_TILE_TIMEOUT,
            headers={"User-Agent": "CICADA/export-plan-de-gestion"},
        )
        if resp.status_code == 200 and resp.content:
            data = resp.content
    except Exception:            # réseau coupé, DNS, proxy… → pas de fond
        data = None
    if len(_TILE_CACHE) > 512:   # borne mémoire sur un export volumineux
        _TILE_CACHE.clear()
    _TILE_CACHE[url] = data
    return data


def _merc(lon, lat):
    """Web Mercator normalisé : (x, y) dans [0, 1], origine en haut à gauche."""
    lat = max(min(lat, 85.05112878), -85.05112878)
    s = math.sin(math.radians(lat))
    return ((lon + 180.0) / 360.0,
            0.5 - math.log((1 + s) / (1 - s)) / (4 * math.pi))


def _fit_zoom(span_x, span_y, width, height):
    """Plus grand zoom XYZ où l'emprise tient dans l'image, marges comprises."""
    usable_w, usable_h = width * (1 - 2 * _TILE_PAD), height * (1 - 2 * _TILE_PAD)
    for z in range(_TILE_ZOOM_MAX, _TILE_ZOOM_MIN - 1, -1):
        world = _TILE_SIZE * (2 ** z)
        if span_x * world <= usable_w and span_y * world <= usable_h:
            return z
    return _TILE_ZOOM_MIN


def _paste_basemap(img, left, top, zoom):
    """Colle les tuiles couvrant la fenêtre (left, top, +taille de `img`).
    Retourne True si au moins une tuile a pu être posée."""
    from PIL import Image as PILImage

    template = _tile_url_template()
    if not template:         # fond de carte desactivé par configuration
        return False
    width, height = img.size
    ntiles = 2 ** zoom
    x0, x1 = math.floor(left / _TILE_SIZE), math.floor((left + width - 1) / _TILE_SIZE)
    y0, y1 = math.floor(top / _TILE_SIZE), math.floor((top + height - 1) / _TILE_SIZE)
    if (x1 - x0 + 1) * (y1 - y0 + 1) > _TILE_MAX:
        return False
    pasted = False
    for ty in range(y0, y1 + 1):
        if not 0 <= ty < ntiles:
            continue
        for tx in range(x0, x1 + 1):
            url = template.format(z=zoom, x=tx % ntiles, y=ty)
            raw = _fetch_tile(url)
            if not raw:
                continue
            try:
                tile = PILImage.open(io.BytesIO(raw)).convert("RGB")
            except Exception:
                continue
            img.paste(tile, (int(tx * _TILE_SIZE - left), int(ty * _TILE_SIZE - top)))
            pasted = True
    return pasted


def _geom_rings(geom, out=None):
    """Anneaux [(lon, lat), …] d'une géométrie GEOS (polygones, lignes, points)."""
    out = [] if out is None else out
    if geom is None:
        return out
    gtype = (getattr(geom, "geom_type", "") or "").lower()
    try:
        if gtype == "polygon":
            for ring in geom.coords:
                out.append([(float(x), float(y)) for x, y in ring])
        elif gtype in ("multipolygon", "geometrycollection", "multilinestring", "multipoint"):
            for part in geom:
                _geom_rings(part, out)
        elif gtype == "linestring":
            out.append([(float(x), float(y)) for x, y in geom.coords])
        elif gtype == "point":
            x, y = float(geom.x), float(geom.y)
            out.append([(x, y)])
    except Exception:
        return out
    return out


def _geom_png(geom, width=_MAP_W, height=_MAP_H):
    """Carte de localisation (PNG) : emprise sur fond de carte, ou None."""
    try:
        from PIL import Image as PILImage, ImageDraw
    except ImportError:      # Pillow absent → pas de carte, le reste s'exporte
        return None

    rings = [r for r in _geom_rings(geom) if r]
    if not rings:
        return None

    proj = [[_merc(lon, lat) for lon, lat in ring] for ring in rings]
    xs = [p[0] for ring in proj for p in ring]
    ys = [p[1] for ring in proj for p in ring]
    span_x, span_y = max(xs) - min(xs), max(ys) - min(ys)
    zoom = _fit_zoom(span_x, span_y, width, height)
    world = _TILE_SIZE * (2 ** zoom)
    # fenêtre de rendu centrée sur l'emprise, exprimée en pixels « monde »
    left = (max(xs) + min(xs)) / 2 * world - width / 2
    top = (max(ys) + min(ys)) / 2 * world - height / 2

    img = PILImage.new("RGB", (width, height), "#F3EFEA")
    has_tiles = False
    try:
        has_tiles = _paste_basemap(img, left, top, zoom)
    except Exception:        # un fond de carte indisponible ne casse rien
        has_tiles = False

    # L'emprise passe en surcouche translucide pour laisser voir le fond.
    overlay = PILImage.new("RGBA", (width, height), (0, 0, 0, 0))
    odraw = ImageDraw.Draw(overlay)
    for ring in proj:
        pts = [(x * world - left, y * world - top) for x, y in ring]
        if len(pts) >= 3:
            odraw.polygon(pts, fill=(192, 227, 207, 110))
            odraw.line(pts + [pts[0]], fill=(2, 83, 89, 255), width=3, joint="curve")
        elif len(pts) == 2:
            odraw.line(pts, fill=(2, 83, 89, 255), width=3)
        else:
            x, y = pts[0]
            odraw.ellipse([x - 6, y - 6, x + 6, y + 6],
                          fill=(183, 77, 93, 255), outline=(255, 255, 255, 255), width=2)
    img = PILImage.alpha_composite(img.convert("RGBA"), overlay).convert("RGB")

    draw = ImageDraw.Draw(img)
    draw.rectangle([0, 0, width - 1, height - 1], outline="#9DB3B4")
    if has_tiles:            # attribution obligatoire du fournisseur de tuiles
        label = _tile_attribution()
        try:
            box = draw.textbbox((0, 0), label)
            tw, th = box[2] - box[0], box[3] - box[1]
        except Exception:
            tw, th = 7 * len(label), 10
        draw.rectangle([width - tw - 8, height - th - 7, width - 2, height - 2],
                       fill="#FFFFFF")
        draw.text((width - tw - 5, height - th - 5), label, fill="#343433")

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def _operation_geom(op):
    """Emprise de l'action, à défaut celle de ses sites (#410)."""
    if op.geom:
        return op.geom
    union = None
    for site in op.sites.all():
        if site.geom:
            union = site.geom if union is None else union.union(site.geom)
    return union


# ---------------------------------------------------------------------------
# Cadre de l'action (état ou pression)
# ---------------------------------------------------------------------------

def _is_reponse(ind) -> bool:
    t = getattr(ind, "type_indicateur", None)
    return bool(t and (t.mnemonique or "").upper() == "REPONSE")


def _linked_indicateurs(op):
    inds = {}
    for met in op.metriques.all():
        if met.id_indicateur_id:
            inds[met.id_indicateur_id] = met.id_indicateur
    if op.id_indicateur_id:
        inds[op.id_indicateur_id] = op.id_indicateur
    return list(inds.values())


def _cadre(op):
    """Renvoie les libellés du cadre (métriques, indicateur, NE/RA, OLT/OO, enjeu).

    #626 — les indicateurs de **réponse** (et leurs métriques) sont exclus : ils
    ont leur propre rubrique en bas de fiche et n'ont rien à faire dans la ligne
    « Indicateur d'état / de pression ». Leur chaîne NE/RA → OLT/OO → enjeu reste
    parcourue : elle porte le contexte de l'action.
    """
    metriques = [_txt(m.nom_metrique) for m in op.metriques.all()
                 if not _is_reponse(m.id_indicateur)]
    inds = _linked_indicateurs(op)
    indicateurs, exigences, objectifs, enjeux = [], [], [], []
    for ind in inds:
        if not _is_reponse(ind):
            indicateurs.append(_txt(ind.nom_indicateur))
        ne = getattr(ind, "id_ne", None)
        ra = getattr(ind, "id_resultat_attendu", None)
        if ne:
            exigences.append(_txt(ne.libelle))
            olt = getattr(ne, "id_olt", None)
            if olt:
                objectifs.append(_txt(olt.libelle))
                enj = getattr(olt, "id_enjeu", None)
                if enj:
                    enjeux.append(_txt(enj.libelle) or _txt(enj.intitule_court))
        if ra:
            exigences.append(_txt(ra.libelle))
            oo = getattr(ra, "id_oo", None)
            if oo:
                objectifs.append(_txt(oo.libelle))
                enj = getattr(oo, "id_enjeu", None)
                if enj:
                    enjeux.append(_txt(enj.libelle) or _txt(enj.intitule_court))

    def uniq(seq):
        out = []
        for s in seq:
            if s and s not in out:
                out.append(s)
        return out

    return {
        "metriques": uniq(metriques),
        "indicateurs": uniq(indicateurs),
        "exigences": uniq(exigences),
        "objectifs": uniq(objectifs),
        "enjeux": uniq(enjeux),
    }


# ---------------------------------------------------------------------------
# Volet financier
# ---------------------------------------------------------------------------



# ---------------------------------------------------------------------------
# Rendu d'une fiche
# ---------------------------------------------------------------------------

class _Writer:
    def __init__(self, ws, ncols):
        self.ws = ws
        self.ncols = ncols          # nb total de colonnes (3 + nb années)
        self.r = 1

    def section(self, text):
        ws = self.ws
        ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=self.ncols)
        c = ws.cell(self.r, 1, text)
        c.fill = PatternFill("solid", fgColor=_SECTION)
        c.font = _F_SECTION
        c.alignment = _AL_L
        for col in range(1, self.ncols + 1):
            ws.cell(self.r, col).border = _BM
        ws.row_dimensions[self.r].height = 20
        self.r += 1

    def kv(self, label, value, *, fill=_LABEL_FILL, font_label=None):
        """Ligne « label (A:C) | valeur (D:fin) »."""
        # Résolu à l'appel et non en valeur par défaut : une valeur par défaut
        # est liée à la définition, elle garderait la couleur d'origine après
        # un changement de couleur d'instance (#601).
        font_label = font_label or _F_LABEL
        ws = self.ws
        ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=3)
        lc = ws.cell(self.r, 1, label)
        lc.fill = PatternFill("solid", fgColor=fill)
        lc.font = font_label
        lc.alignment = _AL_LT
        ws.merge_cells(start_row=self.r, start_column=4, end_row=self.r, end_column=self.ncols)
        vc = ws.cell(self.r, 4, value)
        vc.font = _F_VALUE
        vc.alignment = _AL_LT
        for col in range(1, self.ncols + 1):
            ws.cell(self.r, col).border = _B
        # #626 — une valeur multi-lignes (ex. « Détails du suivi » + protocoles)
        # doit disposer de la hauteur nécessaire pour s'afficher.
        nlines = str(value).count("\n") + 1 if value else 1
        if nlines > 1:
            ws.row_dimensions[self.r].height = max(
                ws.row_dimensions[self.r].height or 15, 15 * nlines)
        self.r += 1

    def year_header(self, label, years):
        ws = self.ws
        ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=3)
        lc = ws.cell(self.r, 1, label)
        lc.fill = PatternFill("solid", fgColor=_LABEL_FILL)
        lc.font = _F_LABEL
        lc.alignment = _AL_L
        for i, y in enumerate(years):
            c = ws.cell(self.r, 4 + i, y)
            c.fill = PatternFill("solid", fgColor=_YEARHDR_FILL)
            c.font = _F_YEAR
            c.alignment = _AL_C
        for col in range(1, self.ncols + 1):
            ws.cell(self.r, col).border = _B
        self.r += 1

    def marks_row(self, label, flags, years):
        ws = self.ws
        ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=3)
        lc = ws.cell(self.r, 1, label)
        lc.fill = PatternFill("solid", fgColor=_SUBLABEL_FILL)
        lc.font = _F_SUBLABEL
        lc.alignment = _AL_L
        for i, y in enumerate(years):
            v = "x" if flags.get(y) else ""
            c = ws.cell(self.r, 4 + i, v)
            c.font = _F_X
            c.alignment = _AL_C
        for col in range(1, self.ncols + 1):
            ws.cell(self.r, col).border = _B
        self.r += 1

    def cost_row(self, label, values_by_year, years, *, fmt=_euro, fill=_SUBLABEL_FILL,
                 font=_F_SUBLABEL, label_font=None):
        ws = self.ws
        ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=3)
        lc = ws.cell(self.r, 1, label)
        lc.fill = PatternFill("solid", fgColor=fill)
        lc.font = label_font or _F_SUBLABEL
        lc.alignment = _AL_L
        for i, y in enumerate(years):
            c = ws.cell(self.r, 4 + i, fmt(values_by_year.get(y, 0)))
            c.font = font
            c.alignment = _AL_R
            c.fill = PatternFill("solid", fgColor=fill) if fill == _TOTAL_FILL else PatternFill()
        for col in range(1, self.ncols + 1):
            ws.cell(self.r, col).border = _B
        self.r += 1

    def org_banner(self, text):
        ws = self.ws
        ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=self.ncols)
        c = ws.cell(self.r, 1, text)
        c.fill = PatternFill("solid", fgColor=_PRIMARY)
        c.font = Font(name="Calibri", bold=True, size=10, color=_WHITE)
        c.alignment = _AL_L
        for col in range(1, self.ncols + 1):
            ws.cell(self.r, col).border = _B
        self.r += 1

    def sub_banner(self, text):
        """Bandeau clair pleine largeur (titre d'une grille de métrique, #626)."""
        ws = self.ws
        ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=self.ncols)
        c = ws.cell(self.r, 1, text)
        c.fill = PatternFill("solid", fgColor=_LABEL_FILL)
        c.font = _F_LABEL
        c.alignment = _AL_L
        for col in range(1, self.ncols + 1):
            ws.cell(self.r, col).border = _B
        self.r += 1

    def spans(self, n=5):
        """Découpe les colonnes 1..ncols en n plages contiguës ~égales."""
        base, extra = divmod(self.ncols, n)
        out, start = [], 1
        for i in range(n):
            size = max(1, base + (1 if i < extra else 0))
            out.append((start, min(start + size - 1, self.ncols)))
            start += size
        return out

    def score_grid(self, values):
        """Grille 5 paliers : en-tête coloré (« Très mauvais / = 1 ») + valeurs."""
        ws = self.ws
        spans = self.spans(5)
        for level, (c1, c2) in enumerate(spans, start=1):
            ws.merge_cells(start_row=self.r, start_column=c1, end_row=self.r, end_column=c2)
            cell = ws.cell(self.r, c1, f"{_SCORE_LABELS[level]}\n= {level}")
            cell.fill = PatternFill("solid", fgColor=_SCORE_FILLS[level])
            cell.font = _F_SCOREHDR
            cell.alignment = _AL_C
            for col in range(c1, c2 + 1):
                ws.cell(self.r, col).border = _B
        ws.row_dimensions[self.r].height = 30
        self.r += 1
        for level, (c1, c2) in enumerate(spans, start=1):
            ws.merge_cells(start_row=self.r, start_column=c1, end_row=self.r, end_column=c2)
            cell = ws.cell(self.r, c1, values.get(level) or "—")
            cell.font = _F_VALUE
            cell.alignment = _AL_C
            for col in range(c1, c2 + 1):
                ws.cell(self.r, col).border = _B
        ws.row_dimensions[self.r].height = 34
        self.r += 1

    def picture_row(self, label, png, *, rows=_MAP_ROWS, col=4):
        """
        Ligne « libellé (A:C) | image (D:fin) », l'image **contenue** dans sa
        cellule (#626).

        Une image Excel flotte au-dessus de la grille : elle n'est « dans » une
        case que si la case est aussi grande qu'elle. Poser la vignette sur une
        ligne de hauteur ordinaire, puis sauter des lignes vides, la laissait
        déborder sur la section suivante — d'où « la carte est *sur* la case,
        pas *dans* la case ».

        On construit donc un vrai cadre : le libellé et la zone de valeur sont
        fusionnés sur `rows` lignes, l'image est **réduite à la largeur
        disponible** (elle dépend du nombre d'années du plan) et les hauteurs de
        lignes sont calculées à partir de sa hauteur finale.
        """
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU

        ws, ancre = self.ws, self.r
        img = XLImage(io.BytesIO(png))

        # Largeur utile de la zone de valeur, en pixels. Largeur de colonne
        # Excel → pixels : ~7 px par caractère plus les 5 px de marge interne.
        zone = sum(
            round((ws.column_dimensions[get_column_letter(c)].width or 8.43) * 7) + 5
            for c in range(col, self.ncols + 1)
        )
        largeur = min(img.width, max(zone - 2 * _MAP_PAD, 60))
        hauteur = round(img.height * largeur / img.width)
        img.width, img.height = largeur, hauteur

        # Cadre : libellé à gauche, image à droite, sur la même hauteur.
        ws.merge_cells(start_row=ancre, start_column=1,
                       end_row=ancre + rows - 1, end_column=3)
        lc = ws.cell(ancre, 1, label)
        lc.fill = PatternFill("solid", fgColor=_LABEL_FILL)
        lc.font = _F_LABEL
        lc.alignment = _AL_LT
        ws.merge_cells(start_row=ancre, start_column=col,
                       end_row=ancre + rows - 1, end_column=self.ncols)
        for r in range(ancre, ancre + rows):
            for c in range(1, self.ncols + 1):
                ws.cell(r, c).border = _B

        # Hauteurs de lignes : le cadre fait exactement la hauteur de l'image,
        # marges comprises. En points (1 px = 0,75 pt à 96 ppp).
        total_px = hauteur + 2 * _MAP_PAD
        for r in range(ancre, ancre + rows):
            ws.row_dimensions[r].height = round(total_px * 0.75 / rows, 2)

        marqueur = AnchorMarker(
            col=col - 1, colOff=pixels_to_EMU(_MAP_PAD),
            row=ancre - 1, rowOff=pixels_to_EMU(_MAP_PAD),
        )
        img.anchor = OneCellAnchor(
            _from=marqueur,
            ext=XDRPositiveSize2D(pixels_to_EMU(largeur), pixels_to_EMU(hauteur)),
        )
        ws.add_image(img)
        self.r = ancre + rows

    def blank(self):
        self.r += 1


def _render_action(ws, op, years, *, is_cs, code_local=""):
    ncols = 3 + len(years)
    w = _Writer(ws, ncols)
    # largeurs
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    for i in range(len(years)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 11

    # ---- En-tête : code / intitulé / priorité ----
    ws.merge_cells(start_row=w.r, start_column=1, end_row=w.r, end_column=ncols - 2)
    hc = ws.cell(w.r, 1, _txt(op.libelle))
    hc.fill = PatternFill("solid", fgColor=_PRIMARY)
    hc.font = _F_TITLE
    hc.alignment = _AL_L
    ws.merge_cells(start_row=w.r, start_column=ncols - 1, end_row=w.r, end_column=ncols)
    prio = op.id_priorite.label if op.id_priorite_id else ""
    pc = ws.cell(w.r, ncols - 1, prio or "Priorité —")
    pc.fill = PatternFill("solid", fgColor=_TERRA)
    pc.font = _F_PRIO
    pc.alignment = _AL_C
    for col in range(1, ncols + 1):
        ws.cell(w.r, col).border = _BM
    ws.row_dimensions[w.r].height = 22
    w.r += 1
    # #626 — deux codes distincts, comme sur le modèle : le code local du plan
    # (CS1, IP2… calculé par le parcours de l'arborescence) et le code de l'action
    # dans le référentiel d'actions (Gestref / Eden 62).
    code = code_local or _txt(op.code_operation) or (
        f"n° {op.numero_manuel}" if op.numero_manuel else "")
    w.kv("Code action (local)", code)
    ta = getattr(op, "id_type_action", None)
    code_ref = " — ".join(x for x in (
        _txt(getattr(ta, "cd_nomenclature", "")), _txt(getattr(ta, "label", ""))) if x)
    w.kv("Code action (référentiel Gestref)", code_ref or _txt(op.id_referentiel_operations))

    # ---- 1) Cadre de l'action ----
    w.section("1) Cadre de l'action")
    cadre = _cadre(op)
    w.kv("Métriques", " ; ".join(cadre["metriques"]))
    w.kv("Indicateur d'état" if is_cs else "Indicateur de pression", " ; ".join(cadre["indicateurs"]))
    w.kv("Niveau d'exigence" if is_cs else "Résultats attendus", " ; ".join(cadre["exigences"]))
    w.kv("OLT" if is_cs else "Objectif opérationnel", " ; ".join(cadre["objectifs"]))
    w.kv("Enjeu", " ; ".join(cadre["enjeux"]))

    # ---- 2) Détails de l'opération ----
    w.section("2) Détails de l'opération")
    details = _txt(op.description)
    if is_cs:
        # #626 — protocole(s) intégrés TEXTUELLEMENT dans « Détails du suivi »
        # (nom, descriptif, objectifs) plutôt qu'en lignes séparées.
        blocs = [details] if details else []
        suivi = getattr(op, "id_suivi", None)
        for pr in (suivi.protocoles.all() if suivi else []):
            standardise = bool(pr.protocole_dans_campanule)
            nom = _txt(pr.protocole_campanule_nom) if standardise else _txt(pr.nom_protocole)
            nom = nom or _txt(pr.nom_protocole) or _txt(pr.protocole_campanule_nom)
            libelle = "Protocole standardisé" if standardise else "Protocole non standardisé"
            lignes = [f"{libelle} : {nom}" if nom else libelle]
            desc = _txt(pr.description_protocole)
            obj = _txt(pr.objectif_protocole)
            if desc:
                lignes.append(f"Descriptif du protocole : {desc}")
            if obj:
                lignes.append(f"Objectifs du protocole : {obj}")
            blocs.append("\n".join(lignes))
        w.kv("Détails du suivi", "\n\n".join(blocs))
    else:
        w.kv("Détails de l'action", details)
    w.kv("Opérateurs", _txt(op.operateurs))
    w.kv("Partenaires", _txt(op.partenaires))

    # ---- Localisation de l'action (#626) ----
    sites = list(op.sites.all())
    w.kv("Localisation (site(s))", " ; ".join(_txt(s.nom_site) for s in sites))
    png = None
    try:
        png = _geom_png(_operation_geom(op))
    except Exception:        # une géométrie exotique ne doit pas casser l'export
        png = None
    if png:
        # La vignette occupe le champ « valeur » de la ligne, en face de son
        # libellé (#626) — et *dans* la case, qui est dimensionnée pour elle.
        w.picture_row("Emprise de l'action", png)
    else:
        w.kv("Emprise de l'action", "Non renseignée")

    # ---- 3) Volet administratif et financier ----
    w.section("3) Détail du volet administratif et financier de l'opération")

    # #607 Q4 : la fiche action est **prévisionnelle** (actions saisies, pas le suivi).
    from .models_operations import OperationAnnee
    annees = {oa.annee: oa for oa in OperationAnnee.objects.filter(id_operation=op)}
    prev_flags = {y: bool(getattr(oa, "periodicite", False)) for y, oa in annees.items()}
    w.year_header("Programmation annuelle", years)
    w.marks_row("Périodicité", prev_flags, years)

    org_names: dict = {}
    af = build_action_finance(op, org_names, defaultdict(poste_entry_factory))

    def cost_block(title, cell_for, *, org_ventilated):
        """Rend un bloc fonctionnement/investissement/total pour un organisme
        (ou pour l'opération entière si non ventilée)."""
        w.org_banner(title)
        # Fonctionnement
        w.cost_row("Travail prévisionnel (jours) — fonctionnement", {y: cell_for(y).j_fonct for y in years}, years, fmt=_jours)
        w.cost_row("Coût salarial — fonctionnement", {y: cell_for(y).sal_fonct for y in years}, years)
        w.cost_row("Coût prestataire — fonctionnement", {y: cell_for(y).prest_fonct for y in years}, years)
        w.cost_row("Autres coûts de fonctionnement", {y: cell_for(y).autre_fonct for y in years}, years)
        w.cost_row("Total Budget de fonctionnement (€)", {y: cell_for(y).tot_fonct for y in years}, years, fill=_TOTAL_FILL, font=_F_TOTAL, label_font=_F_TOTAL)
        # Investissement
        w.cost_row("Travail prévisionnel (jours) — investissement", {y: cell_for(y).j_invest for y in years}, years, fmt=_jours)
        w.cost_row("Coût salarial — investissement", {y: cell_for(y).sal_invest for y in years}, years)
        w.cost_row("Coût prestataire — investissement", {y: cell_for(y).prest_invest for y in years}, years)
        w.cost_row("Autres coûts d'investissement", {y: cell_for(y).autre_invest for y in years}, years)
        w.cost_row("Total Budget d'investissement (€)", {y: cell_for(y).tot_invest for y in years}, years, fill=_TOTAL_FILL, font=_F_TOTAL, label_font=_F_TOTAL)
        w.cost_row("Budget total (€)", {y: cell_for(y).tot for y in years}, years, fill=_TOTAL_FILL, font=_F_TOTAL, label_font=_F_TOTAL)

    # #607 Q3 : ventilation par organisme uniquement si le mode le prévoit.
    if af.is_org_ventilated:
        real_orgs = [oid for oid in af.org_ids() if oid != 0]
        for idx, oid in enumerate(real_orgs, 1):
            cost_block(
                f"Organisme gestionnaire {idx} — {org_names.get(oid, '')}",
                lambda y, oid=oid: af.cell(oid, y),
                org_ventilated=True,
            )
        # TOTAL tous organismes
        if len(real_orgs) > 1:
            w.org_banner("TOTAL")
            w.cost_row("Travail prévisionnel (jours)", {y: af.year_total(y).jours for y in years}, years, fmt=_jours, fill=_TOTAL_FILL, font=_F_TOTAL, label_font=_F_TOTAL)
            w.cost_row("Budget de fonctionnement (€)", {y: af.year_total(y).tot_fonct for y in years}, years, fill=_TOTAL_FILL, font=_F_TOTAL, label_font=_F_TOTAL)
            w.cost_row("Budget d'investissement (€)", {y: af.year_total(y).tot_invest for y in years}, years, fill=_TOTAL_FILL, font=_F_TOTAL, label_font=_F_TOTAL)
            w.cost_row("Budget total (€)", {y: af.year_total(y).tot for y in years}, years, fill=_TOTAL_FILL, font=_F_TOTAL, label_font=_F_TOTAL)
    else:
        cost_block("Budget de l'opération", lambda y: af.year_total(y), org_ventilated=False)

    w.cost_row("Jours bénévoles / partenaires", {y: af.year_total(y).j_benevole for y in years}, years, fmt=_jours)

    # Programmation mensuelle (agrégée : mois programmé sur au moins une année)
    w.blank()
    months = ["Janv", "Fév", "Mars", "Avril", "Mai", "Juin", "Juil", "Août", "Sept", "Oct", "Nov", "Déc"]
    month_flags = _monthly_flags(op, annees)
    _render_monthly(w, months, month_flags)

    # Financeurs
    financeurs = []
    for f in op.finances.all():
        lbl = _txt(f.libelle)
        cat = _txt(getattr(getattr(f, "id_categorie", None), "label", ""))
        financeurs.append(" — ".join(x for x in (lbl, cat) if x))
    if _txt(op.financeurs):
        financeurs.append(_txt(op.financeurs))
    w.kv("Financeurs et types de financement", " ; ".join(financeurs), fill=_LABEL_FILL)

    # ---- 4) Indicateurs de réponse (#626) ----
    w.section("4) Indicateurs de réponse")
    simples, grilles = [], []
    for ind in _reponse_indicateurs(op):
        nom = _txt(ind.nom_indicateur)
        metriques = list(ind.metriques.all())
        if not metriques:
            simples.append(nom)
            continue
        for met in metriques:
            libelle = _txt(met.nom_metrique)
            if _is_grille(met):
                grilles.append((nom, met))
            else:
                # Hors grille : intitulé + métrique entre parenthèses.
                simples.append(f"{nom} ({libelle})" if libelle else nom)
    w.kv("Indicateurs de réponse", " ; ".join(dict.fromkeys(simples)), fill=_LABEL_FILL)
    for nom, met in grilles:
        libelle = _txt(met.nom_metrique)
        unite = _txt(met.unite)
        titre = f"{nom} — {libelle}" if libelle else nom
        if unite:
            titre += f" ({unite})"
        w.sub_banner(titre)
        w.score_grid({level: _grid_cell(met, level) for level in range(1, 6)})

    ws.sheet_view.showGridLines = False


def _monthly_flags(op, annees):
    flags = {m: False for m in range(1, 13)}
    default = op.programmation_mensuelle_defaut or {}
    for m, v in (default.items() if isinstance(default, dict) else []):
        if v:
            flags[int(m)] = True
    for oa in annees.values():
        pm = getattr(oa, "periodicite_mensuelle", None) or {}
        if isinstance(pm, dict):
            for m, v in pm.items():
                if v:
                    flags[int(m)] = True
    return flags


def _render_monthly(w, months, month_flags):
    ws = w.ws
    ws.merge_cells(start_row=w.r, start_column=1, end_row=w.r, end_column=3)
    lc = ws.cell(w.r, 1, "Programmation mensuelle")
    lc.fill = PatternFill("solid", fgColor=_LABEL_FILL)
    lc.font = _F_LABEL
    lc.alignment = _AL_L
    # 12 mois répartis sur les colonnes disponibles à partir de D
    for i, name in enumerate(months):
        col = 4 + i
        if col > w.ncols:
            break
        c = ws.cell(w.r, col, name)
        c.fill = PatternFill("solid", fgColor=_YEARHDR_FILL)
        c.font = _F_YEAR
        c.alignment = _AL_C
    for col in range(1, w.ncols + 1):
        ws.cell(w.r, col).border = _B
    w.r += 1
    ws.merge_cells(start_row=w.r, start_column=1, end_row=w.r, end_column=3)
    lc = ws.cell(w.r, 1, "Périodicité")
    lc.fill = PatternFill("solid", fgColor=_SUBLABEL_FILL)
    lc.font = _F_SUBLABEL
    lc.alignment = _AL_L
    for i in range(12):
        col = 4 + i
        if col > w.ncols:
            break
        c = ws.cell(w.r, col, "x" if month_flags.get(i + 1) else "")
        c.font = _F_X
        c.alignment = _AL_C
    for col in range(1, w.ncols + 1):
        ws.cell(w.r, col).border = _B
    w.r += 1


def _reponse_indicateurs(op):
    """Indicateurs de réponse liés à CETTE action (via ses métriques ou son
    indicateur direct), dédupliqués dans l'ordre de rencontre.

    #626 — on ne remonte plus aux indicateurs « frères » partageant le NE/RA :
    comme tout indicateur (réponse compris) a un parent NE/RA, cette expansion
    faisait apparaître un indicateur de réponse ajouté à une action sur TOUTES
    les fiches partageant ce NE/RA."""
    out = {}
    for ind in _linked_indicateurs(op):
        if _is_reponse(ind):
            out[ind.pk] = ind
    return list(out.values())


# ---------------------------------------------------------------------------
# Point d'entrée
# ---------------------------------------------------------------------------

def _is_cs(op) -> bool:
    cat = getattr(op, "id_categorie_action_reserve", None)
    return bool(cat and (cat.mnemonique or "").upper() == "CS")


def build_fiche_action_workbook(plan, operation_ids=None) -> bytes:
    """Construit le classeur des fiches action (un onglet par opération).

    `operation_ids` restreint l'export à certaines actions du plan (#642 :
    export d'UNE fiche depuis sa page de visualisation). Le rendu de chaque
    onglet est strictement identique à l'export complet du plan.
    """
    from .models_operations import Operation

    _appliquer_couleur_instance()

    y0 = plan.annee_debut or 0
    y1 = plan.annee_fin or y0
    years = list(range(y0, y1 + 1)) if y0 else []
    if not years:
        years = [""]

    ops = (
        Operation.objects
        .filter(metriques__id_indicateur__id_ne__id_olt__id_enjeu__id_pg=plan)
        .select_related("id_priorite", "id_categorie_action_reserve", "id_indicateur",
                        "id_suivi", "id_type_action")
        .prefetch_related(
            "metriques__id_indicateur__id_ne__id_olt__id_enjeu",
            "metriques__id_indicateur__id_resultat_attendu__id_oo",
            "metriques__id_indicateur__id_ne__indicateurs__metriques__score_blocks",
            "metriques__id_indicateur__id_resultat_attendu__indicateurs__metriques__score_blocks",
            "finances__id_categorie", "id_suivi__protocoles", "sites",
        )
        .distinct()
    )
    # + opérations rattachées directement à un indicateur (#367)
    ops_direct = (
        Operation.objects
        .filter(id_indicateur__id_ne__id_olt__id_enjeu__id_pg=plan)
        .exclude(id_operation__in=[o.id_operation for o in ops])
        .distinct()
    )
    all_ops = list(ops) + list(ops_direct)
    if operation_ids is not None:
        wanted = {int(i) for i in operation_ids}
        all_ops = [o for o in all_ops if o.id_operation in wanted]
    all_ops.sort(key=lambda o: (_txt(o.code_operation), o.id_operation))

    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    if not all_ops:
        ws = wb.create_sheet(_sanitize_title("Actions", used))
        ws["A1"] = "Ce plan ne contient pas encore d'action."
        ws["A1"].font = Font(bold=True, size=12, color=_PRIMARY)
    # #626 — code local de l'action tel qu'affiché dans le plan (CS1, IP2…),
    # calculé une fois pour tout le plan (même source que l'export budget, #618).
    from .serializers_operations import compute_operation_codes_for_plan
    try:
        codes = compute_operation_codes_for_plan(plan.pk)
    except Exception:
        codes = {}

    for op in all_ops:
        is_cs = _is_cs(op)
        code_local = codes.get(op.id_operation) or _txt(op.code_operation)
        # Onglet : code saisi s'il existe, sinon code local calculé.
        titre = _txt(op.code_operation) or code_local or f"Action {op.id_operation}"
        ws = wb.create_sheet(_sanitize_title(titre, used))
        _render_action(ws, op, years, is_cs=is_cs, code_local=code_local)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
