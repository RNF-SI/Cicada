"""
Import / export Excel de l'arborescence d'un plan de gestion (V1, sans IA).

Ce module fournit, pour l'instant, la construction du **classeur modèle**
(`.xlsx`) au format multi-onglets décrit dans la note projet :

- un onglet par niveau de l'arborescence (Enjeux, Facteurs, Pressions, OLT, NE,
  OO, RA, Indicateurs, Métriques) ;
- des **codes logiques** choisis par le rédacteur (colonne ``code``) et des
  colonnes de rattachement par code (``enjeu``, ``facteur``, ``parent``…) ;
- les liens N-N en **cellules multi-valeurs** (``E1,E3``) — partage d'un facteur
  entre plusieurs enjeux (#552), d'un OO entre plusieurs pressions ;
- un onglet ``Listes`` (masqué) alimentant les **listes déroulantes** Excel des
  colonnes de nomenclature, pour contraindre la saisie dans le fichier (#478).

Le classeur peut être **vide** (modèle neuf) ou **pré-rempli** à partir d'un
plan existant — dans ce cas il sert aussi d'export / sauvegarde et de point de
départ pour dériver un autre plan.

Le module fournit également le **moteur d'import** :

- ``parse_workbook`` lit le fichier et renvoie les lignes par onglet ;
- ``validate_import`` produit un rapport d'anomalies (dry-run, sans écriture) ;
- ``execute_import`` crée l'arborescence dans un plan en brouillon (transaction).

L'import est en **création seule** : il refuse un plan contenant déjà des enjeux.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Callable, Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from apps.core.models import Nomenclature

# Version du format de fichier — incrémentée si la structure des onglets change.
# Écrite dans l'onglet « Lisez-moi » et (à terme) vérifiée à l'import.
FORMAT_VERSION = "1.0"


# ---------------------------------------------------------------------------
# Définition déclarative des colonnes et des onglets
# ---------------------------------------------------------------------------


@dataclass
class Column:
    """Une colonne d'un onglet.

    ``key``       identifiant interne (sert au pré-remplissage et à l'import) ;
    ``header``    en-tête visible ;
    ``required``  colonne obligatoire (en-tête mise en évidence) ;
    ``help``      texte affiché en commentaire de l'en-tête ;
    ``nomenclature`` mnémonique du *type* de nomenclature → liste déroulante ;
    ``boolean``   colonne Oui/Non → liste déroulante ;
    ``multi``     cellule multi-valeurs (codes séparés par des virgules) ;
    ``ref``       clé de l'onglet référencé → liste déroulante des codes de cet
                  onglet (rattachement, ex : la colonne « enjeux » d'un facteur
                  propose tous les codes de l'onglet Enjeux) ;
    ``width``     largeur de colonne ;
    ``value``     fonction ``obj -> valeur`` pour le pré-remplissage.
    """

    key: str
    header: str
    required: bool = False
    help: str = ""
    nomenclature: Optional[str] = None
    boolean: bool = False
    multi: bool = False
    ref: Optional[str] = None
    vocab: Optional[str] = None  # vocabulaire « maison » (ecolo / socio) → dropdown
    width: int = 24
    value: Optional[Callable] = None


@dataclass
class Sheet:
    key: str  # identifiant interne (parsing / import)
    name: str
    description: str
    columns: list[Column]
    rows: Callable = None  # plan -> iterable d'objets (pré-remplissage)


# ---------------------------------------------------------------------------
# Helpers de pré-remplissage
# ---------------------------------------------------------------------------


def _nom_label(nomenclature) -> str:
    """Label lisible d'une FK nomenclature (ou chaîne vide)."""
    if nomenclature is None:
        return ""
    return nomenclature.label or nomenclature.mnemonique or ""


def _oui_non(value) -> str:
    if value is None:
        return ""
    return "Oui" if value else "Non"


def _enjeu_types_ecologiques(enjeu) -> str:
    mapping = [
        (enjeu.habitat, "Habitat"),
        (enjeu.espece, "Espèce"),
        (enjeu.patrimoine_geologique, "Patrimoine géologique"),
        (enjeu.fonctionnalite_ecosysteme, "Fonctionnalité écosystème"),
        (enjeu.autre_ecologique, "Autre"),
    ]
    return ",".join(label for flag, label in mapping if flag)


def _enjeu_types_socioeco(enjeu) -> str:
    mapping = [
        (enjeu.valeur_paysagere, "Valeur paysagère"),
        (enjeu.patrimoine_culturel, "Patrimoine culturel"),
        (enjeu.developpement_durable, "Développement durable"),
        (enjeu.usages, "Usages"),
        (enjeu.valeur_ajoutee, "Valeur ajoutée"),
        (enjeu.autre_socioeco, "Autre"),
    ]
    return ",".join(label for flag, label in mapping if flag)


# Vocabulaires contrôlés « maison » (pas des nomenclatures en base).
TYPES_ECOLOGIQUES = [
    "Habitat",
    "Espèce",
    "Patrimoine géologique",
    "Fonctionnalité écosystème",
    "Autre",
]
TYPES_SOCIOECO = [
    "Valeur paysagère",
    "Patrimoine culturel",
    "Développement durable",
    "Usages",
    "Valeur ajoutée",
    "Autre",
]


# ---------------------------------------------------------------------------
# Construction du modèle de codes pour le pré-remplissage
# ---------------------------------------------------------------------------


class _CodeAllocator:
    """Attribue des codes logiques stables et lisibles aux objets d'un plan.

    Les codes sont séquentiels par type d'entité (E1, F1, P1, O1, N1, OO1, R1,
    I1, M1…). Ils ne portent pas la hiérarchie : le rattachement est exprimé par
    des colonnes dédiées référençant le code du parent.
    """

    PREFIXES = {
        "enjeu": "E",
        "facteur": "F",
        "pression": "P",
        "olt": "O",
        "ne": "N",
        "oo": "OO",
        "ra": "R",
        "indicateur": "I",
        "metrique": "M",
    }

    def __init__(self):
        self._counters: dict[str, int] = {}
        self._codes: dict[tuple[str, int], str] = {}

    def code(self, kind: str, pk: int) -> str:
        existing = self._codes.get((kind, pk))
        if existing:
            return existing
        self._counters[kind] = self._counters.get(kind, 0) + 1
        code = f"{self.PREFIXES[kind]}{self._counters[kind]}"
        self._codes[(kind, pk)] = code
        return code


# ---------------------------------------------------------------------------
# Extraction à plat d'un plan existant (pré-remplissage)
# ---------------------------------------------------------------------------


@dataclass
class _PlanRows:
    """Lignes extraites d'un plan, prêtes à écrire, une liste par onglet."""

    enjeux: list = field(default_factory=list)
    facteurs: list = field(default_factory=list)
    pressions: list = field(default_factory=list)
    olt: list = field(default_factory=list)
    ne: list = field(default_factory=list)
    oo: list = field(default_factory=list)
    ra: list = field(default_factory=list)
    indicateurs: list = field(default_factory=list)
    metriques: list = field(default_factory=list)
    taxons: list = field(default_factory=list)
    habitats: list = field(default_factory=list)


def _extract_plan(plan) -> _PlanRows:
    """Parcourt l'arborescence d'un plan et produit des dicts par onglet.

    Chaque dict a les clés attendues par les ``Column.key`` de l'onglet
    correspondant. Les facteurs et OO partagés (#552) ne sont émis qu'une fois.
    """
    alloc = _CodeAllocator()
    rows = _PlanRows()

    enjeux = list(plan.enjeux.all().order_by("ordre", "id_enjeu"))

    for enjeu in enjeux:
        e_code = alloc.code("enjeu", enjeu.id_enjeu)
        rows.enjeux.append(
            {
                "code": e_code,
                "categorie": _nom_label(enjeu.id_categorie),
                "categorie_fcr": _nom_label(enjeu.id_categorie_fcr),
                "importance": _nom_label(enjeu.id_importance),
                "rang": enjeu.rang if enjeu.rang is not None else "",
                "libelle": enjeu.libelle,
                "intitule_court": enjeu.intitule_court or "",
                "categorie_ecologique": _oui_non(enjeu.categorie_ecologique),
                "types_ecologiques": _enjeu_types_ecologiques(enjeu),
                "types_socioeco": _enjeu_types_socioeco(enjeu),
                "etat_enjeu": enjeu.etat_enjeu or "",
                "description": enjeu.description or "",
            }
        )
        _emit_bio(rows, e_code, enjeu)

        # --- Branche OLT → NE (vision « état ») ---
        for olt in enjeu.objectifs_long_terme.all().order_by("ordre", "id_olt"):
            o_code = alloc.code("olt", olt.id_olt)
            rows.olt.append(
                {
                    "code": o_code,
                    "enjeu": e_code,
                    "libelle": olt.libelle,
                    "numero_manuel": olt.numero_manuel
                    if olt.numero_manuel is not None
                    else "",
                    "description": olt.description or "",
                }
            )
            for ne in olt.niveaux_exigence.all().order_by("ordre", "id_ne"):
                n_code = alloc.code("ne", ne.id_ne)
                rows.ne.append(
                    {
                        "code": n_code,
                        "olt": o_code,
                        "libelle": ne.libelle,
                        "description": ne.description or "",
                    }
                )
                for ind in ne.indicateurs.all().order_by("ordre", "id_indicateur"):
                    _emit_indicateur(alloc, rows, ind, parent_code=n_code)

    # --- Branche opérationnelle : facteurs / pressions / OO / RA ---
    # Parcourue via les enjeux pour garantir un ordre déterministe, mais les
    # facteurs et OO partagés ne sont émis qu'une fois.
    seen_facteurs: set[int] = set()
    seen_oo: set[int] = set()

    for enjeu in enjeux:
        for cor in enjeu.cor_facteurs.all().order_by("ordre", "id"):
            facteur = cor.id_facteur_influence
            f_code = alloc.code("facteur", facteur.id_facteur_influence)
            if facteur.id_facteur_influence not in seen_facteurs:
                seen_facteurs.add(facteur.id_facteur_influence)
                enjeux_codes = ",".join(
                    alloc.code("enjeu", c.id_enjeu_id)
                    for c in facteur.cor_enjeux.all().order_by("ordre", "id")
                )
                rows.facteurs.append(
                    {
                        "code": f_code,
                        "libelle": facteur.libelle,
                        "enjeux": enjeux_codes,
                        "description": facteur.description or "",
                    }
                )
                for pr in facteur.pressions.all().order_by("ordre", "id_pression"):
                    p_code = alloc.code("pression", pr.id_pression)
                    rows.pressions.append(
                        {
                            "code": p_code,
                            "facteur": f_code,
                            "libelle": pr.libelle,
                            "type_pression": _nom_label(pr.id_type_pression),
                            "description": pr.description or "",
                        }
                    )
                    for oo in pr.objectifs_operationnels.all().order_by(
                        "ordre", "id_oo"
                    ):
                        _emit_oo(alloc, rows, oo, seen_oo)

    # --- OO directs (FCR #337) ---
    for enjeu in enjeux:
        for oo in enjeu.objectifs_operationnels_directs.all().order_by(
            "ordre", "id_oo"
        ):
            _emit_oo(
                alloc, rows, oo, seen_oo, enjeu_code=alloc.code("enjeu", enjeu.id_enjeu)
            )

    return rows


def _emit_oo(alloc, rows, oo, seen_oo, enjeu_code: str = "") -> None:
    if oo.id_oo in seen_oo:
        return
    seen_oo.add(oo.id_oo)
    oo_code = alloc.code("oo", oo.id_oo)
    pressions_codes = ",".join(
        alloc.code("pression", pr.id_pression)
        for pr in oo.pressions.all().order_by("ordre", "id_pression")
    )
    rows.oo.append(
        {
            "code": oo_code,
            "pressions": pressions_codes,
            "enjeu": enjeu_code,
            "libelle": oo.libelle,
            "numero_manuel": oo.numero_manuel if oo.numero_manuel is not None else "",
            "description": oo.description or "",
        }
    )
    for ra in oo.resultats_attendus.all().order_by("ordre", "id_ra"):
        ra_code = alloc.code("ra", ra.id_ra)
        rows.ra.append(
            {
                "code": ra_code,
                "oo": oo_code,
                "libelle": ra.libelle,
                "description": ra.description or "",
            }
        )
        for ind in ra.indicateurs.all().order_by("ordre", "id_indicateur"):
            _emit_indicateur(alloc, rows, ind, parent_code=ra_code)


def _emit_bio(rows, cible_code: str, obj) -> None:
    """Émet les taxons et habitats rattachés à un enjeu ou un indicateur."""
    for taxon in obj.taxons.all().order_by("cd_nom"):
        rows.taxons.append(
            {
                "cible": cible_code,
                "cd_nom": taxon.cd_nom,
                "nom": taxon.nom_complet or taxon.nom_vern or "",
            }
        )
    for habitat in obj.habitats.all().order_by("cd_hab"):
        rows.habitats.append(
            {
                "cible": cible_code,
                "cd_hab": habitat.cd_hab or "",
                "nom": habitat.lb_hab_fr or "",
            }
        )


def _emit_indicateur(alloc, rows, ind, parent_code: str) -> None:
    i_code = alloc.code("indicateur", ind.id_indicateur)
    rows.indicateurs.append(
        {
            "code": i_code,
            "parent": parent_code,
            "type": _nom_label(ind.type_indicateur),
            "nom_indicateur": ind.nom_indicateur,
            "description": ind.description or "",
        }
    )
    _emit_bio(rows, i_code, ind)
    for met in ind.metriques.all().order_by("ordre", "id_metrique"):
        m_code = alloc.code("metrique", met.id_metrique)
        rows.metriques.append(
            {
                "code": m_code,
                "indicateur": i_code,
                "nom_metrique": met.nom_metrique,
                "type_metrique": _nom_label(met.type_metrique),
                "unite": met.unite or "",
                "description": met.description or "",
            }
        )


# ---------------------------------------------------------------------------
# Schéma des onglets
# ---------------------------------------------------------------------------


EXAMPLE_PLAN_NAME = "Réserve naturelle d'une zone humide (exemple)"


def _example_plan_rows() -> "_PlanRows":
    """Contenu d'exemple, cohérent et complet, illustrant les liens entre onglets.

    Thème fictif : réserve de zone humide tourbeuse. Montre notamment un facteur
    partagé entre deux enjeux (F1 → E1,E3 ; F3 → E1,E2), les deux branches de
    l'arborescence, un FCR (OO4 rattaché directement à l'enjeu E4) et des
    taxons/habitats sur enjeux et indicateurs.
    """
    return _PlanRows(
        enjeux=[
            {
                "code": "E1",
                "categorie": "Enjeu de conservation",
                "importance": "Priorité 1",
                "rang": 1,
                "libelle": "Habitats tourbeux et prairies humides",
                "intitule_court": "Habitats tourbeux",
                "categorie_ecologique": "Oui",
                "types_ecologiques": "Habitat",
                "etat_enjeu": "État de conservation variable : tourbière active bien conservée, marge en voie d'assèchement.",
                "description": "Préserver et restaurer les habitats tourbeux et prairies humides de la réserve.",
            },
            {
                "code": "E2",
                "categorie": "Enjeu de conservation",
                "importance": "Priorité 1",
                "rang": 1,
                "libelle": "Avifaune paludicole nicheuse",
                "intitule_court": "Avifaune paludicole",
                "categorie_ecologique": "Oui",
                "types_ecologiques": "Espèce",
                "etat_enjeu": "Population de passereaux paludicoles fragile, dépendante des roselières.",
                "description": "Conserver des populations nicheuses viables d'oiseaux des roselières.",
            },
            {
                "code": "E3",
                "categorie": "Enjeu de conservation",
                "importance": "Priorité 2",
                "rang": 2,
                "libelle": "Fonctionnalité hydrologique de la zone humide",
                "intitule_court": "Hydrologie",
                "categorie_ecologique": "Oui",
                "types_ecologiques": "Fonctionnalité écosystème",
                "etat_enjeu": "Alimentation en eau perturbée par d'anciens drainages.",
                "description": "Restaurer un fonctionnement hydrologique naturel de la zone humide.",
            },
            {
                "code": "E4",
                "categorie": "Facteur Clé de Réussite",
                "categorie_fcr": "Ancrage territorial",
                "importance": "Priorité 2",
                "rang": 2,
                "libelle": "Ancrage territorial et adhésion des acteurs",
                "intitule_court": "Ancrage territorial",
                "categorie_ecologique": "Non",
                "types_socioeco": "Usages,Développement durable",
                "etat_enjeu": "Relations établies avec les communes et exploitants, à consolider.",
                "description": "Impliquer durablement les acteurs du territoire dans la gestion de la réserve.",
            },
        ],
        facteurs=[
            {
                "code": "F1",
                "libelle": "Gestion des niveaux d'eau",
                "enjeux": "E1,E3",  # facteur partagé entre deux enjeux (#552)
                "description": "La maîtrise des niveaux d'eau conditionne à la fois les habitats tourbeux et l'hydrologie.",
            },
            {
                "code": "F2",
                "libelle": "Fréquentation et dérangement",
                "enjeux": "E2",
                "description": "La fréquentation du public peut déranger l'avifaune nicheuse.",
            },
            {
                "code": "F3",
                "libelle": "Espèces exotiques envahissantes",
                "enjeux": "E1,E2",  # facteur partagé
                "description": "La Jussie et l'écrevisse de Californie dégradent habitats et ressources.",
            },
        ],
        pressions=[
            {
                "code": "P1",
                "facteur": "F1",
                "libelle": "Assèchement estival",
                "type_pression": "Modifications hydrologiques",
                "description": "Baisse des niveaux d'eau en été, défavorable aux sphaignes.",
            },
            {
                "code": "P2",
                "facteur": "F1",
                "libelle": "Atterrissement et comblement",
                "type_pression": "Modification des apports en matériel organique dans le milieu",
                "description": "Accumulation de matière organique et fermeture des milieux ouverts.",
            },
            {
                "code": "P3",
                "facteur": "F2",
                "libelle": "Dérangement en période de nidification",
                "type_pression": "Dérangement des espèces sauvages",
                "description": "Passages hors sentiers au printemps sur les zones de nidification.",
            },
            {
                "code": "P4",
                "facteur": "F3",
                "libelle": "Colonisation par la Jussie",
                "type_pression": "Introduction ou propagation d'espèces non indigènes-exotiques",
                "description": "Progression de la Jussie dans les gouilles et fossés.",
            },
        ],
        olt=[
            {
                "code": "O1",
                "enjeu": "E1",
                "libelle": "Maintenir les habitats tourbeux en bon état de conservation",
                "description": "Vision à long terme : une tourbière active fonctionnelle.",
            },
            {
                "code": "O2",
                "enjeu": "E2",
                "libelle": "Conserver des populations nicheuses viables de passereaux paludicoles",
                "description": "Des roselières accueillantes et une reproduction régulière.",
            },
            {
                "code": "O3",
                "enjeu": "E3",
                "libelle": "Restaurer un régime hydrologique naturel",
                "description": "Des niveaux d'eau proches du fonctionnement de référence.",
            },
        ],
        ne=[
            {
                "code": "N1",
                "olt": "O1",
                "libelle": "Au moins 80 % de la surface de tourbière en bon état",
                "description": "Seuil d'exigence sur l'état de conservation des habitats.",
            },
            {
                "code": "N2",
                "olt": "O2",
                "libelle": "Au moins 5 espèces paludicoles nicheuses chaque année",
                "description": "Diversité minimale attendue de l'avifaune nicheuse.",
            },
            {
                "code": "N3",
                "olt": "O3",
                "libelle": "Niveau d'eau estival maintenu au-dessus de -20 cm/sol",
                "description": "Seuil piézométrique compatible avec les tourbières.",
            },
        ],
        oo=[
            {
                "code": "OO1",
                "pressions": "P1,P2",
                "libelle": "Restaurer un régime hydraulique favorable",
                "description": "Agir sur les ouvrages et le règlement d'eau.",
            },
            {
                "code": "OO2",
                "pressions": "P3",
                "libelle": "Maîtriser la fréquentation en période sensible",
                "description": "Canaliser le public et créer des zones de quiétude.",
            },
            {
                "code": "OO3",
                "pressions": "P4",
                "libelle": "Contenir la Jussie",
                "description": "Limiter puis réduire l'emprise de la Jussie.",
            },
            {
                "code": "OO4",
                "enjeu": "E4",  # FCR : rattaché directement à l'enjeu
                "libelle": "Animer un comité local de concertation",
                "description": "Réunir élus, agriculteurs et usagers autour du plan de gestion.",
            },
        ],
        ra=[
            {
                "code": "R1",
                "oo": "OO1",
                "libelle": "Niveaux d'eau conformes au règlement d'eau",
                "description": "Les niveaux respectent la consigne saisonnière.",
            },
            {
                "code": "R2",
                "oo": "OO1",
                "libelle": "Ouvrages hydrauliques fonctionnels et entretenus",
                "description": "Vannes et seuils opérationnels toute l'année.",
            },
            {
                "code": "R3",
                "oo": "OO2",
                "libelle": "Zones de quiétude respectées au printemps",
                "description": "Absence de pénétration sur les secteurs sensibles.",
            },
            {
                "code": "R4",
                "oo": "OO3",
                "libelle": "Surface de Jussie réduite de 50 %",
                "description": "Réduction mesurée de l'emprise sur cinq ans.",
            },
            {
                "code": "R5",
                "oo": "OO4",
                "libelle": "Un comité de concertation réuni au moins une fois par an",
                "description": "Gouvernance locale active et régulière.",
            },
        ],
        indicateurs=[
            {
                "code": "I1",
                "parent": "N1",
                "type": "État",
                "nom_indicateur": "État de conservation des habitats tourbeux",
                "description": "Évaluation périodique de l'état des habitats.",
            },
            {
                "code": "I2",
                "parent": "N2",
                "type": "État",
                "nom_indicateur": "Richesse en passereaux paludicoles nicheurs",
                "description": "Nombre d'espèces paludicoles nicheuses.",
            },
            {
                "code": "I3",
                "parent": "N3",
                "type": "État",
                "nom_indicateur": "Niveau piézométrique estival",
                "description": "Suivi du niveau de la nappe en été.",
            },
            {
                "code": "I4",
                "parent": "R1",
                "type": "Réponse",
                "nom_indicateur": "Conformité des niveaux d'eau au règlement",
                "description": "Écart aux consignes du règlement d'eau.",
            },
            {
                "code": "I5",
                "parent": "R3",
                "type": "Pression",
                "nom_indicateur": "Dérangement observé sur les zones de quiétude",
                "description": "Événements de dérangement recensés au printemps.",
            },
            {
                "code": "I6",
                "parent": "R4",
                "type": "Réponse",
                "nom_indicateur": "Surface colonisée par la Jussie",
                "description": "Emprise de la Jussie mesurée chaque année.",
            },
            {
                "code": "I7",
                "parent": "R5",
                "type": "Réponse",
                "nom_indicateur": "Nombre de réunions du comité local",
                "description": "Fréquence des réunions de concertation.",
            },
        ],
        metriques=[
            {
                "code": "M1",
                "indicateur": "I1",
                "nom_metrique": "Surface en bon état de conservation",
                "unite": "ha",
            },
            {
                "code": "M2",
                "indicateur": "I2",
                "nom_metrique": "Nombre d'espèces paludicoles nicheuses",
                "unite": "espèces",
            },
            {
                "code": "M3",
                "indicateur": "I3",
                "nom_metrique": "Niveau piézométrique moyen estival",
                "unite": "cm",
            },
            {
                "code": "M4",
                "indicateur": "I4",
                "nom_metrique": "Jours hors consigne du règlement d'eau",
                "unite": "jours",
            },
            {
                "code": "M5",
                "indicateur": "I5",
                "nom_metrique": "Événements de dérangement recensés",
                "unite": "évén./an",
            },
            {
                "code": "M6",
                "indicateur": "I6",
                "nom_metrique": "Surface de Jussie",
                "unite": "m²",
            },
            {
                "code": "M7",
                "indicateur": "I7",
                "nom_metrique": "Réunions du comité par an",
                "unite": "réunions/an",
            },
        ],
        taxons=[
            {"cible": "E1", "cd_nom": 104398, "nom": "Drosera rotundifolia"},
            {"cible": "E2", "cd_nom": 2878, "nom": "Circus aeruginosus"},
            {"cible": "I2", "cd_nom": 4187, "nom": "Acrocephalus schoenobaenus"},
        ],
        habitats=[
            {"cible": "E1", "cd_hab": "7110", "nom": "Tourbières hautes actives"},
            {"cible": "E1", "cd_hab": "7230", "nom": "Tourbières basses alcalines"},
            {"cible": "E3", "cd_hab": "3150", "nom": "Lacs eutrophes naturels"},
        ],
    )


def _build_schema() -> list[Sheet]:
    """Définit les onglets et leurs colonnes (structure du format V1)."""
    return [
        Sheet(
            key="enjeux",
            name="Enjeux",
            description="Un enjeu (ou FCR) par ligne. Le code sert à rattacher facteurs, OLT et OO.",
            rows=lambda rows: rows.enjeux,
            columns=[
                Column(
                    "code",
                    "code",
                    required=True,
                    width=10,
                    help="Identifiant libre et unique de l'enjeu (ex : E1). "
                    "Réutilisé dans les autres onglets pour le rattachement.",
                ),
                Column(
                    "categorie",
                    "catégorie",
                    required=True,
                    nomenclature="CATEGORIE_ENJEU",
                    help="Enjeu écologique/socio-éco ou FCR (facteur clé de réussite).",
                ),
                Column(
                    "categorie_fcr",
                    "catégorie FCR",
                    nomenclature="CATEGORIE_FCR",
                    help="À renseigner uniquement pour les FCR.",
                ),
                Column("importance", "importance", nomenclature="IMPORTANCE_ENJEU"),
                Column(
                    "rang",
                    "rang (priorité)",
                    width=14,
                    help="Priorité de l'enjeu (1 à 3). Facultatif.",
                ),
                Column(
                    "libelle",
                    "libellé",
                    required=True,
                    width=45,
                    help="Libellé de l'enjeu. Unique au sein du plan.",
                ),
                Column(
                    "intitule_court",
                    "intitulé court",
                    width=20,
                    help="25 caractères maximum.",
                ),
                Column(
                    "categorie_ecologique",
                    "écologique ?",
                    boolean=True,
                    width=14,
                    help="Oui = enjeu de conservation du patrimoine naturel ; "
                    "Non = enjeu socio-économique.",
                ),
                Column(
                    "types_ecologiques",
                    "types écologiques",
                    multi=True,
                    vocab="ecolo",
                    width=30,
                    help="Un ou plusieurs, séparés par des virgules : "
                    + ", ".join(TYPES_ECOLOGIQUES),
                ),
                Column(
                    "types_socioeco",
                    "types socio-éco",
                    multi=True,
                    vocab="socio",
                    width=30,
                    help="Un ou plusieurs, séparés par des virgules : "
                    + ", ".join(TYPES_SOCIOECO),
                ),
                Column("etat_enjeu", "état de l'enjeu", width=40),
                Column("description", "description", width=40),
            ],
        ),
        Sheet(
            key="facteurs",
            name="Facteurs",
            description="Facteurs d'influence. Un facteur peut être partagé entre plusieurs enjeux.",
            rows=lambda rows: rows.facteurs,
            columns=[
                Column(
                    "code",
                    "code",
                    required=True,
                    width=10,
                    help="Identifiant libre et unique du facteur (ex : F1).",
                ),
                Column("libelle", "libellé", required=True, width=45),
                Column(
                    "enjeux",
                    "enjeux",
                    required=True,
                    multi=True,
                    ref="enjeux",
                    width=20,
                    help="Code(s) d'enjeu, séparés par des virgules pour un facteur "
                    "partagé (ex : E1,E3).",
                ),
                Column("description", "description", width=40),
            ],
        ),
        Sheet(
            key="pressions",
            name="Pressions",
            description="Pressions rattachées à un facteur d'influence.",
            rows=lambda rows: rows.pressions,
            columns=[
                Column(
                    "code",
                    "code",
                    required=True,
                    width=10,
                    help="Identifiant libre et unique de la pression (ex : P1).",
                ),
                Column(
                    "facteur",
                    "facteur",
                    required=True,
                    ref="facteurs",
                    width=12,
                    help="Code du facteur parent (ex : F1).",
                ),
                Column("libelle", "libellé", required=True, width=45),
                Column(
                    "type_pression",
                    "type de pression",
                    nomenclature="TYPE_PRESSION",
                    width=30,
                    help="Type PressRef (facultatif).",
                ),
                Column("description", "description", width=40),
            ],
        ),
        Sheet(
            key="olt",
            name="OLT",
            description="Objectifs à long terme, rattachés à un enjeu (vision « état »).",
            rows=lambda rows: rows.olt,
            columns=[
                Column(
                    "code",
                    "code",
                    required=True,
                    width=10,
                    help="Identifiant libre et unique de l'OLT (ex : O1).",
                ),
                Column(
                    "enjeu",
                    "enjeu",
                    required=True,
                    ref="enjeux",
                    width=12,
                    help="Code de l'enjeu parent (ex : E1).",
                ),
                Column("libelle", "libellé", required=True, width=45),
                Column(
                    "numero_manuel",
                    "numéro manuel",
                    width=14,
                    help="Numéro imposé (facultatif). Vide = numérotation automatique.",
                ),
                Column("description", "description", width=40),
            ],
        ),
        Sheet(
            key="ne",
            name="NE",
            description="Niveaux d'exigence, rattachés à un OLT.",
            rows=lambda rows: rows.ne,
            columns=[
                Column(
                    "code",
                    "code",
                    required=True,
                    width=10,
                    help="Identifiant libre et unique du niveau d'exigence (ex : N1).",
                ),
                Column(
                    "olt",
                    "OLT",
                    required=True,
                    ref="olt",
                    width=12,
                    help="Code de l'OLT parent (ex : O1).",
                ),
                Column("libelle", "libellé", required=True, width=45),
                Column("description", "description", width=40),
            ],
        ),
        Sheet(
            key="oo",
            name="OO",
            description="Objectifs opérationnels : rattachés à une/des pression(s), ou "
            "directement à un FCR.",
            rows=lambda rows: rows.oo,
            columns=[
                Column(
                    "code",
                    "code",
                    required=True,
                    width=10,
                    help="Identifiant libre et unique de l'OO (ex : OO1).",
                ),
                Column(
                    "pressions",
                    "pressions",
                    multi=True,
                    ref="pressions",
                    width=20,
                    help="Code(s) de pression, séparés par des virgules (ex : P1,P2). "
                    "Laisser vide si l'OO est rattaché à un FCR.",
                ),
                Column(
                    "enjeu",
                    "enjeu (FCR)",
                    ref="enjeux",
                    width=14,
                    help="Code de l'enjeu FCR parent, uniquement pour un OO sans pression.",
                ),
                Column("libelle", "libellé", required=True, width=45),
                Column(
                    "numero_manuel",
                    "numéro manuel",
                    width=14,
                    help="Numéro imposé (facultatif).",
                ),
                Column("description", "description", width=40),
            ],
        ),
        Sheet(
            key="ra",
            name="RA",
            description="Résultats attendus, rattachés à un objectif opérationnel.",
            rows=lambda rows: rows.ra,
            columns=[
                Column(
                    "code",
                    "code",
                    required=True,
                    width=10,
                    help="Identifiant libre et unique du résultat attendu (ex : R1).",
                ),
                Column(
                    "oo",
                    "OO",
                    required=True,
                    ref="oo",
                    width=12,
                    help="Code de l'OO parent (ex : OO1).",
                ),
                Column("libelle", "libellé", required=True, width=45),
                Column("description", "description", width=40),
            ],
        ),
        Sheet(
            key="indicateurs",
            name="Indicateurs",
            description="Indicateurs. Rattachés à un niveau d'exigence (état) OU à un "
            "résultat attendu (pression/réponse).",
            rows=lambda rows: rows.indicateurs,
            columns=[
                Column(
                    "code",
                    "code",
                    required=True,
                    width=10,
                    help="Identifiant libre et unique de l'indicateur (ex : I1).",
                ),
                Column(
                    "parent",
                    "parent (NE ou RA)",
                    required=True,
                    width=18,
                    help="Code d'un niveau d'exigence (ex : N1) OU d'un résultat "
                    "attendu (ex : R1) — un seul des deux.",
                ),
                Column(
                    "type",
                    "type",
                    nomenclature="TYPE_INDICATEUR",
                    width=16,
                    help="État, Pression ou Réponse.",
                ),
                Column("nom_indicateur", "nom", required=True, width=45),
                Column("description", "description", width=40),
            ],
        ),
        Sheet(
            key="metriques",
            name="Metriques",
            description="Métriques d'un indicateur. En V1, seul le libellé est importé : "
            "l'indicateur reste « indéterminé » tant qu'aucune grille n'est saisie.",
            rows=lambda rows: rows.metriques,
            columns=[
                Column(
                    "code",
                    "code",
                    required=True,
                    width=10,
                    help="Identifiant libre et unique de la métrique (ex : M1).",
                ),
                Column(
                    "indicateur",
                    "indicateur",
                    required=True,
                    ref="indicateurs",
                    width=14,
                    help="Code de l'indicateur parent (ex : I1).",
                ),
                Column("nom_metrique", "nom", required=True, width=45),
                Column("type_metrique", "type", nomenclature="TYPE_METRIQUE", width=16),
                Column("unite", "unité", width=16),
                Column("description", "description", width=40),
            ],
        ),
        Sheet(
            key="taxons",
            name="Taxons",
            description="Espèces (TaxRef) rattachées à un enjeu ou à un indicateur, "
            "via le code de la cible.",
            rows=lambda rows: rows.taxons,
            columns=[
                Column(
                    "cible",
                    "cible (enjeu ou indicateur)",
                    required=True,
                    width=22,
                    help="Code d'un enjeu (ex : E1) OU d'un indicateur (ex : I1).",
                ),
                Column(
                    "cd_nom",
                    "cd_nom",
                    required=True,
                    width=14,
                    help="Code TaxRef de l'espèce (nombre entier).",
                ),
                Column("nom", "nom", width=45, help="Nom de l'espèce (facultatif)."),
            ],
        ),
        Sheet(
            key="habitats",
            name="Habitats",
            description="Habitats (HabRef) rattachés à un enjeu ou à un indicateur, "
            "via le code de la cible.",
            rows=lambda rows: rows.habitats,
            columns=[
                Column(
                    "cible",
                    "cible (enjeu ou indicateur)",
                    required=True,
                    width=22,
                    help="Code d'un enjeu (ex : E1) OU d'un indicateur (ex : I1).",
                ),
                Column(
                    "cd_hab",
                    "cd_hab",
                    required=True,
                    width=14,
                    help="Code HabRef de l'habitat.",
                ),
                Column(
                    "nom", "nom", width=45, help="Libellé de l'habitat (facultatif)."
                ),
            ],
        ),
    ]


# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

_PRIMARY = "025359"
_TERRA = "B74D5D"
_HEADER_FILL = PatternFill("solid", fgColor=_PRIMARY)
_REQUIRED_FILL = PatternFill("solid", fgColor=_TERRA)
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=_PRIMARY)
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP_TOP = Alignment(vertical="top", wrap_text=True)

# Ligne « exemple » du modèle vide : stylée en gris italique, jamais importée.
_HINT_FONT = Font(name="Calibri", italic=True, color="9A8F86", size=10)
_HINT_FILL = PatternFill("solid", fgColor="F7F3EE")
# Marqueur placé dans la 1re colonne (code / cible) de la ligne exemple. Toute
# ligne dont la 1re colonne commence par ce marqueur est ignorée à l'import.
_EXAMPLE_MARKER = "(exemple)"


# ---------------------------------------------------------------------------
# Construction du classeur
# ---------------------------------------------------------------------------


def _load_nomenclature_values() -> dict[str, list[str]]:
    """Retourne, par type de nomenclature, la liste des labels actifs."""
    types = {
        col.nomenclature
        for sheet in _build_schema()
        for col in sheet.columns
        if col.nomenclature
    }
    values: dict[str, list[str]] = {}
    for type_mnemo in types:
        labels = list(
            Nomenclature.objects.filter(id_type__mnemonique=type_mnemo, actif=True)
            .order_by("id_nomenclature")
            .values_list("label", flat=True)
        )
        values[type_mnemo] = [lbl for lbl in labels if lbl]
    return values


def _write_lisez_moi(wb: Workbook, plan, example_name=None, with_hints=False) -> None:
    ws = wb.create_sheet("Lisez-moi", 0)
    ws.sheet_properties.tabColor = _PRIMARY
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110

    lines = [
        ("Import de l'arborescence d'un plan de gestion", _TITLE_FONT),
        (f"Format version {FORMAT_VERSION}", Font(italic=True, color="746F6E")),
        ("", None),
        ("Comment remplir ce fichier", Font(bold=True, size=12, color=_PRIMARY)),
        (
            "• Un onglet par niveau de l'arborescence : Enjeux, Facteurs, Pressions, "
            "OLT, NE, OO, RA, Indicateurs, Métriques.",
            None,
        ),
        (
            "• Chaque ligne a un code unique (colonne « code ») que vous choisissez "
            "librement (ex : E1, F1, P1…).",
            None,
        ),
        (
            "• Pour rattacher une ligne à son parent, reportez le code du parent dans "
            "la colonne prévue (ex : la pression P1 a « F1 » dans sa colonne facteur).",
            None,
        ),
        (
            "• Les colonnes de rattachement multiples (enjeux d'un facteur, pressions "
            "d'un OO) acceptent plusieurs codes séparés par des virgules : E1,E3.",
            None,
        ),
        (
            "• Les colonnes avec une liste déroulante n'acceptent que les valeurs "
            "proposées (catégorie, type, importance…).",
            None,
        ),
        (
            "• Les colonnes de rattachement proposent aussi une liste : le code "
            "d'un parent se choisit dans la liste des codes de l'onglet concerné "
            "(ex : la colonne « enjeux » d'un facteur propose tous les enjeux). "
            "Cette liste se met à jour au fur et à mesure que vous ajoutez des "
            "lignes dans l'onglet référencé.",
            None,
        ),
        (
            "• Les colonnes « types écologiques » et « types socio-éco » proposent "
            "aussi une liste : choisissez un type, ou saisissez-en plusieurs "
            "séparés par des virgules (ex : Habitat,Espèce).",
            None,
        ),
        (
            "• Astuce codes : tapez « E1 » dans la première ligne puis faites "
            "glisser la poignée de recopie (petit carré en bas à droite de la "
            "cellule) vers le bas — Excel incrémente automatiquement E2, E3, E4…",
            None,
        ),
        (
            "• Les en-têtes en rouge sont obligatoires ; survolez un en-tête pour "
            "afficher son aide.",
            None,
        ),
        ("", None),
        ("Structure de l'arborescence", Font(bold=True, size=12, color=_PRIMARY)),
        ("Enjeu → OLT → Niveau d'exigence → Indicateur (état) → Métrique", None),
        (
            "Enjeu → Facteur → Pression → Objectif opérationnel → Résultat attendu "
            "→ Indicateur (pression/réponse) → Métrique",
            None,
        ),
        (
            "Un FCR relie directement l'enjeu à ses objectifs opérationnels (sans "
            "facteur ni pression).",
            None,
        ),
        ("", None),
        ("À noter", Font(bold=True, size=12, color=_PRIMARY)),
        (
            "• En V1, une métrique n'importe que son libellé : l'indicateur reste "
            "« indéterminé » (aucune grille de scoring).",
            None,
        ),
        (
            "• L'import remplit un plan de gestion en brouillon ne contenant pas "
            "encore d'arborescence.",
            None,
        ),
    ]
    if with_hints:
        lines += [
            ("", None),
            (
                "• La première ligne grisée de chaque onglet est un EXEMPLE (elle "
                "commence par « (exemple) ») : elle montre quoi écrire et n'est "
                "jamais importée. Vous pouvez la laisser, la remplacer ou la "
                "supprimer ; saisissez vos données sur les lignes suivantes.",
                Font(italic=True, color="9A8F86"),
            ),
        ]
    if plan is not None:
        lines += [
            ("", None),
            (
                f"Ce fichier a été pré-rempli à partir du plan : {plan.nom}",
                Font(italic=True, color="746F6E"),
            ),
        ]
    if example_name is not None:
        lines += [
            ("", None),
            (
                f"⚠ EXEMPLE PÉDAGOGIQUE FICTIF — {example_name}. "
                "Ce fichier illustre le format et les liens entre onglets ; "
                "remplacez son contenu par le vôtre avant de l'importer.",
                Font(bold=True, italic=True, color="B74D5D"),
            ),
        ]

    for i, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=2, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.sheet_view.showGridLines = False


def _write_listes(
    wb: Workbook, nomenclature_values: dict[str, list[str]]
) -> dict[str, str]:
    """Écrit l'onglet masqué « Listes » et retourne, par clé de liste, la
    référence de plage Excel absolue (ex : 'Listes'!$A$2:$A$8)."""
    ws = wb.create_sheet("Listes")
    ranges: dict[str, str] = {}

    col_idx = 1
    # Nomenclatures
    for type_mnemo, labels in sorted(nomenclature_values.items()):
        letter = get_column_letter(col_idx)
        ws.cell(row=1, column=col_idx, value=type_mnemo).font = Font(bold=True)
        for r, label in enumerate(labels, start=2):
            ws.cell(row=r, column=col_idx, value=label)
        if labels:
            ranges[
                f"nom:{type_mnemo}"
            ] = f"'Listes'!${letter}$2:${letter}${len(labels) + 1}"
        ws.column_dimensions[letter].width = 28
        col_idx += 1

    # Vocabulaires maison (types écologiques / socio-éco) : source des listes
    # déroulantes (aide à la saisie, non bloquante car colonnes multi-valeurs).
    for vocab_key, name, vocab in (
        ("ecolo", "Types écologiques", TYPES_ECOLOGIQUES),
        ("socio", "Types socio-éco", TYPES_SOCIOECO),
    ):
        letter = get_column_letter(col_idx)
        ws.cell(row=1, column=col_idx, value=name).font = Font(bold=True)
        for r, label in enumerate(vocab, start=2):
            ws.cell(row=r, column=col_idx, value=label)
        if vocab:
            ranges[
                f"vocab:{vocab_key}"
            ] = f"'Listes'!${letter}$2:${letter}${len(vocab) + 1}"
        ws.column_dimensions[letter].width = 28
        col_idx += 1

    ws.sheet_state = "hidden"
    ws.protection.sheet = True  # onglet de référence : non modifiable
    return ranges


# Nombre de lignes vierges (avec validation) dans un modèle vide.
_BLANK_ROWS = 100


def _sheet_code_ranges(schema: list[Sheet], plan_rows) -> dict[str, str]:
    """Retourne, par clé d'onglet, la plage Excel absolue de sa colonne « code ».

    Sert de source aux listes déroulantes de rattachement (ex : la colonne
    « enjeux » d'un facteur propose les codes de l'onglet Enjeux). La plage
    couvre les lignes pré-remplies **et** la zone de saisie vierge, pour que les
    codes ajoutés par l'utilisateur apparaissent dans les listes.
    """
    ranges: dict[str, str] = {}
    for sheet in schema:
        data_rows = (
            sheet.rows(plan_rows) if (plan_rows is not None and sheet.rows) else []
        )
        last_row = 2 + max(len(data_rows), 0) + _BLANK_ROWS
        # La colonne « code » est toujours la première (A), les données à L3.
        ranges[sheet.key] = f"'{sheet.name}'!$A$3:$A${last_row}"
    return ranges


def _write_sheet(
    wb: Workbook, sheet: Sheet, plan_rows, list_ranges, code_ranges=None, hint_row=None
) -> None:
    ws = wb.create_sheet(sheet.name)

    # Ligne 1 : description de l'onglet.
    desc = ws.cell(row=1, column=1, value=sheet.description)
    desc.font = Font(italic=True, color="746F6E", size=10)
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=len(sheet.columns)
    )

    header_row = 2
    for c, col in enumerate(sheet.columns, start=1):
        cell = ws.cell(row=header_row, column=c, value=col.header)
        cell.font = _HEADER_FONT
        cell.fill = _REQUIRED_FILL if col.required else _HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = _BORDER
        if col.help:
            note = col.help + (" (obligatoire)" if col.required else "")
            cell.comment = Comment(note, "CICADA")
            cell.comment.width = 260
            cell.comment.height = 120
        ws.column_dimensions[get_column_letter(c)].width = col.width

    # Données (pré-remplissage) ou lignes vierges.
    data_rows = sheet.rows(plan_rows) if (plan_rows is not None and sheet.rows) else []
    first_data = header_row + 1
    r = first_data

    # Ligne « exemple » (modèle vide uniquement) : stylée à part, jamais importée.
    n_hint = 0
    if hint_row and not data_rows:
        for c, col in enumerate(sheet.columns, start=1):
            cell = ws.cell(row=r, column=c, value=hint_row.get(col.key, ""))
            cell.alignment = _WRAP_TOP
            cell.border = _BORDER
            cell.font = _HINT_FONT
            cell.fill = _HINT_FILL
        r += 1
        n_hint = 1

    for row in data_rows:
        for c, col in enumerate(sheet.columns, start=1):
            cell = ws.cell(row=r, column=c, value=row.get(col.key, ""))
            cell.alignment = _WRAP_TOP
            cell.border = _BORDER
        r += 1

    # Validation (listes déroulantes) sur toute la zone de saisie.
    n_rows = n_hint + max(len(data_rows), 0) + _BLANK_ROWS
    last_row = header_row + n_rows
    code_ranges = code_ranges or {}
    for c, col in enumerate(sheet.columns, start=1):
        letter = get_column_letter(c)
        dv = None
        strict = True  # la liste rejette une valeur hors liste
        multi_example = "E1,E3"  # exemple de saisie multiple (aide non bloquante)
        if col.boolean:
            dv = DataValidation(type="list", formula1='"Oui,Non"', allow_blank=True)
        elif col.nomenclature and not col.multi:
            ref = list_ranges.get(f"nom:{col.nomenclature}")
            if ref:
                dv = DataValidation(type="list", formula1=ref, allow_blank=True)
        elif col.ref:
            ref = code_ranges.get(col.ref)
            if ref:
                dv = DataValidation(type="list", formula1=ref, allow_blank=True)
                # Colonnes multi-valeurs (E1,E3) : la liste est une aide à la
                # saisie, elle ne doit pas bloquer une saisie multiple.
                strict = not col.multi
        elif col.vocab:
            ref = list_ranges.get(f"vocab:{col.vocab}")
            if ref:
                dv = DataValidation(type="list", formula1=ref, allow_blank=True)
                strict = False  # types multiples possibles → aide non bloquante
                multi_example = "Habitat,Espèce"
        if dv is not None:
            if strict:
                dv.error = "Choisissez une valeur dans la liste proposée."
                dv.errorTitle = "Valeur non autorisée"
                dv.showErrorMessage = True
            else:
                dv.showErrorMessage = False
                dv.promptTitle = "Sélection multiple"
                dv.prompt = (
                    "Choisissez une valeur dans la liste, ou saisissez-en "
                    f"plusieurs séparées par des virgules (ex : {multi_example})."
                )
                dv.showInputMessage = True
            ws.add_data_validation(dv)
            dv.add(f"{letter}{first_data}:{letter}{last_row}")

    ws.freeze_panes = f"A{first_data}"
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(sheet.columns))}{header_row}"
    )
    ws.sheet_view.showGridLines = False


def _example_hint_rows() -> dict[str, dict]:
    """Une ligne « exemple » par onglet, pour le modèle vide.

    Reprend la 1re ligne du contenu d'exemple et remplace la 1re colonne (code
    ou cible) par le marqueur ``(exemple)``, afin que la ligne soit ignorée à
    l'import tout en montrant, sur les autres colonnes, ce qu'il faut écrire et
    comment rattacher (colonnes de rattachement renseignées avec des codes).
    """
    example = _example_plan_rows()
    hints: dict[str, dict] = {}
    for sheet in _build_schema():
        rows = sheet.rows(example) if sheet.rows else []
        if not rows:
            continue
        hint = dict(rows[0])
        hint[sheet.columns[0].key] = _EXAMPLE_MARKER
        hints[sheet.key] = hint
    return hints


def _render_workbook(
    plan_rows, plan=None, example_name=None, with_hints=False
) -> bytes:
    """Assemble le classeur à partir de lignes déjà extraites (ou ``None``).

    :param plan_rows: un ``_PlanRows`` (pré-rempli/exemple) ou ``None`` (vide).
    :param plan: plan source, pour la mention de l'onglet « Lisez-moi ».
    :param example_name: nom d'exemple à mentionner dans « Lisez-moi ».
    :param with_hints: ajoute une ligne « exemple » (modèle vide uniquement).
    """
    wb = Workbook()
    # Retirer la feuille par défaut ; les onglets sont créés explicitement.
    wb.remove(wb.active)

    nomenclature_values = _load_nomenclature_values()
    schema = _build_schema()
    code_ranges = _sheet_code_ranges(schema, plan_rows)
    hints = _example_hint_rows() if with_hints else {}

    _write_lisez_moi(wb, plan, example_name=example_name, with_hints=with_hints)
    list_ranges = _write_listes(wb, nomenclature_values)
    for sheet in schema:
        _write_sheet(
            wb, sheet, plan_rows, list_ranges, code_ranges, hints.get(sheet.key)
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_arborescence_workbook(plan=None) -> bytes:
    """Construit le classeur modèle d'import d'arborescence.

    :param plan: si fourni, le classeur est pré-rempli avec l'arborescence du
        plan ; sinon un modèle vide (avec ligne exemple) est produit.
    :returns: le contenu binaire du fichier ``.xlsx``.
    """
    plan_rows = _extract_plan(plan) if plan is not None else None
    # Ligne exemple tant que le plan n'a pas d'arborescence (modèle vierge ET
    # export d'un plan encore vide) ; jamais sur un plan déjà rempli.
    with_hints = plan_rows is None or not plan_rows.enjeux
    return _render_workbook(plan_rows, plan=plan, with_hints=with_hints)


def build_example_workbook() -> bytes:
    """Construit un classeur **exemple**, entièrement pré-rempli et cohérent.

    Contenu pédagogique fictif (réserve de zone humide) illustrant tous les
    onglets et surtout les **liens entre onglets** : un facteur partagé entre
    deux enjeux (#552), les deux branches (état via OLT/NE, pression via
    Facteur/Pression/OO/RA), un FCR relié directement à un enjeu, des taxons et
    habitats. Indépendant de la base : sert de référence téléchargeable.
    """
    return _render_workbook(_example_plan_rows(), example_name=EXAMPLE_PLAN_NAME)


# ===========================================================================
# Import : parsing, validation (dry-run) et exécution
# ===========================================================================

import re as _re
import unicodedata as _ud

from django.db import transaction

from openpyxl import load_workbook

from .models_enjeux import (
    Enjeu,
    FacteurInfluence,
    CorFacteurEnjeu,
    Pression,
    ObjectifLongTerme,
    NiveauExigence,
    ObjectifOperationnel,
    CorOoPression,
    ResultatAttendu,
    CorEnjeuTaxon,
    CorEnjeuHabitat,
)
from .models_indicateurs import (
    Indicateur,
    Metrique,
    CorIndicateurTaxon,
    CorIndicateurHabitat,
)


ERROR = "error"
WARNING = "warning"

# Première ligne de données (ligne 1 = description, ligne 2 = en-têtes).
_HEADER_ROW = 2
_FIRST_DATA_ROW = 3


def _norm(value) -> str:
    """Normalise une chaîne pour comparaison : sans accents, minuscule, espaces
    compressés. Sert au rapprochement des en-têtes et des labels de nomenclature."""
    if value is None:
        return ""
    text = str(value).strip()
    text = _ud.normalize("NFKD", text)
    text = "".join(c for c in text if not _ud.combining(c))
    text = _re.sub(r"\s+", " ", text)
    return text.lower()


def _cell_str(value) -> str:
    """Valeur de cellule en chaîne nettoyée (gère les nombres, None)."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_example_row(value) -> bool:
    """Vrai si la 1re colonne porte le marqueur « (exemple) » du modèle vide."""
    s = _cell_str(value).lower().lstrip("(").strip()
    return s.startswith("exemple")


def _split_multi(value) -> list[str]:
    """Découpe une cellule multi-valeurs (codes séparés par des virgules)."""
    raw = _cell_str(value)
    if not raw:
        return []
    parts = _re.split(r"[,;]", raw)
    return [p.strip() for p in parts if p.strip()]


class _NomenclatureResolver:
    """Rapproche un label saisi de la nomenclature correspondante (par type)."""

    def __init__(self):
        self._cache: dict[str, dict[str, object]] = {}

    def _index(self, type_mnemo: str) -> dict[str, object]:
        if type_mnemo not in self._cache:
            index: dict[str, object] = {}
            qs = Nomenclature.objects.filter(id_type__mnemonique=type_mnemo, actif=True)
            for nom in qs:
                for key in (nom.label, nom.mnemonique, nom.cd_nomenclature):
                    if key:
                        index.setdefault(_norm(key), nom)
            self._cache[type_mnemo] = index
        return self._cache[type_mnemo]

    def resolve(self, type_mnemo: str, label: str):
        return self._index(type_mnemo).get(_norm(label))


# Rapprochement label → champ booléen de l'Enjeu (colonnes multi-valeurs).
_ECOLO_FIELD = {
    _norm("Habitat"): "habitat",
    _norm("Espèce"): "espece",
    _norm("Patrimoine géologique"): "patrimoine_geologique",
    _norm("Fonctionnalité écosystème"): "fonctionnalite_ecosysteme",
    _norm("Autre"): "autre_ecologique",
}
_SOCIO_FIELD = {
    _norm("Valeur paysagère"): "valeur_paysagere",
    _norm("Patrimoine culturel"): "patrimoine_culturel",
    _norm("Développement durable"): "developpement_durable",
    _norm("Usages"): "usages",
    _norm("Valeur ajoutée"): "valeur_ajoutee",
    _norm("Autre"): "autre_socioeco",
}


class ArborescenceImportError(Exception):
    """Erreur de haut niveau (fichier illisible, onglet manquant…)."""


class ImportReport:
    """Rapport de validation : anomalies par onglet/ligne + décompte à créer."""

    def __init__(self):
        self.issues: list[dict] = []
        self.summary: dict[str, int] = {}

    def add(self, sheet, row, column, level, message):
        self.issues.append(
            {
                "sheet": sheet,
                "row": row,
                "column": column,
                "level": level,
                "message": message,
            }
        )

    @property
    def errors(self):
        return [i for i in self.issues if i["level"] == ERROR]

    @property
    def warnings(self):
        return [i for i in self.issues if i["level"] == WARNING]

    @property
    def can_import(self):
        return not self.errors

    def as_dict(self):
        return {
            "can_import": self.can_import,
            "n_errors": len(self.errors),
            "n_warnings": len(self.warnings),
            "issues": self.issues,
            "summary": self.summary,
        }


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------


def parse_workbook(source) -> dict[str, list[dict]]:
    """Lit le classeur et renvoie, par clé d'onglet, la liste des lignes.

    ``source`` : bytes ou objet fichier. Chaque ligne est un dict
    ``{col_key: valeur_brute, "_row": numéro_de_ligne_excel}``. Les lignes
    entièrement vides sont ignorées.

    Lève ``ArborescenceImportError`` si le fichier est illisible ou si un onglet attendu
    (ou une colonne obligatoire) est absent.
    """
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    elif hasattr(source, "read"):
        source = io.BytesIO(source.read())
    try:
        wb = load_workbook(source, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ArborescenceImportError(
            "Le fichier n'a pas pu être lu. Vérifiez qu'il s'agit bien d'un "
            "classeur Excel (.xlsx) issu du modèle d'import."
        ) from exc

    # Index des feuilles par nom normalisé.
    ws_by_name = {_norm(name): wb[name] for name in wb.sheetnames}

    parsed: dict[str, list[dict]] = {}
    for sheet in _build_schema():
        ws = ws_by_name.get(_norm(sheet.name))
        if ws is None:
            raise ArborescenceImportError(
                f"Onglet « {sheet.name} » introuvable dans le fichier."
            )

        # Rapprochement en-tête → clé de colonne.
        header_to_key = {_norm(col.header): col.key for col in sheet.columns}
        col_index_to_key: dict[int, str] = {}
        header_cells = next(
            ws.iter_rows(min_row=_HEADER_ROW, max_row=_HEADER_ROW, values_only=True),
            (),
        )
        for idx, cell in enumerate(header_cells):
            key = header_to_key.get(_norm(cell))
            if key:
                col_index_to_key[idx] = key

        # Colonnes obligatoires présentes ?
        present_keys = set(col_index_to_key.values())
        for col in sheet.columns:
            if col.required and col.key not in present_keys:
                raise ArborescenceImportError(
                    f"Colonne obligatoire « {col.header} » absente de l'onglet "
                    f"« {sheet.name} »."
                )

        first_key = sheet.columns[0].key  # code (ou cible) : porte le marqueur exemple
        rows: list[dict] = []
        for r, values in enumerate(
            ws.iter_rows(min_row=_FIRST_DATA_ROW, values_only=True),
            start=_FIRST_DATA_ROW,
        ):
            record = {}
            has_value = False
            for idx, key in col_index_to_key.items():
                value = values[idx] if idx < len(values) else None
                if _cell_str(value):
                    has_value = True
                record[key] = value
            if not has_value:
                continue
            # Ligne « exemple » du modèle : jamais importée.
            if _is_example_row(record.get(first_key)):
                continue
            record["_row"] = r
            rows.append(record)
        parsed[sheet.key] = rows

    return parsed


# ---------------------------------------------------------------------------
# Validation (dry-run)
# ---------------------------------------------------------------------------


def _as_int(value):
    text = _cell_str(value)
    if not text:
        return None
    try:
        return int(float(text.replace(",", ".")))
    except (ValueError, TypeError):
        return None


def _parse_bool(value):
    text = _norm(value)
    if text in ("oui", "vrai", "true", "1", "x"):
        return True
    if text in ("non", "faux", "false", "0"):
        return False
    return None


def _resolve_bio_refs(parsed: dict[str, list[dict]]) -> None:
    """Valide/corrige taxons et habitats contre TaxRef/HabRef (in place).

    Ajoute à chaque ligne ``_cd_nom`` / ``_cd_hab`` (code retenu, ou None) et
    ``_bio`` (statut : ``ok``, ``resolved`` = corrigé par le nom / le code
    typologique, ``unknown`` = code fourni introuvable, ``ambiguous`` = plusieurs
    correspondances, ``missing`` = ni code ni nom). Idempotent.

    **Défensif** : si le référentiel n'est pas chargé en base (dev/test sans
    import INPN), la validation est neutralisée (tout ``ok``) — on ne peut pas
    valider contre un référentiel absent.
    """
    from django.db.models import Q

    from apps.taxonomy.models import Taxref
    from apps.habitats.models import Habref

    taxons = parsed.get("taxons", [])
    habitats = parsed.get("habitats", [])
    if not taxons and not habitats:
        return
    if any("_bio" in r for r in taxons) or any("_bio" in r for r in habitats):
        return  # déjà résolu

    # --- Taxons : cd_nom ∈ TaxRef, sinon résolution par nom (unique) ---
    taxref_loaded = Taxref.objects.exists()
    cd_noms = {_as_int(r.get("cd_nom")) for r in taxons}
    cd_noms.discard(None)
    known = (
        set(Taxref.objects.filter(cd_nom__in=cd_noms).values_list("cd_nom", flat=True))
        if (taxref_loaded and cd_noms)
        else set()
    )
    names = {_cell_str(r.get("nom")) for r in taxons if _cell_str(r.get("nom"))}
    by_name: dict[str, set] = {}
    if taxref_loaded and names:
        for cd, lb, nc in Taxref.objects.filter(
            Q(lb_nom__in=names) | Q(nom_complet__in=names)
        ).values_list("cd_nom", "lb_nom", "nom_complet"):
            for label in (lb, nc):
                if label:
                    by_name.setdefault(_norm(label), set()).add(cd)
    for r in taxons:
        cd = _as_int(r.get("cd_nom"))
        name = _norm(_cell_str(r.get("nom")))
        if not taxref_loaded:
            r["_cd_nom"], r["_bio"] = cd, "ok"
        elif cd is not None and cd in known:
            r["_cd_nom"], r["_bio"] = cd, "ok"
        else:
            cands = by_name.get(name, set()) if name else set()
            if len(cands) == 1:
                r["_cd_nom"], r["_bio"] = next(iter(cands)), "resolved"
            elif len(cands) > 1:
                r["_cd_nom"], r["_bio"] = None, "ambiguous"
            elif cd is not None:
                r["_cd_nom"], r["_bio"] = None, "unknown"
            else:
                r["_cd_nom"], r["_bio"] = None, "missing"

    # --- Habitats : cd_hab (entier HabRef) ∈ HabRef, sinon résolution par le
    #     code typologique (lb_code, ex : « 7110 ») ou le libellé (lb_hab_fr) ---
    habref_loaded = Habref.objects.exists()
    hab_ints = {_as_int(r.get("cd_hab")) for r in habitats}
    hab_ints.discard(None)
    known_hab = (
        set(Habref.objects.filter(cd_hab__in=hab_ints).values_list("cd_hab", flat=True))
        if (habref_loaded and hab_ints)
        else set()
    )
    codes = {_cell_str(r.get("cd_hab")) for r in habitats if _cell_str(r.get("cd_hab"))}
    hab_names = {_cell_str(r.get("nom")) for r in habitats if _cell_str(r.get("nom"))}
    hab_by_key: dict[str, set] = {}
    if habref_loaded and (codes or hab_names):
        for cd, code, lbfr in Habref.objects.filter(
            Q(lb_code__in=codes) | Q(lb_hab_fr__in=hab_names)
        ).values_list("cd_hab", "lb_code", "lb_hab_fr"):
            for label in (code, lbfr):
                if label:
                    hab_by_key.setdefault(_norm(label), set()).add(cd)
    for r in habitats:
        raw = _cell_str(r.get("cd_hab"))
        cd = _as_int(raw)
        key = _norm(raw) or _norm(_cell_str(r.get("nom")))
        if not habref_loaded:
            r["_cd_hab"], r["_bio"] = raw or None, "ok"
        elif cd is not None and cd in known_hab:
            r["_cd_hab"], r["_bio"] = str(cd), "ok"
        else:
            cands = hab_by_key.get(_norm(raw), set()) or hab_by_key.get(
                _norm(_cell_str(r.get("nom"))), set()
            )
            if len(cands) == 1:
                r["_cd_hab"], r["_bio"] = str(next(iter(cands))), "resolved"
            elif len(cands) > 1:
                r["_cd_hab"], r["_bio"] = None, "ambiguous"
            elif raw:
                r["_cd_hab"], r["_bio"] = None, "unknown"
            else:
                r["_cd_hab"], r["_bio"] = None, "missing"


def validate_import(plan, parsed: dict[str, list[dict]]) -> ImportReport:
    """Valide les lignes parsées et produit un rapport (sans rien écrire)."""
    report = ImportReport()
    resolver = _NomenclatureResolver()

    # Création seule : refuser si le plan a déjà une arborescence.
    if plan.enjeux.exists():
        report.add(
            None,
            None,
            None,
            ERROR,
            "Ce plan contient déjà une arborescence. L'import n'est possible "
            "que sur un plan en brouillon sans enjeux.",
        )

    # --- Collecte des codes par onglet (unicité) ---
    code_sets: dict[str, set[str]] = {}
    for key in (
        "enjeux",
        "facteurs",
        "pressions",
        "olt",
        "ne",
        "oo",
        "ra",
        "indicateurs",
        "metriques",
    ):
        codes: set[str] = set()
        for row in parsed.get(key, []):
            code = _cell_str(row.get("code"))
            sheet_name = _sheet_name(key)
            if not code:
                report.add(sheet_name, row["_row"], "code", ERROR, "Code manquant.")
                continue
            if code in codes:
                report.add(
                    sheet_name,
                    row["_row"],
                    "code",
                    ERROR,
                    f"Code « {code} » en double dans l'onglet.",
                )
            codes.add(code)
        code_sets[key] = codes

    def _req(sheet_key, row, col, label):
        if not _cell_str(row.get(col)):
            report.add(
                _sheet_name(sheet_key),
                row["_row"],
                col,
                ERROR,
                f"« {label} » est obligatoire.",
            )
            return False
        return True

    def _ref(sheet_key, row, col, target_key, label, required=True):
        value = _cell_str(row.get(col))
        if not value:
            if required:
                report.add(
                    _sheet_name(sheet_key),
                    row["_row"],
                    col,
                    ERROR,
                    f"« {label} » est obligatoire.",
                )
            return
        if value not in code_sets.get(target_key, set()):
            report.add(
                _sheet_name(sheet_key),
                row["_row"],
                col,
                ERROR,
                f"Code « {value} » introuvable dans l'onglet "
                f"{_sheet_name(target_key)}.",
            )

    def _nomenclature(sheet_key, row, col, type_mnemo, label, required=False):
        value = _cell_str(row.get(col))
        if not value:
            if required:
                report.add(
                    _sheet_name(sheet_key),
                    row["_row"],
                    col,
                    ERROR,
                    f"« {label} » est obligatoire.",
                )
            return
        if resolver.resolve(type_mnemo, value) is None:
            report.add(
                _sheet_name(sheet_key),
                row["_row"],
                col,
                ERROR,
                f"Valeur « {value} » non reconnue pour « {label} ».",
            )

    # --- Enjeux ---
    libelles = {}
    for row in parsed.get("enjeux", []):
        _req("enjeux", row, "libelle", "libellé")
        _nomenclature(
            "enjeux", row, "categorie", "CATEGORIE_ENJEU", "catégorie", required=True
        )
        _nomenclature("enjeux", row, "categorie_fcr", "CATEGORIE_FCR", "catégorie FCR")
        _nomenclature("enjeux", row, "importance", "IMPORTANCE_ENJEU", "importance")
        libelle = _cell_str(row.get("libelle"))
        if libelle:
            if _norm(libelle) in libelles:
                report.add(
                    "Enjeux",
                    row["_row"],
                    "libelle",
                    ERROR,
                    f"Libellé « {libelle} » en double (déjà ligne "
                    f"{libelles[_norm(libelle)]}).",
                )
            libelles[_norm(libelle)] = row["_row"]
        court = _cell_str(row.get("intitule_court"))
        if len(court) > 25:
            report.add(
                "Enjeux",
                row["_row"],
                "intitule_court",
                ERROR,
                "L'intitulé court dépasse 25 caractères.",
            )
        rang = _cell_str(row.get("rang"))
        if rang and _as_int(rang) is None:
            report.add(
                "Enjeux",
                row["_row"],
                "rang",
                ERROR,
                "Le rang doit être un nombre (1 à 3).",
            )
        for col, vocab in (
            ("types_ecologiques", _ECOLO_FIELD),
            ("types_socioeco", _SOCIO_FIELD),
        ):
            for token in _split_multi(row.get(col)):
                if _norm(token) not in vocab:
                    report.add(
                        "Enjeux",
                        row["_row"],
                        col,
                        WARNING,
                        f"Type « {token} » non reconnu, ignoré.",
                    )

    # --- Facteurs ---
    for row in parsed.get("facteurs", []):
        _req("facteurs", row, "libelle", "libellé")
        codes = _split_multi(row.get("enjeux"))
        if not codes:
            report.add(
                "Facteurs",
                row["_row"],
                "enjeux",
                ERROR,
                "Au moins un enjeu est requis.",
            )
        for c in codes:
            if c not in code_sets["enjeux"]:
                report.add(
                    "Facteurs",
                    row["_row"],
                    "enjeux",
                    ERROR,
                    f"Enjeu « {c} » introuvable.",
                )

    # --- Pressions ---
    for row in parsed.get("pressions", []):
        _req("pressions", row, "libelle", "libellé")
        _ref("pressions", row, "facteur", "facteurs", "facteur")
        _nomenclature(
            "pressions", row, "type_pression", "TYPE_PRESSION", "type de pression"
        )

    # --- OLT ---
    for row in parsed.get("olt", []):
        _req("olt", row, "libelle", "libellé")
        _ref("olt", row, "enjeu", "enjeux", "enjeu")

    # --- NE ---
    for row in parsed.get("ne", []):
        _req("ne", row, "libelle", "libellé")
        _ref("ne", row, "olt", "olt", "OLT")

    # --- OO ---
    for row in parsed.get("oo", []):
        _req("oo", row, "libelle", "libellé")
        pressions = _split_multi(row.get("pressions"))
        enjeu = _cell_str(row.get("enjeu"))
        if not pressions and not enjeu:
            report.add(
                "OO",
                row["_row"],
                "pressions",
                ERROR,
                "Un OO doit être rattaché à au moins une pression, ou à "
                "un enjeu (cas FCR).",
            )
        if pressions and enjeu:
            report.add(
                "OO",
                row["_row"],
                "enjeu",
                WARNING,
                "OO rattaché à la fois à des pressions et à un enjeu : "
                "le rattachement direct à l'enjeu sera ignoré.",
            )
        for c in pressions:
            if c not in code_sets["pressions"]:
                report.add(
                    "OO",
                    row["_row"],
                    "pressions",
                    ERROR,
                    f"Pression « {c} » introuvable.",
                )
        if enjeu and not pressions and enjeu not in code_sets["enjeux"]:
            report.add(
                "OO", row["_row"], "enjeu", ERROR, f"Enjeu « {enjeu} » introuvable."
            )

    # --- RA ---
    for row in parsed.get("ra", []):
        _req("ra", row, "libelle", "libellé")
        _ref("ra", row, "oo", "oo", "OO")

    # --- Indicateurs ---
    for row in parsed.get("indicateurs", []):
        _req("indicateurs", row, "nom_indicateur", "nom")
        parent = _cell_str(row.get("parent"))
        if not parent:
            report.add(
                "Indicateurs",
                row["_row"],
                "parent",
                ERROR,
                "Le parent (NE ou RA) est obligatoire.",
            )
        else:
            in_ne = parent in code_sets["ne"]
            in_ra = parent in code_sets["ra"]
            if in_ne and in_ra:
                report.add(
                    "Indicateurs",
                    row["_row"],
                    "parent",
                    ERROR,
                    f"Le code « {parent} » existe à la fois comme NE et "
                    "comme RA : rendez-le unique.",
                )
            elif not in_ne and not in_ra:
                report.add(
                    "Indicateurs",
                    row["_row"],
                    "parent",
                    ERROR,
                    f"Parent « {parent} » introuvable (ni NE ni RA).",
                )
        _nomenclature("indicateurs", row, "type", "TYPE_INDICATEUR", "type")

    # --- Métriques ---
    for row in parsed.get("metriques", []):
        _req("metriques", row, "nom_metrique", "nom")
        _ref("metriques", row, "indicateur", "indicateurs", "indicateur")
        _nomenclature("metriques", row, "type_metrique", "TYPE_METRIQUE", "type")

    # --- Taxons / Habitats (rattachés à un enjeu OU un indicateur) ---
    def _check_cible(sheet_key, row):
        value = _cell_str(row.get("cible"))
        if not value:
            report.add(
                _sheet_name(sheet_key),
                row["_row"],
                "cible",
                ERROR,
                "La cible (enjeu ou indicateur) est obligatoire.",
            )
            return
        in_e = value in code_sets.get("enjeux", set())
        in_i = value in code_sets.get("indicateurs", set())
        if in_e and in_i:
            report.add(
                _sheet_name(sheet_key),
                row["_row"],
                "cible",
                ERROR,
                f"Le code « {value} » existe comme enjeu ET comme "
                "indicateur : rendez-le unique.",
            )
        elif not in_e and not in_i:
            report.add(
                _sheet_name(sheet_key),
                row["_row"],
                "cible",
                ERROR,
                f"Cible « {value} » introuvable (ni enjeu ni indicateur).",
            )

    # Validation / correction des taxons et habitats contre TaxRef / HabRef.
    _resolve_bio_refs(parsed)

    for row in parsed.get("taxons", []):
        _check_cible("taxons", row)
        raw = _cell_str(row.get("cd_nom"))
        name = _cell_str(row.get("nom"))
        # 1) Format / obligation (toujours vérifié).
        if not raw and not name:
            report.add(
                "Taxons",
                row["_row"],
                "cd_nom",
                ERROR,
                "Renseignez le cd_nom (entier, code TaxRef) ou le nom de l'espèce.",
            )
            continue
        if raw and _as_int(raw) is None and not name:
            report.add(
                "Taxons",
                row["_row"],
                "cd_nom",
                ERROR,
                "Le cd_nom doit être un nombre entier (code TaxRef), ou renseignez "
                "le nom.",
            )
            continue
        # 2) Existence / correction contre TaxRef (si le référentiel est chargé).
        status = row.get("_bio")
        if status == "resolved":
            report.add(
                "Taxons",
                row["_row"],
                "cd_nom",
                WARNING,
                f"Taxon reconnu par son nom → cd_nom {row.get('_cd_nom')} retenu.",
            )
        elif status == "ambiguous":
            report.add(
                "Taxons",
                row["_row"],
                "nom",
                ERROR,
                f"Le nom « {name} » correspond à plusieurs taxons : précisez le "
                "cd_nom.",
            )
        elif status == "unknown":
            report.add(
                "Taxons",
                row["_row"],
                "cd_nom",
                ERROR,
                f"Le cd_nom « {raw} » est introuvable dans TaxRef.",
            )

    for row in parsed.get("habitats", []):
        _check_cible("habitats", row)
        raw = _cell_str(row.get("cd_hab"))
        name = _cell_str(row.get("nom"))
        if not raw and not name:
            report.add(
                "Habitats",
                row["_row"],
                "cd_hab",
                ERROR,
                "Renseignez le cd_hab (code HabRef) ou le libellé de l'habitat.",
            )
            continue
        status = row.get("_bio")
        if status == "resolved":
            report.add(
                "Habitats",
                row["_row"],
                "cd_hab",
                WARNING,
                f"Habitat reconnu → cd_hab {row.get('_cd_hab')} retenu.",
            )
        elif status == "ambiguous":
            report.add(
                "Habitats",
                row["_row"],
                "cd_hab",
                ERROR,
                f"« {raw} » correspond à plusieurs habitats : précisez le cd_hab.",
            )
        elif status == "unknown":
            report.add(
                "Habitats",
                row["_row"],
                "cd_hab",
                ERROR,
                f"Le cd_hab « {raw} » est introuvable dans HabRef.",
            )

    # --- Décompte de ce qui serait créé ---
    report.summary = {
        key: len(parsed.get(key, []))
        for key in (
            "enjeux",
            "facteurs",
            "pressions",
            "olt",
            "ne",
            "oo",
            "ra",
            "indicateurs",
            "metriques",
            "taxons",
            "habitats",
        )
    }
    return report


def _sheet_name(sheet_key: str) -> str:
    for sheet in _build_schema():
        if sheet.key == sheet_key:
            return sheet.name
    return sheet_key


# ---------------------------------------------------------------------------
# Exécution
# ---------------------------------------------------------------------------


@transaction.atomic
def execute_import(plan, parsed: dict[str, list[dict]], user) -> dict:
    """Crée l'arborescence dans le plan (transaction). Suppose la validation OK.

    Retourne un décompte des objets créés. Lève ``ValueError`` si la validation
    échoue (le rapport est joint via ``args``).
    """
    report = validate_import(plan, parsed)
    if not report.can_import:
        raise ValueError(report)

    resolver = _NomenclatureResolver()
    counts = {}

    def nom(type_mnemo, value):
        v = _cell_str(value)
        return resolver.resolve(type_mnemo, v) if v else None

    # --- Enjeux ---
    enjeu_map: dict[str, Enjeu] = {}
    for i, row in enumerate(parsed.get("enjeux", [])):
        enjeu = Enjeu(
            id_pg=plan,
            id_categorie=nom("CATEGORIE_ENJEU", row.get("categorie")),
            id_categorie_fcr=nom("CATEGORIE_FCR", row.get("categorie_fcr")),
            id_importance=nom("IMPORTANCE_ENJEU", row.get("importance")),
            libelle=_cell_str(row.get("libelle")),
            intitule_court=_cell_str(row.get("intitule_court")) or None,
            rang=_as_int(row.get("rang")),
            etat_enjeu=_cell_str(row.get("etat_enjeu")) or None,
            description=_cell_str(row.get("description")) or None,
            ordre=i,
            id_utilisateur_ajout=user,
        )
        eco = _parse_bool(row.get("categorie_ecologique"))
        if eco is not None:
            enjeu.categorie_ecologique = eco
        for token in _split_multi(row.get("types_ecologiques")):
            attr = _ECOLO_FIELD.get(_norm(token))
            if attr:
                setattr(enjeu, attr, True)
        for token in _split_multi(row.get("types_socioeco")):
            attr = _SOCIO_FIELD.get(_norm(token))
            if attr:
                setattr(enjeu, attr, True)
        enjeu.slug = ""  # régénéré par save()
        enjeu.save()
        enjeu_map[_cell_str(row.get("code"))] = enjeu
    counts["enjeux"] = len(enjeu_map)

    # --- Facteurs + liaison enjeux ---
    facteur_map: dict[str, FacteurInfluence] = {}
    for i, row in enumerate(parsed.get("facteurs", [])):
        facteur = FacteurInfluence.objects.create(
            libelle=_cell_str(row.get("libelle")),
            description=_cell_str(row.get("description")) or None,
            id_utilisateur_ajout=user,
        )
        code = _cell_str(row.get("code"))
        facteur_map[code] = facteur
        for j, enjeu_code in enumerate(_split_multi(row.get("enjeux"))):
            enjeu = enjeu_map.get(enjeu_code)
            if enjeu is not None:
                CorFacteurEnjeu.objects.create(
                    id_facteur_influence=facteur,
                    id_enjeu=enjeu,
                    ordre=i,
                )
    counts["facteurs"] = len(facteur_map)

    # --- Pressions ---
    pression_map: dict[str, Pression] = {}
    for i, row in enumerate(parsed.get("pressions", [])):
        facteur = facteur_map.get(_cell_str(row.get("facteur")))
        pression = Pression.objects.create(
            id_facteur_influence=facteur,
            libelle=_cell_str(row.get("libelle")),
            id_type_pression=nom("TYPE_PRESSION", row.get("type_pression")),
            description=_cell_str(row.get("description")) or None,
            ordre=i,
            id_utilisateur_ajout=user,
        )
        pression_map[_cell_str(row.get("code"))] = pression
    counts["pressions"] = len(pression_map)

    # --- OLT ---
    olt_map: dict[str, ObjectifLongTerme] = {}
    for i, row in enumerate(parsed.get("olt", [])):
        olt = ObjectifLongTerme.objects.create(
            id_enjeu=enjeu_map.get(_cell_str(row.get("enjeu"))),
            libelle=_cell_str(row.get("libelle")),
            numero_manuel=_as_int(row.get("numero_manuel")),
            description=_cell_str(row.get("description")) or None,
            ordre=i,
            id_utilisateur_ajout=user,
        )
        olt_map[_cell_str(row.get("code"))] = olt
    counts["olt"] = len(olt_map)

    # --- NE ---
    ne_map: dict[str, NiveauExigence] = {}
    for i, row in enumerate(parsed.get("ne", [])):
        ne = NiveauExigence.objects.create(
            id_olt=olt_map.get(_cell_str(row.get("olt"))),
            libelle=_cell_str(row.get("libelle")),
            description=_cell_str(row.get("description")) or None,
            ordre=i,
            id_utilisateur_ajout=user,
        )
        ne_map[_cell_str(row.get("code"))] = ne
    counts["ne"] = len(ne_map)

    # --- OO + liaisons pressions / enjeu direct ---
    oo_map: dict[str, ObjectifOperationnel] = {}
    for i, row in enumerate(parsed.get("oo", [])):
        pressions = _split_multi(row.get("pressions"))
        enjeu_code = _cell_str(row.get("enjeu"))
        oo = ObjectifOperationnel(
            libelle=_cell_str(row.get("libelle")),
            numero_manuel=_as_int(row.get("numero_manuel")),
            description=_cell_str(row.get("description")) or None,
            ordre=i,
            id_utilisateur_ajout=user,
        )
        if not pressions and enjeu_code:
            oo.id_enjeu = enjeu_map.get(enjeu_code)
        oo.save()
        for pression_code in pressions:
            pression = pression_map.get(pression_code)
            if pression is not None:
                CorOoPression.objects.create(id_oo=oo, id_pression=pression)
        oo_map[_cell_str(row.get("code"))] = oo
    counts["oo"] = len(oo_map)

    # --- RA ---
    ra_map: dict[str, ResultatAttendu] = {}
    for i, row in enumerate(parsed.get("ra", [])):
        ra = ResultatAttendu.objects.create(
            id_oo=oo_map.get(_cell_str(row.get("oo"))),
            libelle=_cell_str(row.get("libelle")),
            description=_cell_str(row.get("description")) or None,
            ordre=i,
            id_utilisateur_ajout=user,
        )
        ra_map[_cell_str(row.get("code"))] = ra
    counts["ra"] = len(ra_map)

    # --- Indicateurs (parent NE ou RA) ---
    indicateur_map: dict[str, Indicateur] = {}
    for i, row in enumerate(parsed.get("indicateurs", [])):
        parent = _cell_str(row.get("parent"))
        ne = ne_map.get(parent)
        ra = ra_map.get(parent)
        indicateur = Indicateur.objects.create(
            id_ne=ne,
            id_resultat_attendu=ra,
            nom_indicateur=_cell_str(row.get("nom_indicateur")),
            type_indicateur=nom("TYPE_INDICATEUR", row.get("type")),
            description=_cell_str(row.get("description")) or None,
            ordre=i,
            id_utilisateur_ajout=user,
        )
        indicateur_map[_cell_str(row.get("code"))] = indicateur
    counts["indicateurs"] = len(indicateur_map)

    # --- Métriques ---
    n_metriques = 0
    for i, row in enumerate(parsed.get("metriques", [])):
        Metrique.objects.create(
            id_indicateur=indicateur_map.get(_cell_str(row.get("indicateur"))),
            nom_metrique=_cell_str(row.get("nom_metrique")),
            type_metrique=nom("TYPE_METRIQUE", row.get("type_metrique")),
            unite=_cell_str(row.get("unite")) or None,
            description=_cell_str(row.get("description")) or None,
            ordre=i,
            id_utilisateur_ajout=user,
        )
        n_metriques += 1
    counts["metriques"] = n_metriques

    # --- Taxons / Habitats (cible = enjeu ou indicateur) ---
    def _resolve_cible(code):
        """Renvoie ('enjeu', obj) ou ('indicateur', obj) selon le code."""
        if code in enjeu_map:
            return "enjeu", enjeu_map[code]
        if code in indicateur_map:
            return "indicateur", indicateur_map[code]
        return None, None

    n_taxons = 0
    for row in parsed.get("taxons", []):
        kind, obj = _resolve_cible(_cell_str(row.get("cible")))
        # Code retenu par la résolution TaxRef (corrigé le cas échéant).
        cd_nom = row.get("_cd_nom", _as_int(row.get("cd_nom")))
        if obj is None or cd_nom is None:
            continue
        nom_complet = _cell_str(row.get("nom")) or None
        if kind == "enjeu":
            CorEnjeuTaxon.objects.create(
                id_enjeu=obj, cd_nom=cd_nom, nom_complet=nom_complet
            )
        else:
            CorIndicateurTaxon.objects.create(
                id_indicateur=obj, cd_nom=cd_nom, nom_complet=nom_complet
            )
        n_taxons += 1
    counts["taxons"] = n_taxons

    n_habitats = 0
    for row in parsed.get("habitats", []):
        kind, obj = _resolve_cible(_cell_str(row.get("cible")))
        # Code retenu par la résolution HabRef (corrigé le cas échéant).
        cd_hab = row.get("_cd_hab") or _cell_str(row.get("cd_hab"))
        if obj is None or not cd_hab:
            continue
        lb_hab_fr = _cell_str(row.get("nom")) or None
        if kind == "enjeu":
            CorEnjeuHabitat.objects.create(
                id_enjeu=obj, cd_hab=cd_hab, lb_hab_fr=lb_hab_fr
            )
        else:
            CorIndicateurHabitat.objects.create(
                id_indicateur=obj, cd_hab=cd_hab, lb_hab_fr=lb_hab_fr
            )
        n_habitats += 1
    counts["habitats"] = n_habitats

    return counts
