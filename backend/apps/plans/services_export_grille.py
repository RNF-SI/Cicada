"""
Rendu Excel d'une **grille affichée** : tableau de bord (#638), suivi des
actions (#637).

Ces tableaux sont filtrés côté client et leurs valeurs (scores, agrégats
budget/RH) calculées à l'affichage. C'est donc le client qui envoie **les
lignes telles qu'il les montre**, et ce module se charge de la seule chose
qu'un CSV ne sait pas faire : la mise en forme. L'export reste ainsi exactement
ce que l'utilisateur a sous les yeux — la promesse d'origine de #637/#638 —
sans réimplémenter filtrage et calculs une seconde fois côté serveur, où ils
divergeraient au premier écart.

Mise en forme (retour de recette : « dans le même esprit que l'interface ») :
en-têtes à la couleur de l'instance (#601), **cases de score reprenant la
palette de l'application** et lignes de total détachées du reste. Les couleurs
de score ne sont volontairement pas paramétrables — elles forment la grille de
lecture, cf. :mod:`export_theme`.

Contrat de la charge utile
--------------------------

``{'titre': str, 'onglet': str, 'meta': [[libellé, valeur]], 'entetes': [str],
   'gel': int, 'formats': [str | None], 'lignes': [{'type': ..., 'cellules': [...]}]}``

``formats`` est aligné sur ``entetes`` et n'a qu'une valeur utile : ``'euro'``,
qui suffixe les montants d'un ``€`` **sans les sortir du numérique** (le client
seul sait quelles colonnes sont des montants).

Types de ligne :

- ``normal``  — ligne courante (défaut) ;
- ``detail``  — sous-ligne rattachée à la précédente (métrique d'un
  indicateur), tramée pour se lire comme un repli ;
- ``total``   — ligne de synthèse, détachée par un aplat.

Une cellule est soit un texte, soit un nombre (écrit **comme un nombre**, pour
rester sommable dans le tableur), soit ``{'t': texte, 's': niveau de score}``
pour une case colorée, soit ``None`` pour une case vide.
"""

from __future__ import annotations

import io
import math
import numbers

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import export_theme

# --- palette ---------------------------------------------------------------

#: Couleurs de score de l'application (`SCORE_PALETTE`, design system).
#: Texte toujours noir : les fonds sont clairs, cf. docs/DESIGN_SYSTEM.md.
SCORE_FILLS = {
    'very-bad': 'FFFF7579',
    'bad': 'FFFA9965',
    'neutral': 'FFF7D35C',
    'good': 'FF82DB8A',
    'very-good': 'FF81C9D8',
    'no-data': 'FFDADADA',
}

_BLACK = 'FF343433'
_WHITE = 'FFFFFFFF'
_GRAY = 'FF746F6E'
_DETAIL_FILL = 'FFF8F5F1'   # beige du design system, pour les sous-lignes
_TOTAL_FILL = 'FFC0E3CF'    # vert pâle, réservé aux lignes de synthèse
_BORDER_COLOR = 'FFE4E4E4'

_PRIMARY = 'FF025359'
_F_TITLE = Font(name='Calibri', bold=True, size=14, color=_PRIMARY)
_F_META = Font(name='Calibri', size=9, color=_GRAY)
_F_HDR = Font(name='Calibri', bold=True, size=9, color=_WHITE)
_F_NORMAL = Font(name='Calibri', bold=True, size=9, color=_BLACK)
_F_DETAIL = Font(name='Calibri', size=9, color=_BLACK)
_F_TOTAL = Font(name='Calibri', bold=True, size=9, color=_PRIMARY)
_F_SCORE = Font(name='Calibri', size=9, color=_BLACK)

_THIN = Side(style='thin', color=_BORDER_COLOR)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

#: Séparateur de milliers, décimales seulement si utiles : montants et jours se
#: lisent en entiers à l'écran.
#:
#: Le format est choisi **par cellule** : un tableur affiche le séparateur
#: décimal même quand aucune décimale ne suit (`#,##0.##` donnait « 12 345, »,
#: lu comme un € raté — #644). Un entier reçoit donc un format sans décimales.
#: Le séparateur de milliers s'écrit `,` : c'est le code invariant du format
#: xlsx, que le tableur rend avec le séparateur de la langue (espace en FR).
_NUM_ENTIER = '#,##0'
_NUM_DECIMAL = '#,##0.##'
_EURO_ENTIER = '#,##0 "€"'
_EURO_DECIMAL = '#,##0.## "€"'


def _appliquer_couleur_instance():
    """Réaligne les styles sur la couleur d'export de l'instance (#601)."""
    global _PRIMARY, _F_TITLE, _F_TOTAL

    couleur = export_theme.argb()
    if couleur == _PRIMARY:
        return
    _PRIMARY = couleur
    _F_TITLE = Font(name='Calibri', bold=True, size=14, color=couleur)
    _F_TOTAL = Font(name='Calibri', bold=True, size=9, color=couleur)


# --- lecture de la charge utile --------------------------------------------

def _score(cellule) -> str | None:
    """Niveau de score d'une cellule, s'il y en a un et qu'il est connu."""
    if not isinstance(cellule, dict):
        return None
    niveau = cellule.get('s')
    return niveau if niveau in SCORE_FILLS else None


def _valeur(cellule):
    """Valeur écrite dans la cellule : nombre conservé, reste en texte."""
    brute = cellule.get('t') if isinstance(cellule, dict) else cellule
    if brute is None:
        return None
    if isinstance(brute, bool):
        return str(brute)
    if isinstance(brute, numbers.Number):
        return brute
    return str(brute)


def _format_nombre(valeur, euro: bool) -> str:
    """Format d'affichage d'un nombre : décimales et « € » seulement si utiles."""
    brut = float(valeur)
    # Un NaN / infini venant du client ne doit pas faire échouer tout l'export.
    entier = not math.isfinite(brut) or brut == int(brut)
    if euro:
        return _EURO_ENTIER if entier else _EURO_DECIMAL
    return _NUM_ENTIER if entier else _NUM_DECIMAL


def _texte(cellule) -> str:
    """Représentation textuelle, pour le calcul des largeurs de colonnes."""
    valeur = _valeur(cellule)
    return '' if valeur is None else str(valeur)


# --- construction ----------------------------------------------------------

def build_grille_workbook(payload: dict) -> bytes:
    """Rend la grille affichée en classeur Excel mis en forme."""
    _appliquer_couleur_instance()

    entetes = [str(h) for h in (payload.get('entetes') or [])]
    lignes = payload.get('lignes') or []
    meta = payload.get('meta') or []
    formats = payload.get('formats') or []

    wb = Workbook()
    ws = wb.active
    ws.title = str(payload.get('onglet') or 'Export')[:31]

    ligne = 1
    titre = payload.get('titre')
    if titre:
        ws.cell(row=ligne, column=1, value=str(titre)).font = _F_TITLE
        ligne += 1

    for libelle, valeur in ((list(m) + ['', ''])[:2] for m in meta):
        ws.cell(row=ligne, column=1, value=str(libelle)).font = _F_META
        ws.cell(row=ligne, column=2, value=str(valeur)).font = _F_META
        ligne += 1

    if titre or meta:
        ligne += 1  # respiration avant le tableau

    # En-têtes : aplat à la couleur de l'instance, texte blanc.
    entete_ligne = ligne
    for col, libelle in enumerate(entetes, start=1):
        cell = ws.cell(row=entete_ligne, column=col, value=libelle)
        cell.font = _F_HDR
        cell.fill = PatternFill('solid', fgColor=_PRIMARY)
        cell.alignment = _CENTER
        cell.border = _BORDER
    ligne += 1

    for row in lignes:
        _ecrire_ligne(ws, ligne, row, formats)
        ligne += 1

    _mettre_en_page(ws, entetes, lignes, entete_ligne, payload.get('gel'))

    flux = io.BytesIO()
    wb.save(flux)
    return flux.getvalue()


def _ecrire_ligne(ws, ligne: int, row: dict, formats=()):
    type_ligne = row.get('type') or 'normal'
    fond = {'detail': _DETAIL_FILL, 'total': _TOTAL_FILL}.get(type_ligne)
    police = {'detail': _F_DETAIL, 'total': _F_TOTAL}.get(type_ligne, _F_NORMAL)

    for col, brute in enumerate(row.get('cellules') or [], start=1):
        cell = ws.cell(row=ligne, column=col, value=_valeur(brute))
        niveau = _score(brute)
        if niveau and not fond:
            # Case colorée comme la pastille de score de l'interface.
            cell.fill = PatternFill('solid', fgColor=SCORE_FILLS[niveau])
            cell.font = _F_SCORE
            cell.alignment = _CENTER
        else:
            # Une ligne de total garde son aplat de synthèse : sa couleur dit
            # « c'est une somme », elle ne doit pas être lue comme un score.
            cell.font = police
            cell.alignment = _CENTER if isinstance(cell.value, numbers.Number) else _LEFT
            if fond:
                cell.fill = PatternFill('solid', fgColor=fond)
        if isinstance(cell.value, numbers.Number):
            euro = col <= len(formats) and formats[col - 1] == 'euro'
            cell.number_format = _format_nombre(cell.value, euro)
        cell.border = _BORDER


def _mettre_en_page(ws, entetes, lignes, entete_ligne, gel=None):
    """Largeurs de colonnes, hauteur d'en-tête, volets figés et filtre."""
    if not entetes:
        return

    # Largeur : au plus large des libellés de la colonne, dans des bornes qui
    # gardent les colonnes chiffrées compactes et les libellés lisibles.
    for col, libelle in enumerate(entetes, start=1):
        largeur = len(libelle)
        for row in lignes:
            cellules = row.get('cellules') or []
            if col <= len(cellules):
                largeur = max(largeur, len(_texte(cellules[col - 1])))
        ws.column_dimensions[get_column_letter(col)].width = min(max(largeur + 2, 8), 42)

    ws.row_dimensions[entete_ligne].height = 30
    # Fige l'en-tête ET les colonnes d'identification annoncées par le client :
    # sans cela, faire défiler les années ou les périodes fait perdre de vue la
    # ligne qu'on lit.
    colonnes_gelees = max(0, min(int(gel or 0), len(entetes)))
    ws.freeze_panes = ws.cell(row=entete_ligne + 1, column=colonnes_gelees + 1)
    ws.auto_filter.ref = (
        f'A{entete_ligne}:{get_column_letter(len(entetes))}{entete_ligne + len(lignes)}'
    )
