"""
Import / export Excel des **actions** (opérations) d'un plan de gestion — module 2
du dispositif d'import (#478). Fait suite au module « arborescence »
(``services_import.py``).

Une action s'importe dans un plan qui possède **déjà son arborescence**. Comme
les codes logiques de l'arborescence (I1, I2…) ne sont pas persistés en base, le
classeur d'actions est **généré depuis le plan** : il embarque un onglet de
référence ``Indicateurs`` listant les indicateurs existants (code lisible + nom
+ enjeu + identifiant technique). L'onglet ``Actions`` référence l'indicateur
parent par ce code (liste déroulante). À l'import, on relit les deux onglets
pour rattacher chaque action au bon indicateur du plan.

Périmètre : libellé, type d'action, priorité, années (``annee_min`` /
``annee_max`` → ``OperationAnnee``), opérateurs / financeurs (texte libre).
Deux onglets facultatifs complètent la saisie :

- ``Budgets`` : budget de fonctionnement / investissement par (action, année)
  → renseigne l'``OperationAnnee`` et bascule l'opération en ``ventilation_mode
  = 'by_type'`` ;
- ``RH`` : temps de travail en jours par (action, année, poste) → crée des
  ``OperationAnneeRH`` (#560) et active ``declinaison_par_poste``. Les postes du
  plan sont listés dans un onglet de référence ``Postes``.

Les actions sont créées en statut ``draft``.
"""

from __future__ import annotations

import io
from decimal import Decimal, InvalidOperation

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from apps.core.models import Nomenclature

from .models_indicateurs import Indicateur
from .models_operations import Operation, OperationAnnee, OperationAnneeRH, Poste
from .services_import import (
    ERROR,
    FORMAT_VERSION,
    WARNING,
    ArborescenceImportError,
    ImportReport,
    _NomenclatureResolver,
    _as_int,
    _cell_str,
    _is_example_row,
    _norm,
    _parse_bool,
    _BORDER,
    _EXAMPLE_MARKER,
    _HEADER_FILL,
    _HEADER_FONT,
    _HINT_FILL,
    _HINT_FONT,
    _PRIMARY,
    _REQUIRED_FILL,
    _TITLE_FONT,
    _WRAP_TOP,
)


def _as_decimal(value):
    """Convertit une cellule en Decimal (ou None si vide/invalide)."""
    text = _cell_str(value)
    if not text:
        return None
    try:
        return Decimal(text.replace(",", ".").replace(" ", ""))
    except (InvalidOperation, ValueError):
        return None


_HEADER_ROW = 2
_FIRST_DATA_ROW = 3
_BLANK_ROWS = 200

# Types de nomenclature utilisés par les actions.
_TYPE_ACTION = "TYPE_ACTION"
_PRIORITE = "PRIORITE_OPERATION"


# ---------------------------------------------------------------------------
# Traversée des indicateurs d'un plan (référence stable code → indicateur)
# ---------------------------------------------------------------------------


def _plan_indicateurs(plan) -> list[tuple[Indicateur, str]]:
    """Retourne les indicateurs du plan avec le libellé de leur enjeu, dans un
    ordre déterministe (branche état puis branche opérationnelle)."""
    result: list[tuple[Indicateur, str]] = []
    seen: set[int] = set()

    def _add(ind, enjeu_libelle):
        if ind.id_indicateur not in seen:
            seen.add(ind.id_indicateur)
            result.append((ind, enjeu_libelle))

    enjeux = list(plan.enjeux.all().order_by("ordre", "id_enjeu"))

    # Branche état : Enjeu → OLT → NE → Indicateur
    for enjeu in enjeux:
        for olt in enjeu.objectifs_long_terme.all().order_by("ordre", "id_olt"):
            for ne in olt.niveaux_exigence.all().order_by("ordre", "id_ne"):
                for ind in ne.indicateurs.all().order_by("ordre", "id_indicateur"):
                    _add(ind, enjeu.libelle)

    # Branche opérationnelle : Enjeu → Facteur → Pression → OO → RA → Indicateur
    for enjeu in enjeux:
        for cor in enjeu.cor_facteurs.all().order_by("ordre", "id"):
            facteur = cor.id_facteur_influence
            for pr in facteur.pressions.all().order_by("ordre", "id_pression"):
                for oo in pr.objectifs_operationnels.all().order_by("ordre", "id_oo"):
                    for ra in oo.resultats_attendus.all().order_by("ordre", "id_ra"):
                        for ind in ra.indicateurs.all().order_by(
                            "ordre", "id_indicateur"
                        ):
                            _add(ind, enjeu.libelle)
        # OO rattachés directement au FCR (#337)
        for oo in enjeu.objectifs_operationnels_directs.all().order_by(
            "ordre", "id_oo"
        ):
            for ra in oo.resultats_attendus.all().order_by("ordre", "id_ra"):
                for ind in ra.indicateurs.all().order_by("ordre", "id_indicateur"):
                    _add(ind, enjeu.libelle)

    return result


def _plan_postes(plan) -> list[Poste]:
    """Postes du plan (référence pour la ventilation RH #560)."""
    return list(plan.postes.all().order_by("id_poste"))


# ---------------------------------------------------------------------------
# Colonnes de l'onglet « Actions »
# ---------------------------------------------------------------------------

# (key, header, required, help, nomenclature_type, width)
_ACTION_COLUMNS = [
    (
        "code",
        "code",
        True,
        "Identifiant libre et unique de l'action (ex : A1).",
        None,
        10,
    ),
    (
        "indicateur",
        "indicateur",
        True,
        "Code de l'indicateur auquel l'action se rattache (voir l'onglet "
        "« Indicateurs »).",
        None,
        14,
    ),
    ("libelle", "libellé", True, "Intitulé de l'action.", None, 45),
    (
        "type_action",
        "type d'action",
        False,
        "Type d'action (codification Eden 62).",
        _TYPE_ACTION,
        26,
    ),
    ("priorite", "priorité", False, "Priorité de l'action.", _PRIORITE, 18),
    ("annee_min", "année début", False, "Première année (ex : 2024).", None, 14),
    ("annee_max", "année fin", False, "Dernière année (ex : 2028).", None, 14),
    ("operateurs", "opérateurs", False, "Texte libre.", None, 30),
    ("financeurs", "financeurs", False, "Texte libre.", None, 30),
    ("description", "description", False, "", None, 40),
]

# Colonnes de l'onglet de référence « Indicateurs » (lecture seule à la saisie).
_REF_COLUMNS = [
    ("code", "code", 10),
    ("indicateur", "indicateur", 45),
    ("enjeu", "enjeu", 35),
    ("id", "id (technique — ne pas modifier)", 26),
]

# Colonnes de l'onglet de référence « Postes » (lecture seule à la saisie).
_POSTE_REF_COLUMNS = [
    ("code", "code", 10),
    ("poste", "poste", 45),
    ("organisme", "organisme", 30),
    ("id", "id (technique — ne pas modifier)", 26),
]

# Onglet « Budgets » : budget par (action, année). Ventilation par type
# (fonctionnement / investissement), au niveau de l'OperationAnnee.
# (key, header, required, help, width)
_BUDGET_COLUMNS = [
    ("action", "action", True, "Code de l'action (onglet « Actions »).", 12),
    ("annee", "année", True, "Année concernée (ex : 2024).", 12),
    ("budget_fonctionnement", "budget fonctionnement (€)", False, "", 24),
    ("budget_investissement", "budget investissement (€)", False, "", 24),
]

# Onglet « RH » : temps de travail par (action, année, poste), en jours (#560).
_RH_COLUMNS = [
    ("action", "action", True, "Code de l'action (onglet « Actions »).", 12),
    ("annee", "année", True, "Année concernée (ex : 2024).", 12),
    ("poste", "poste", True, "Code du poste (onglet « Postes »).", 12),
    ("jours", "jours", False, "Nombre de jours travaillés.", 12),
    ("finance", "financé ?", False, "Oui / Non (temps financé ou non).", 12),
]

# En-têtes normalisés → clé de colonne, pour le parsing des onglets à plat.
_BUDGET_HEADERS = {_norm(c[1]): c[0] for c in _BUDGET_COLUMNS}
_RH_HEADERS = {_norm(c[1]): c[0] for c in _RH_COLUMNS}


# ---------------------------------------------------------------------------
# Construction du classeur
# ---------------------------------------------------------------------------


def _nomenclature_values() -> dict[str, list[str]]:
    values: dict[str, list[str]] = {}
    for type_mnemo in (_TYPE_ACTION, _PRIORITE):
        labels = list(
            Nomenclature.objects.filter(id_type__mnemonique=type_mnemo, actif=True)
            .order_by("id_nomenclature")
            .values_list("label", flat=True)
        )
        values[type_mnemo] = [lbl for lbl in labels if lbl]
    return values


def _actions_code_range(n_action_rows: int) -> str:
    """Plage de la colonne « code » de l'onglet Actions (source des dropdowns
    « action » des onglets Budgets et RH)."""
    last = _HEADER_ROW + max(n_action_rows, 0) + _BLANK_ROWS
    return f"'Actions'!$A${_FIRST_DATA_ROW}:$A${last}"


def _render_actions_workbook(
    indicateurs,
    code_by_id,
    postes,
    poste_code_by_id,
    action_rows,
    budget_rows,
    rh_rows,
    plan=None,
    with_hints=False,
    example_name=None,
) -> bytes:
    """Assemble le classeur des actions à partir de données déjà préparées."""
    wb = Workbook()
    wb.remove(wb.active)

    # Codes de rattachement pour les dropdowns.
    first_ind_code = next(iter(code_by_id.values()), "I1")
    first_poste_code = next(iter(poste_code_by_id.values()), "Q1")
    actions_range = _actions_code_range(len(action_rows) + (1 if with_hints else 0))

    hints = _actions_hint_rows(first_ind_code, first_poste_code) if with_hints else {}

    _write_lisez_moi(wb, plan, example_name=example_name, with_hints=with_hints)
    ref_code_range = _write_indicateurs_ref(wb, indicateurs, code_by_id)
    poste_code_range = _write_postes_ref(wb, postes, poste_code_by_id)
    list_ranges = _write_listes(wb)
    _write_actions(
        wb,
        action_rows,
        ref_code_range,
        list_ranges,
        hint_row=hints.get("actions"),
    )
    _write_budgets(wb, budget_rows, actions_range, hint_row=hints.get("budgets"))
    _write_rh(
        wb,
        rh_rows,
        poste_code_range,
        actions_range,
        hint_row=hints.get("rh"),
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_actions_workbook(plan) -> bytes:
    """Construit le classeur d'import des actions pour un plan donné.

    L'onglet ``Indicateurs`` est pré-rempli avec les indicateurs existants du
    plan ; l'onglet ``Actions`` est vierge (avec une ligne exemple) ou
    pré-rempli si des actions existent déjà (export/sauvegarde).
    """
    indicateurs = _plan_indicateurs(plan)
    code_by_id = {
        ind.id_indicateur: f"I{i}" for i, (ind, _) in enumerate(indicateurs, start=1)
    }
    postes = _plan_postes(plan)
    poste_code_by_id = {p.id_poste: f"Q{i}" for i, p in enumerate(postes, start=1)}

    action_rows = _extract_actions(plan, code_by_id)
    budget_rows = _extract_budgets(plan)
    rh_rows = _extract_rh(plan, poste_code_by_id)

    return _render_actions_workbook(
        indicateurs,
        code_by_id,
        postes,
        poste_code_by_id,
        action_rows,
        budget_rows,
        rh_rows,
        plan=plan,
        with_hints=not action_rows,
    )


ACTIONS_EXAMPLE_NAME = "Réserve naturelle d'une zone humide (exemple)"


def _actions_hint_rows(first_ind_code: str, first_poste_code: str) -> dict[str, dict]:
    """Une ligne « exemple » par onglet de saisie (Actions, Budgets, RH).

    La 1re colonne porte le marqueur ``(exemple)`` (ignoré à l'import) ; les
    colonnes de rattachement montrent des codes réels (indicateur, poste) pour
    illustrer les liens.
    """
    return {
        "actions": {
            "code": _EXAMPLE_MARKER,
            "indicateur": first_ind_code,
            "libelle": "Faucher tardivement les prairies humides",
            "annee_min": 2024,
            "annee_max": 2028,
            "operateurs": "Équipe technique de la réserve",
            "financeurs": "Agence de l'eau",
            "description": "Fauche annuelle avec exportation, après le 15 juillet.",
        },
        "budgets": {
            "action": _EXAMPLE_MARKER,
            "annee": 2024,
            "budget_fonctionnement": 1500,
            "budget_investissement": 0,
        },
        "rh": {
            "action": _EXAMPLE_MARKER,
            "annee": 2024,
            "poste": first_poste_code,
            "jours": 6,
            "finance": "Oui",
        },
    }


def _actions_example_data():
    """Données fictives (indicateurs/postes/actions/budgets/RH) pour l'exemple."""
    from types import SimpleNamespace

    indicateurs = [
        (
            SimpleNamespace(
                id_indicateur=9001,
                nom_indicateur="État de conservation des habitats tourbeux",
            ),
            "Habitats tourbeux",
        ),
        (
            SimpleNamespace(
                id_indicateur=9002,
                nom_indicateur="Richesse en passereaux paludicoles nicheurs",
            ),
            "Avifaune paludicole",
        ),
        (
            SimpleNamespace(
                id_indicateur=9003,
                nom_indicateur="Surface colonisée par la Jussie",
            ),
            "Habitats tourbeux",
        ),
    ]
    code_by_id = {9001: "I1", 9002: "I2", 9003: "I3"}
    postes = [
        SimpleNamespace(
            id_poste=8001, libelle="Chargé·e de mission", id_organisme=None
        ),
        SimpleNamespace(
            id_poste=8002, libelle="Garde technicien·ne", id_organisme=None
        ),
    ]
    poste_code_by_id = {8001: "Q1", 8002: "Q2"}

    # Valeurs de nomenclature réelles (si disponibles) pour un rendu réaliste.
    noms = _nomenclature_values()
    type_action = (noms.get(_TYPE_ACTION) or [""])[0]
    priorite = (noms.get(_PRIORITE) or [""])[0]

    action_rows = [
        {
            "code": "A1",
            "indicateur": "I1",
            "libelle": "Faucher tardivement les prairies humides",
            "type_action": type_action,
            "priorite": priorite,
            "annee_min": 2024,
            "annee_max": 2028,
            "operateurs": "Équipe technique de la réserve",
            "financeurs": "Agence de l'eau",
            "description": "Fauche annuelle avec exportation, après le 15 juillet.",
        },
        {
            "code": "A2",
            "indicateur": "I2",
            "libelle": "Poser des panneaux de mise en défens des roselières",
            "type_action": type_action,
            "priorite": priorite,
            "annee_min": 2024,
            "annee_max": 2025,
            "operateurs": "Garde technicien·ne",
            "financeurs": "Région",
            "description": "Balisage des zones de quiétude au printemps.",
        },
        {
            "code": "A3",
            "indicateur": "I3",
            "libelle": "Arrachage manuel de la Jussie",
            "type_action": type_action,
            "priorite": priorite,
            "annee_min": 2024,
            "annee_max": 2029,
            "operateurs": "Chantier bénévole",
            "financeurs": "Agence de l'eau",
            "description": "Campagnes d'arrachage estivales répétées.",
        },
    ]
    budget_rows = [
        {
            "action": "A1",
            "annee": 2024,
            "budget_fonctionnement": 1500,
            "budget_investissement": 0,
        },
        {
            "action": "A1",
            "annee": 2025,
            "budget_fonctionnement": 1500,
            "budget_investissement": 0,
        },
        {
            "action": "A3",
            "annee": 2024,
            "budget_fonctionnement": 800,
            "budget_investissement": 1200,
        },
    ]
    rh_rows = [
        {"action": "A1", "annee": 2024, "poste": "Q1", "jours": 6, "finance": "Oui"},
        {"action": "A1", "annee": 2024, "poste": "Q2", "jours": 4, "finance": "Non"},
        {"action": "A3", "annee": 2024, "poste": "Q2", "jours": 10, "finance": "Oui"},
    ]
    return (
        indicateurs,
        code_by_id,
        postes,
        poste_code_by_id,
        action_rows,
        budget_rows,
        rh_rows,
    )


def build_actions_example_workbook() -> bytes:
    """Classeur **exemple** des actions, entièrement rempli (indépendant d'un plan).

    Reprend le thème de l'exemple d'arborescence : montre des actions rattachées
    à des indicateurs (onglet de référence), avec budgets et RH renseignés, et
    illustre les liens entre onglets. Fictif : à consulter, pas à importer tel
    quel (les indicateurs de référence n'existent dans aucun plan réel).
    """
    (
        indicateurs,
        code_by_id,
        postes,
        poste_code_by_id,
        action_rows,
        budget_rows,
        rh_rows,
    ) = _actions_example_data()
    return _render_actions_workbook(
        indicateurs,
        code_by_id,
        postes,
        poste_code_by_id,
        action_rows,
        budget_rows,
        rh_rows,
        with_hints=False,
        example_name=ACTIONS_EXAMPLE_NAME,
    )


def _write_lisez_moi(wb: Workbook, plan, example_name=None, with_hints=False) -> None:
    ws = wb.create_sheet("Lisez-moi", 0)
    ws.sheet_properties.tabColor = _PRIMARY
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 110
    lines = [
        ("Import des actions d'un plan de gestion", _TITLE_FONT),
        (f"Format version {FORMAT_VERSION}", Font(italic=True, color="746F6E")),
        ("", None),
        ("• Remplissez l'onglet « Actions » : une action par ligne.", None),
        (
            "• Rattachez chaque action à un indicateur en reportant son code "
            "(colonne « indicateur ») depuis l'onglet « Indicateurs ».",
            None,
        ),
        (
            "• L'onglet « Indicateurs » liste les indicateurs existants du plan. "
            "Ne modifiez pas la colonne « id (technique) ».",
            None,
        ),
        (
            "• Les colonnes type d'action et priorité proposent une liste déroulante.",
            None,
        ),
        ("• Les actions importées sont créées en brouillon.", None),
        ("", None),
        (
            "Budgets et ressources humaines (facultatif)",
            Font(bold=True, color=_PRIMARY),
        ),
        (
            "• Onglet « Budgets » : un budget (fonctionnement / investissement) par "
            "action et par année. Reportez le code de l'action et l'année.",
            None,
        ),
        (
            "• Onglet « RH » : temps de travail en jours par action, année et poste. "
            "Les postes du plan sont listés dans l'onglet « Postes ».",
            None,
        ),
        (
            "• Créez d'abord vos postes (page « Postes / RH » du plan) pour pouvoir "
            "les référencer ici.",
            None,
        ),
        (
            "• Les onglets Budgets et RH proposent une liste déroulante « action » "
            "avec les codes que vous saisissez dans l'onglet « Actions ».",
            None,
        ),
    ]
    if with_hints:
        lines += [
            ("", None),
            (
                "• La première ligne grisée des onglets Actions, Budgets et RH est "
                "un EXEMPLE (commence par « (exemple) ») : elle montre quoi écrire "
                "et n'est jamais importée. Saisissez vos données sur les lignes "
                "suivantes.",
                Font(italic=True, color="9A8F86"),
            ),
        ]
    if example_name is not None:
        lines += [
            ("", None),
            (
                f"⚠ EXEMPLE PÉDAGOGIQUE FICTIF — {example_name}. Illustre le format "
                "et les liens entre onglets ; remplacez son contenu par le vôtre.",
                Font(bold=True, italic=True, color="B74D5D"),
            ),
        ]
    if plan is not None:
        lines += [("", None), (f"Plan : {plan.nom}", Font(italic=True, color="746F6E"))]
    for i, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=2, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.sheet_view.showGridLines = False


def _write_indicateurs_ref(wb, indicateurs, code_by_id) -> str:
    ws = wb.create_sheet("Indicateurs")
    desc = ws.cell(
        row=1,
        column=1,
        value="Indicateurs existants du plan (référence). "
        "Ne modifiez pas la colonne « id ».",
    )
    desc.font = Font(italic=True, color="746F6E", size=10)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_REF_COLUMNS))
    for c, (key, header, width) in enumerate(_REF_COLUMNS, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=c, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(c)].width = width
    for r, (ind, enjeu_libelle) in enumerate(indicateurs, start=_FIRST_DATA_ROW):
        ws.cell(row=r, column=1, value=code_by_id[ind.id_indicateur])
        ws.cell(row=r, column=2, value=ind.nom_indicateur)
        ws.cell(row=r, column=3, value=enjeu_libelle)
        ws.cell(row=r, column=4, value=ind.id_indicateur)
        for c in range(1, len(_REF_COLUMNS) + 1):
            ws.cell(row=r, column=c).border = _BORDER
            ws.cell(row=r, column=c).alignment = _WRAP_TOP
    ws.freeze_panes = f"A{_FIRST_DATA_ROW}"
    ws.sheet_view.showGridLines = False
    ws.protection.sheet = True  # onglet de référence : non modifiable
    last = max(len(indicateurs) + _HEADER_ROW, _HEADER_ROW + 1)
    return f"'Indicateurs'!$A${_FIRST_DATA_ROW}:$A${last}"


def _write_listes(wb) -> dict[str, str]:
    ws = wb.create_sheet("Listes")
    ranges: dict[str, str] = {}
    for col_idx, (type_mnemo, labels) in enumerate(
        _nomenclature_values().items(), start=1
    ):
        letter = get_column_letter(col_idx)
        ws.cell(row=1, column=col_idx, value=type_mnemo).font = Font(bold=True)
        for r, label in enumerate(labels, start=2):
            ws.cell(row=r, column=col_idx, value=label)
        if labels:
            ranges[type_mnemo] = f"'Listes'!${letter}$2:${letter}${len(labels) + 1}"
        ws.column_dimensions[letter].width = 28
    ws.sheet_state = "hidden"
    ws.protection.sheet = True  # onglet de référence : non modifiable
    return ranges


def _write_actions(wb, rows, ref_code_range, list_ranges, hint_row=None) -> None:
    ws = wb.create_sheet("Actions")
    desc = ws.cell(
        row=1,
        column=1,
        value="Une action par ligne, rattachée à un " "indicateur par son code.",
    )
    desc.font = Font(italic=True, color="746F6E", size=10)
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=len(_ACTION_COLUMNS)
    )

    for c, (key, header, required, help_text, nomencl, width) in enumerate(
        _ACTION_COLUMNS, start=1
    ):
        cell = ws.cell(row=_HEADER_ROW, column=c, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _REQUIRED_FILL if required else _HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = _BORDER
        if help_text:
            cell.comment = Comment(
                help_text + (" (obligatoire)" if required else ""), "CICADA"
            )
        ws.column_dimensions[get_column_letter(c)].width = width

    r = _FIRST_DATA_ROW
    n_hint = 0
    if hint_row and not rows:
        for c, (key, *_rest) in enumerate(_ACTION_COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=hint_row.get(key, ""))
            cell.alignment = _WRAP_TOP
            cell.border = _BORDER
            cell.font = _HINT_FONT
            cell.fill = _HINT_FILL
        r += 1
        n_hint = 1
    for row in rows:
        for c, (key, *_rest) in enumerate(_ACTION_COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=row.get(key, ""))
            cell.alignment = _WRAP_TOP
            cell.border = _BORDER
        r += 1

    last_row = _HEADER_ROW + n_hint + max(len(rows), 0) + _BLANK_ROWS
    for c, (key, header, required, help_text, nomencl, width) in enumerate(
        _ACTION_COLUMNS, start=1
    ):
        letter = get_column_letter(c)
        dv = None
        if key == "indicateur":
            dv = DataValidation(type="list", formula1=ref_code_range, allow_blank=True)
        elif nomencl and list_ranges.get(nomencl):
            dv = DataValidation(
                type="list", formula1=list_ranges[nomencl], allow_blank=True
            )
        if dv is not None:
            dv.error = "Choisissez une valeur dans la liste proposée."
            dv.errorTitle = "Valeur non autorisée"
            ws.add_data_validation(dv)
            dv.add(f"{letter}{_FIRST_DATA_ROW}:{letter}{last_row}")

    ws.freeze_panes = f"A{_FIRST_DATA_ROW}"
    ws.auto_filter.ref = (
        f"A{_HEADER_ROW}:{get_column_letter(len(_ACTION_COLUMNS))}{_HEADER_ROW}"
    )
    ws.sheet_view.showGridLines = False


def _extract_actions(plan, code_by_id) -> list[dict]:
    """Actions existantes du plan, pour le pré-remplissage (export/sauvegarde)."""
    ids = list(code_by_id.keys())
    if not ids:
        return []
    rows = []
    ops = Operation.objects.filter(id_indicateur_id__in=ids).order_by(
        "ordre", "id_operation"
    )
    for op in ops:
        rows.append(
            {
                "code": op.code_operation or f"A{op.id_operation}",
                "indicateur": code_by_id.get(op.id_indicateur_id, ""),
                "libelle": op.libelle,
                "type_action": op.id_type_action.label if op.id_type_action else "",
                "priorite": op.id_priorite.label if op.id_priorite else "",
                "annee_min": op.annee_min if op.annee_min is not None else "",
                "annee_max": op.annee_max if op.annee_max is not None else "",
                "operateurs": op.operateurs or "",
                "financeurs": op.financeurs or "",
                "description": op.description or "",
            }
        )
    return rows


def _write_postes_ref(wb, postes, poste_code_by_id) -> str:
    ws = wb.create_sheet("Postes")
    desc = ws.cell(
        row=1,
        column=1,
        value="Postes existants du plan (référence RH). Ne modifiez pas la "
        "colonne « id ». Créez vos postes via la page « Postes / RH » du plan.",
    )
    desc.font = Font(italic=True, color="746F6E", size=10)
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=len(_POSTE_REF_COLUMNS)
    )
    for c, (key, header, width) in enumerate(_POSTE_REF_COLUMNS, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=c, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(c)].width = width
    for r, poste in enumerate(postes, start=_FIRST_DATA_ROW):
        ws.cell(row=r, column=1, value=poste_code_by_id[poste.id_poste])
        ws.cell(row=r, column=2, value=poste.libelle)
        ws.cell(
            row=r, column=3, value=str(poste.id_organisme) if poste.id_organisme else ""
        )
        ws.cell(row=r, column=4, value=poste.id_poste)
        for c in range(1, len(_POSTE_REF_COLUMNS) + 1):
            ws.cell(row=r, column=c).border = _BORDER
            ws.cell(row=r, column=c).alignment = _WRAP_TOP
    ws.freeze_panes = f"A{_FIRST_DATA_ROW}"
    ws.sheet_view.showGridLines = False
    ws.protection.sheet = True  # onglet de référence : non modifiable
    last = max(len(postes) + _HEADER_ROW, _HEADER_ROW + 1)
    return f"'Postes'!$A${_FIRST_DATA_ROW}:$A${last}"


def _write_simple_sheet(
    wb, name, description, columns, rows, dropdowns=None, hint_row=None
) -> None:
    """Écrit un onglet de saisie « à plat » (Budgets, RH).

    ``columns`` : liste de (key, header, required, help, width).
    ``dropdowns`` : {key: (type_or_range)} — 'oui_non' ou une référence de plage.
    ``hint_row`` : ligne exemple (grisée, jamais importée) si l'onglet est vide.
    """
    dropdowns = dropdowns or {}
    ws = wb.create_sheet(name)
    desc = ws.cell(row=1, column=1, value=description)
    desc.font = Font(italic=True, color="746F6E", size=10)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))

    for c, (key, header, required, help_text, width) in enumerate(columns, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=c, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _REQUIRED_FILL if required else _HEADER_FILL
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = _BORDER
        if help_text:
            cell.comment = Comment(
                help_text + (" (obligatoire)" if required else ""), "CICADA"
            )
        ws.column_dimensions[get_column_letter(c)].width = width

    r = _FIRST_DATA_ROW
    n_hint = 0
    if hint_row and not rows:
        for c, (key, *_rest) in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=c, value=hint_row.get(key, ""))
            cell.alignment = _WRAP_TOP
            cell.border = _BORDER
            cell.font = _HINT_FONT
            cell.fill = _HINT_FILL
        r += 1
        n_hint = 1
    for row in rows:
        for c, (key, *_rest) in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=c, value=row.get(key, ""))
            cell.alignment = _WRAP_TOP
            cell.border = _BORDER
        r += 1

    last_row = _HEADER_ROW + n_hint + max(len(rows), 0) + _BLANK_ROWS
    for c, (key, *_rest) in enumerate(columns, start=1):
        spec = dropdowns.get(key)
        if not spec:
            continue
        formula = '"Oui,Non"' if spec == "oui_non" else spec
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Choisissez une valeur dans la liste proposée."
        dv.errorTitle = "Valeur non autorisée"
        ws.add_data_validation(dv)
        letter = get_column_letter(c)
        dv.add(f"{letter}{_FIRST_DATA_ROW}:{letter}{last_row}")

    ws.freeze_panes = f"A{_FIRST_DATA_ROW}"
    ws.auto_filter.ref = (
        f"A{_HEADER_ROW}:{get_column_letter(len(columns))}{_HEADER_ROW}"
    )
    ws.sheet_view.showGridLines = False


def _write_budgets(wb, rows, actions_range, hint_row=None) -> None:
    _write_simple_sheet(
        wb,
        "Budgets",
        "Un budget par action et par année (ventilation fonctionnement / "
        "investissement).",
        _BUDGET_COLUMNS,
        rows,
        dropdowns={"action": actions_range},
        hint_row=hint_row,
    )


def _write_rh(wb, rows, poste_code_range, actions_range, hint_row=None) -> None:
    _write_simple_sheet(
        wb,
        "RH",
        "Temps de travail en jours par action, année et poste (#560).",
        _RH_COLUMNS,
        rows,
        dropdowns={
            "action": actions_range,
            "poste": poste_code_range,
            "finance": "oui_non",
        },
        hint_row=hint_row,
    )


def _plan_operations(plan):
    """Opérations rattachées aux indicateurs du plan (pour le pré-remplissage)."""
    ids = [ind.id_indicateur for ind, _ in _plan_indicateurs(plan)]
    if not ids:
        return Operation.objects.none()
    return Operation.objects.filter(id_indicateur_id__in=ids).order_by(
        "ordre", "id_operation"
    )


def _op_code(op) -> str:
    return op.code_operation or f"A{op.id_operation}"


def _extract_budgets(plan) -> list[dict]:
    rows = []
    for op in _plan_operations(plan):
        for oa in op.operation_annees.all().order_by("annee"):
            if oa.budget_fonctionnement is None and oa.budget_investissement is None:
                continue
            rows.append(
                {
                    "action": _op_code(op),
                    "annee": oa.annee if oa.annee is not None else "",
                    "budget_fonctionnement": oa.budget_fonctionnement
                    if oa.budget_fonctionnement is not None
                    else "",
                    "budget_investissement": oa.budget_investissement
                    if oa.budget_investissement is not None
                    else "",
                }
            )
    return rows


def _extract_rh(plan, poste_code_by_id) -> list[dict]:
    rows = []
    for op in _plan_operations(plan):
        for oa in op.operation_annees.all().order_by("annee"):
            for rh in oa.rh_lignes.all().order_by("id_operation_annee_rh"):
                rows.append(
                    {
                        "action": _op_code(op),
                        "annee": oa.annee if oa.annee is not None else "",
                        "poste": poste_code_by_id.get(rh.id_poste_id, ""),
                        "jours": rh.jours if rh.jours is not None else "",
                        "finance": "Oui" if rh.finance else "Non",
                    }
                )
    return rows


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------


def parse_actions_workbook(source) -> dict:
    """Lit le classeur d'actions : renvoie ``{"actions": [...], "indicateurs":
    {code: id_indicateur}}``. Lève ``ArborescenceImportError`` si illisible."""
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    elif hasattr(source, "read"):
        source = io.BytesIO(source.read())
    try:
        wb = load_workbook(source, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ArborescenceImportError(
            "Le fichier n'a pas pu être lu. Vérifiez qu'il s'agit bien d'un "
            "classeur Excel (.xlsx) issu du modèle d'import des actions."
        ) from exc

    ws_by_name = {_norm(name): wb[name] for name in wb.sheetnames}

    ref = ws_by_name.get(_norm("Indicateurs"))
    actions_ws = ws_by_name.get(_norm("Actions"))
    if ref is None or actions_ws is None:
        raise ArborescenceImportError(
            "Onglets « Indicateurs » et « Actions » attendus dans le fichier."
        )

    # Référence code → id_indicateur (colonne 1 = code, colonne 4 = id).
    ref_map: dict[str, int] = {}
    for values in ref.iter_rows(min_row=_FIRST_DATA_ROW, values_only=True):
        code = _cell_str(values[0] if len(values) > 0 else None)
        ident = _as_int(values[3] if len(values) > 3 else None)
        if code and ident is not None:
            ref_map[code] = ident

    # Actions.
    header_to_key = {_norm(header): key for key, header, *_ in _ACTION_COLUMNS}
    col_index_to_key: dict[int, str] = {}
    header_cells = next(
        actions_ws.iter_rows(
            min_row=_HEADER_ROW, max_row=_HEADER_ROW, values_only=True
        ),
        (),
    )
    for idx, cell in enumerate(header_cells):
        key = header_to_key.get(_norm(cell))
        if key:
            col_index_to_key[idx] = key
    for key, header, required, *_ in _ACTION_COLUMNS:
        if required and key not in col_index_to_key.values():
            raise ArborescenceImportError(
                f"Colonne obligatoire « {header} » absente de l'onglet « Actions »."
            )

    actions = []
    for r, values in enumerate(
        actions_ws.iter_rows(min_row=_FIRST_DATA_ROW, values_only=True),
        start=_FIRST_DATA_ROW,
    ):
        record, has_value = {}, False
        for idx, key in col_index_to_key.items():
            value = values[idx] if idx < len(values) else None
            if _cell_str(value):
                has_value = True
            record[key] = value
        if not has_value:
            continue
        if _is_example_row(record.get("code")):  # ligne exemple du modèle
            continue
        record["_row"] = r
        actions.append(record)

    # Référence postes (facultative) : code (col 1) → id_poste (col 4).
    poste_map: dict[str, int] = {}
    postes_ws = ws_by_name.get(_norm("Postes"))
    if postes_ws is not None:
        for values in postes_ws.iter_rows(min_row=_FIRST_DATA_ROW, values_only=True):
            code = _cell_str(values[0] if len(values) > 0 else None)
            ident = _as_int(values[3] if len(values) > 3 else None)
            if code and ident is not None:
                poste_map[code] = ident

    budgets = _parse_flat_sheet(ws_by_name.get(_norm("Budgets")), _BUDGET_HEADERS)
    rh = _parse_flat_sheet(ws_by_name.get(_norm("RH")), _RH_HEADERS)

    return {
        "actions": actions,
        "indicateurs": ref_map,
        "postes": poste_map,
        "budgets": budgets,
        "rh": rh,
    }


def _parse_flat_sheet(ws, headers_map) -> list[dict]:
    """Lit un onglet à plat (en-tête ligne 2, données dès la ligne 3) en
    rapprochant les en-têtes de ``headers_map`` (``{en-tête normalisé: clé}``)."""
    if ws is None:
        return []
    col_index_to_key: dict[int, str] = {}
    header_cells = next(
        ws.iter_rows(min_row=_HEADER_ROW, max_row=_HEADER_ROW, values_only=True), ()
    )
    for idx, cell in enumerate(header_cells):
        key = headers_map.get(_norm(cell))
        if key:
            col_index_to_key[idx] = key

    rows = []
    for r, values in enumerate(
        ws.iter_rows(min_row=_FIRST_DATA_ROW, values_only=True), start=_FIRST_DATA_ROW
    ):
        record, has_value = {}, False
        for idx, key in col_index_to_key.items():
            value = values[idx] if idx < len(values) else None
            if _cell_str(value):
                has_value = True
            record[key] = value
        if not has_value:
            continue
        if _is_example_row(record.get("action")):  # ligne exemple du modèle
            continue
        record["_row"] = r
        rows.append(record)
    return rows


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------


def validate_actions_import(plan, parsed: dict) -> ImportReport:
    report = ImportReport()
    resolver = _NomenclatureResolver()

    plan_indicateur_ids = {ind.id_indicateur for ind, _ in _plan_indicateurs(plan)}
    if not plan_indicateur_ids:
        report.add(
            None,
            None,
            None,
            ERROR,
            "Ce plan n'a pas encore d'indicateurs : importez d'abord "
            "l'arborescence.",
        )

    # Création seule : refuser si des actions existent déjà.
    if Operation.objects.filter(id_indicateur_id__in=plan_indicateur_ids).exists():
        report.add(
            None,
            None,
            None,
            ERROR,
            "Ce plan contient déjà des actions. L'import n'est possible "
            "que sur un plan sans action.",
        )

    ref_map = parsed.get("indicateurs", {})
    codes: set[str] = set()
    for row in parsed.get("actions", []):
        r = row["_row"]
        code = _cell_str(row.get("code"))
        if not code:
            report.add("Actions", r, "code", ERROR, "Code manquant.")
        elif code in codes:
            report.add("Actions", r, "code", ERROR, f"Code « {code} » en double.")
        else:
            codes.add(code)

        if not _cell_str(row.get("libelle")):
            report.add("Actions", r, "libelle", ERROR, "« libellé » est obligatoire.")

        ind_code = _cell_str(row.get("indicateur"))
        if not ind_code:
            report.add(
                "Actions",
                r,
                "indicateur",
                ERROR,
                "L'indicateur de rattachement est obligatoire.",
            )
        else:
            ident = ref_map.get(ind_code)
            if ident is None:
                report.add(
                    "Actions",
                    r,
                    "indicateur",
                    ERROR,
                    f"Indicateur « {ind_code} » introuvable dans l'onglet "
                    "« Indicateurs ».",
                )
            elif ident not in plan_indicateur_ids:
                report.add(
                    "Actions",
                    r,
                    "indicateur",
                    ERROR,
                    f"L'indicateur « {ind_code} » n'appartient pas à ce plan.",
                )

        for col, type_mnemo, label in (
            ("type_action", _TYPE_ACTION, "type d'action"),
            ("priorite", _PRIORITE, "priorité"),
        ):
            value = _cell_str(row.get(col))
            if value and resolver.resolve(type_mnemo, value) is None:
                report.add(
                    "Actions",
                    r,
                    col,
                    ERROR,
                    f"Valeur « {value} » non reconnue pour « {label} ».",
                )

        amin, amax = _as_int(row.get("annee_min")), _as_int(row.get("annee_max"))
        for col, raw in (
            ("annee_min", row.get("annee_min")),
            ("annee_max", row.get("annee_max")),
        ):
            if _cell_str(raw) and _as_int(raw) is None:
                report.add("Actions", r, col, ERROR, "L'année doit être un nombre.")
        if amin is not None and amax is not None and amin > amax:
            report.add(
                "Actions",
                r,
                "annee_max",
                ERROR,
                "L'année de fin est antérieure à l'année de début.",
            )

    # --- Budgets et RH (facultatifs), rattachés aux actions par leur code ---
    action_codes = {
        _cell_str(r.get("code"))
        for r in parsed.get("actions", [])
        if _cell_str(r.get("code"))
    }
    spans = {
        _cell_str(r.get("code")): (
            _as_int(r.get("annee_min")),
            _as_int(r.get("annee_max")),
        )
        for r in parsed.get("actions", [])
        if _cell_str(r.get("code"))
    }
    poste_ref = parsed.get("postes", {})
    plan_poste_ids = {p.id_poste for p in _plan_postes(plan)}

    def _check_action_annee(sheet, row, ac):
        rr = row["_row"]
        if not ac:
            report.add(sheet, rr, "action", ERROR, "Code d'action manquant.")
        elif ac not in action_codes:
            report.add(
                sheet,
                rr,
                "action",
                ERROR,
                f"Action « {ac} » introuvable dans l'onglet « Actions ».",
            )
        annee = _as_int(row.get("annee"))
        if not _cell_str(row.get("annee")):
            report.add(sheet, rr, "annee", ERROR, "L'année est obligatoire.")
        elif annee is None:
            report.add(sheet, rr, "annee", ERROR, "L'année doit être un nombre.")
        elif ac in spans:
            amin, amax = spans[ac]
            if amin is not None and amax is not None and not (amin <= annee <= amax):
                report.add(
                    sheet,
                    rr,
                    "annee",
                    WARNING,
                    f"L'année {annee} est hors de la période de l'action "
                    f"({amin}–{amax}).",
                )
        return annee

    def _check_amount(sheet, row, col, label):
        raw = row.get(col)
        if not _cell_str(raw):
            return
        value = _as_decimal(raw)
        if value is None:
            report.add(sheet, row["_row"], col, ERROR, f"{label} doit être un nombre.")
        elif value < 0:
            report.add(
                sheet, row["_row"], col, ERROR, f"{label} ne peut pas être négatif."
            )

    for row in parsed.get("budgets", []):
        _check_action_annee("Budgets", row, _cell_str(row.get("action")))
        _check_amount(
            "Budgets", row, "budget_fonctionnement", "Le budget de fonctionnement"
        )
        _check_amount(
            "Budgets", row, "budget_investissement", "Le budget d'investissement"
        )

    for row in parsed.get("rh", []):
        _check_action_annee("RH", row, _cell_str(row.get("action")))
        poste_code = _cell_str(row.get("poste"))
        if not poste_code:
            report.add("RH", row["_row"], "poste", ERROR, "Le poste est obligatoire.")
        else:
            pid = poste_ref.get(poste_code)
            if pid is None:
                report.add(
                    "RH",
                    row["_row"],
                    "poste",
                    ERROR,
                    f"Poste « {poste_code} » introuvable dans l'onglet « Postes ».",
                )
            elif pid not in plan_poste_ids:
                report.add(
                    "RH",
                    row["_row"],
                    "poste",
                    ERROR,
                    f"Le poste « {poste_code} » n'appartient pas à ce plan.",
                )
        _check_amount("RH", row, "jours", "Le nombre de jours")
        finance = _cell_str(row.get("finance"))
        if finance and _parse_bool(finance) is None:
            report.add(
                "RH", row["_row"], "finance", ERROR, "Valeur attendue : Oui ou Non."
            )

    report.summary = {
        "actions": len(parsed.get("actions", [])),
        "budgets": len(parsed.get("budgets", [])),
        "rh": len(parsed.get("rh", [])),
    }
    return report


# ---------------------------------------------------------------------------
# Exécution
# ---------------------------------------------------------------------------


@transaction.atomic
def execute_actions_import(plan, parsed: dict, user) -> dict:
    report = validate_actions_import(plan, parsed)
    if not report.can_import:
        raise ValueError(report)

    resolver = _NomenclatureResolver()
    ref_map = parsed.get("indicateurs", {})
    indicateurs = {ind.id_indicateur: ind for ind, _ in _plan_indicateurs(plan)}

    def nom(type_mnemo, value):
        v = _cell_str(value)
        return resolver.resolve(type_mnemo, v) if v else None

    op_by_code: dict[str, Operation] = {}
    annee_index: dict[tuple[str, int], OperationAnnee] = {}

    n_actions = n_annees = 0
    for i, row in enumerate(parsed.get("actions", [])):
        ident = ref_map.get(_cell_str(row.get("indicateur")))
        indicateur = indicateurs.get(ident)
        amin = _as_int(row.get("annee_min"))
        amax = _as_int(row.get("annee_max"))
        code = _cell_str(row.get("code"))
        operation = Operation.objects.create(
            id_indicateur=indicateur,
            libelle=_cell_str(row.get("libelle")),
            code_operation=code or None,
            id_type_action=nom(_TYPE_ACTION, row.get("type_action")),
            id_priorite=nom(_PRIORITE, row.get("priorite")),
            annee_min=amin,
            annee_max=amax,
            operateurs=_cell_str(row.get("operateurs")) or None,
            financeurs=_cell_str(row.get("financeurs")) or None,
            description=_cell_str(row.get("description")) or None,
            statut="draft",
            ordre=i,
            id_utilisateur_ajout=user,
        )
        op_by_code[code] = operation
        n_actions += 1
        if amin is not None and amax is not None:
            for annee in range(amin, amax + 1):
                oa = OperationAnnee.objects.create(id_operation=operation, annee=annee)
                annee_index[(code, annee)] = oa
                n_annees += 1

    def _get_annee(code, annee):
        """OperationAnnee de (action, année), créée à la volée si nécessaire."""
        nonlocal n_annees
        key = (code, annee)
        oa = annee_index.get(key)
        if oa is None:
            oa = OperationAnnee.objects.create(
                id_operation=op_by_code[code], annee=annee
            )
            annee_index[key] = oa
            n_annees += 1
        return oa

    # --- Budgets (ventilation par type au niveau de l'OperationAnnee) ---
    ops_with_budget: set[str] = set()
    n_budgets = 0
    for row in parsed.get("budgets", []):
        code = _cell_str(row.get("action"))
        annee = _as_int(row.get("annee"))
        if code not in op_by_code or annee is None:
            continue
        oa = _get_annee(code, annee)
        bf = _as_decimal(row.get("budget_fonctionnement"))
        bi = _as_decimal(row.get("budget_investissement"))
        oa.budget_fonctionnement = bf
        oa.budget_investissement = bi
        oa.save(update_fields=["budget_fonctionnement", "budget_investissement"])
        ops_with_budget.add(code)
        n_budgets += 1

    # --- RH (temps de travail par poste, #560) ---
    postes_by_id = {p.id_poste: p for p in _plan_postes(plan)}
    poste_ref = parsed.get("postes", {})
    ops_with_rh: set[str] = set()
    n_rh = 0
    for row in parsed.get("rh", []):
        code = _cell_str(row.get("action"))
        annee = _as_int(row.get("annee"))
        if code not in op_by_code or annee is None:
            continue
        oa = _get_annee(code, annee)
        poste = postes_by_id.get(poste_ref.get(_cell_str(row.get("poste"))))
        finance = _parse_bool(row.get("finance"))
        OperationAnneeRH.objects.create(
            id_operation_annee=oa,
            id_poste=poste,
            id_organisme=poste.id_organisme if poste else None,
            jours=_as_decimal(row.get("jours")),
            finance=finance if finance is not None else True,
        )
        ops_with_rh.add(code)
        n_rh += 1

    # Marqueurs de ventilation sur les opérations concernées.
    for code in ops_with_budget:
        op = op_by_code[code]
        op.ventilation_mode = "by_type"
        op.save(update_fields=["ventilation_mode"])
    for code in ops_with_rh:
        op = op_by_code[code]
        op.declinaison_par_poste = True
        op.save(update_fields=["declinaison_par_poste"])

    return {
        "actions": n_actions,
        "annees": n_annees,
        "budgets": n_budgets,
        "rh": n_rh,
    }


# ---------------------------------------------------------------------------
# Couture JSON (sans fichier) — parité avec l'arborescence (grille #9, IA)
# ---------------------------------------------------------------------------
#
# Ces fonctions permettent d'alimenter le pipeline validate/execute des actions
# à partir de données JSON plutôt que d'un classeur : grille de correction
# interactive (#9) et surtout **extraction IA** (PDF → JSON au format ci-dessous
# → relecture → import). Le contrat est identique à celui de l'arborescence
# (``describe_schema`` / ``public_parsed`` / ``sanitize_parsed`` dans
# ``services_import``), transposé aux onglets Actions / Budgets / RH.

# Clés de colonne autorisées par onglet de saisie (les onglets « Indicateurs » et
# « Postes » sont des références en lecture seule, exposées séparément).
_ACTIONS_SHEET_KEYS = {
    "actions": [c[0] for c in _ACTION_COLUMNS],
    "budgets": [c[0] for c in _BUDGET_COLUMNS],
    "rh": [c[0] for c in _RH_COLUMNS],
}


def _actions_reference(plan):
    """Codes de rattachement (identiques à ceux des onglets de référence du
    classeur) et maps ``{code: id}`` pour indicateurs et postes du plan."""
    indicateurs = _plan_indicateurs(plan)
    ind_code_by_id = {
        ind.id_indicateur: f"I{i}" for i, (ind, _) in enumerate(indicateurs, start=1)
    }
    postes = _plan_postes(plan)
    poste_code_by_id = {p.id_poste: f"Q{i}" for i, p in enumerate(postes, start=1)}
    return indicateurs, ind_code_by_id, postes, poste_code_by_id


def _describe_action_columns() -> list[dict]:
    refs = {"indicateur": "indicateurs"}
    return [
        {
            "key": key,
            "header": header,
            "required": bool(required),
            "nomenclature": nomenclature,
            "ref": refs.get(key),
            "help": help_,
        }
        for key, header, required, help_, nomenclature, _width in _ACTION_COLUMNS
    ]


def _describe_flat_columns(columns, refs) -> list[dict]:
    return [
        {
            "key": key,
            "header": header,
            "required": bool(required),
            "nomenclature": None,
            "ref": refs.get(key),
            "help": help_,
        }
        for key, header, required, help_, _width in columns
    ]


def describe_actions_schema(plan) -> dict:
    """Décrit les onglets/colonnes du format actions et les listes de référence
    du plan (indicateurs, postes).

    Sert à piloter la grille de correction (#9) et à fournir à l'extracteur IA la
    liste des **codes valides** (indicateurs/postes) pour le rattachement.
    """
    indicateurs, ind_code_by_id, postes, poste_code_by_id = _actions_reference(plan)
    ref_indicateurs = [
        {
            "code": ind_code_by_id[ind.id_indicateur],
            "indicateur": ind.nom_indicateur,
            "enjeu": enjeu_libelle,
            "id": ind.id_indicateur,
        }
        for ind, enjeu_libelle in indicateurs
    ]
    ref_postes = [
        {
            "code": poste_code_by_id[p.id_poste],
            "poste": p.libelle,
            "organisme": str(p.id_organisme) if p.id_organisme else "",
            "id": p.id_poste,
        }
        for p in postes
    ]
    sheets = [
        {"key": "actions", "name": "Actions", "columns": _describe_action_columns()},
        {
            "key": "budgets",
            "name": "Budgets",
            "columns": _describe_flat_columns(_BUDGET_COLUMNS, {"action": "actions"}),
        },
        {
            "key": "rh",
            "name": "RH",
            "columns": _describe_flat_columns(
                _RH_COLUMNS, {"action": "actions", "poste": "postes"}
            ),
        },
    ]
    return {
        "sheets": sheets,
        "references": {"indicateurs": ref_indicateurs, "postes": ref_postes},
    }


def sanitize_actions_parsed(plan, data: dict) -> dict:
    """Convertit des données JSON (grille #9 / extraction IA) → format ``parsed``
    interne attendu par ``validate_actions_import`` / ``execute_actions_import``.

    Ne conserve que les colonnes connues + ``_row`` ; ignore les lignes vides et
    les clés internes. Les maps de référence ``indicateurs`` / ``postes``
    (``{code: id}``) sont **reconstruites depuis le plan** (autorité serveur), et
    non lues depuis le client.
    """
    data = data if isinstance(data, dict) else {}
    _indics, ind_code_by_id, _postes, poste_code_by_id = _actions_reference(plan)

    out: dict = {}
    for sheet_key, keys in _ACTIONS_SHEET_KEYS.items():
        rows = data.get(sheet_key)
        clean = []
        for i, r in enumerate(rows or []):
            if not isinstance(r, dict):
                continue
            row = {k: r.get(k) for k in keys}
            if not any(_cell_str(v) for v in row.values()):
                continue  # ligne entièrement vide
            row["_row"] = _as_int(r.get("_row")) or (i + _FIRST_DATA_ROW)
            clean.append(row)
        out[sheet_key] = clean

    out["indicateurs"] = {code: ident for ident, code in ind_code_by_id.items()}
    out["postes"] = {code: ident for ident, code in poste_code_by_id.items()}
    return out


def public_actions_parsed(parsed: dict) -> dict:
    """Copie « publique » des lignes parsées (colonnes + ``_row``) pour renvoyer
    les données à éditer dans la grille (#9).

    Ne renvoie que les onglets de saisie (comme ``public_parsed`` pour l'arbo) :
    les maps de référence indicateurs/postes sont reconstruites côté serveur par
    ``sanitize_actions_parsed`` à chaque appel, inutile de les exposer.
    """
    out: dict = {}
    for sheet_key, keys in _ACTIONS_SHEET_KEYS.items():
        rows = parsed.get(sheet_key) or []
        clean = []
        for r in rows:
            row = {k: r.get(k) for k in keys}
            row["_row"] = r.get("_row")
            clean.append(row)
        out[sheet_key] = clean
    return out
