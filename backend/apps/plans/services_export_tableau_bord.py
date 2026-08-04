"""
Export Excel du **tableau de bord** des indicateurs (#638).

Le tableau de bord est filtré côté client (onglet État / Pression / Ensemble,
nom d'objectif, enjeu, recherche) et ses scores sont calculés à l'affichage.
C'est donc le client qui envoie **les lignes telles qu'il les montre**, et ce
module se charge de la seule chose qu'un CSV ne sait pas faire : la mise en
forme. L'export reste ainsi exactement ce que l'utilisateur a sous les yeux —
la promesse d'origine de #638 — sans réimplémenter le filtrage et le calcul de
score une seconde fois côté serveur, où ils divergeraient au premier écart.

Mise en forme (retour de recette : « dans le même esprit que l'interface ») :
en-têtes à la couleur de l'instance (#601) et **cases de score reprenant la
palette de l'application**. Ces couleurs de score ne sont volontairement pas
paramétrables — elles forment la grille de lecture, cf. :mod:`export_theme`.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import export_theme

# --- palette ---------------------------------------------------------------

#: Couleurs de score de l'application (`SCORE_PALETTE`, design system).
#: Texte toujours noir : les fonds sont clairs, cf. docs/DESIGN_SYSTEM.md.
SCORE_FILLS = {
    'very-bad': 'FFFF7579',
    'bad': 'FFFA9965',
    'neutral': 'FFF7D35C',
    'good': 'FF82DB8A',
    'very-good': 'FF81C9D8',
    'no-data': 'FFDADADA',
}

_BLACK = 'FF343433'
_WHITE = 'FFFFFFFF'
_GRAY = 'FF746F6E'
_METRIQUE_FILL = 'FFF8F5F1'   # beige du design system, pour les lignes de détail
_BORDER_COLOR = 'FFE4E4E4'

_PRIMARY = 'FF025359'
_F_TITLE = Font(name='Calibri', bold=True, size=14, color=_PRIMARY)
_F_META = Font(name='Calibri', size=9, color=_GRAY)
_F_HDR = Font(name='Calibri', bold=True, size=9, color=_WHITE)
_F_INDIC = Font(name='Calibri', bold=True, size=9, color=_BLACK)
_F_METRIQUE = Font(name='Calibri', size=9, color=_BLACK)
_F_SCORE = Font(name='Calibri', size=9, color=_BLACK)

_THIN = Side(style='thin', color=_BORDER_COLOR)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def _appliquer_couleur_instance():
    """Réaligne les styles sur la couleur d'export de l'instance (#601)."""
    global _PRIMARY, _F_TITLE

    couleur = export_theme.argb()
    if couleur == _PRIMARY:
        return
    _PRIMARY = couleur
    _F_TITLE = Font(name='Calibri', bold=True, size=14, color=couleur)


# --- lecture de la charge utile --------------------------------------------

def _texte(cellule) -> str:
    """Texte d'une cellule, qu'elle soit nue ou porteuse d'un score."""
    if isinstance(cellule, dict):
        return str(cellule.get('t') or '')
    return '' if cellule is None else str(cellule)


def _score(cellule) -> str | None:
    """Niveau de score d'une cellule, s'il y en a un et qu'il est connu."""
    if not isinstance(cellule, dict):
        return None
    niveau = cellule.get('s')
    return niveau if niveau in SCORE_FILLS else None


# --- construction ----------------------------------------------------------

def build_tableau_bord_workbook(payload: dict) -> bytes:
    """
    Rend le tableau de bord affiché en classeur Excel mis en forme.

    ``payload`` (envoyé par le client) :

    - ``titre``    : titre du classeur ;
    - ``meta``     : lignes « libellé / valeur » rappelant les filtres actifs ;
    - ``entetes``  : libellés de colonnes ;
    - ``lignes``   : ``[{'type': 'indicateur'|'metrique', 'cellules': [...]}]``,
      chaque cellule étant une chaîne ou ``{'t': texte, 's': niveau de score}``.
    """
    _appliquer_couleur_instance()

    entetes = [str(h) for h in (payload.get('entetes') or [])]
    lignes = payload.get('lignes') or []
    meta = payload.get('meta') or []

    wb = Workbook()
    ws = wb.active
    ws.title = 'Tableau de bord'

    ligne = 1
    titre = payload.get('titre')
    if titre:
        cell = ws.cell(row=ligne, column=1, value=str(titre))
        cell.font = _F_TITLE
        ligne += 1

    for libelle, valeur in ((m + ['', ''])[:2] for m in meta):
        ws.cell(row=ligne, column=1, value=str(libelle)).font = _F_META
        ws.cell(row=ligne, column=2, value=str(valeur)).font = _F_META
        ligne += 1

    if titre or meta:
        ligne += 1  # respiration avant le tableau

    # En-têtes : aplat à la couleur de l'instance, texte blanc.
    entete_ligne = ligne
    for col, libelle in enumerate(entetes, start=1):
        cell = ws.cell(row=entete_ligne, column=col, value=libelle)
        cell.font = _F_HDR
        cell.fill = PatternFill('solid', fgColor=_PRIMARY)
        cell.alignment = _CENTER
        cell.border = _BORDER
    ligne += 1

    for row in lignes:
        est_metrique = row.get('type') == 'metrique'
        cellules = row.get('cellules') or []
        for col, brute in enumerate(cellules, start=1):
            cell = ws.cell(row=ligne, column=col, value=_texte(brute))
            niveau = _score(brute)
            if niveau:
                # Case colorée comme la pastille de score de l'interface.
                cell.fill = PatternFill('solid', fgColor=SCORE_FILLS[niveau])
                cell.font = _F_SCORE
                cell.alignment = _CENTER
            else:
                cell.font = _F_METRIQUE if est_metrique else _F_INDIC
                cell.alignment = _LEFT
                if est_metrique:
                    cell.fill = PatternFill('solid', fgColor=_METRIQUE_FILL)
            cell.border = _BORDER
        ligne += 1

    _mettre_en_page(ws, entetes, lignes, entete_ligne)

    flux = io.BytesIO()
    wb.save(flux)
    return flux.getvalue()


def _mettre_en_page(ws, entetes, lignes, entete_ligne):
    """Largeurs de colonnes, hauteur d'en-tête et volets figés."""
    if not entetes:
        return

    # Largeur : au plus large des libellés de la colonne, dans des bornes qui
    # gardent les colonnes d'années compactes et les libellés lisibles.
    for col, libelle in enumerate(entetes, start=1):
        largeur = len(libelle)
        for row in lignes:
            cellules = row.get('cellules') or []
            if col <= len(cellules):
                largeur = max(largeur, len(_texte(cellules[col - 1])))
        ws.column_dimensions[get_column_letter(col)].width = min(max(largeur + 2, 8), 42)

    ws.row_dimensions[entete_ligne].height = 30
    # Fige l'en-tête ET les colonnes d'identification (jusqu'à « Métrique ») :
    # sans cela, faire défiler les années fait perdre de vue l'indicateur lu.
    ws.freeze_panes = ws.cell(row=entete_ligne + 1, column=min(6, len(entetes) + 1))
    ws.auto_filter.ref = (
        f'A{entete_ligne}:{get_column_letter(len(entetes))}{entete_ligne + len(lignes)}'
    )
